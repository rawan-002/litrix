"""Dashboard affiliation-policy integration check — a release-gate guard.

Read-only integration test for the tri-state affiliation policy
(analytics/stats.py :: verified_affil_clause / active_affil_clause). It calls the
REAL DRF endpoints (force-authenticated as an admin) against the configured DB
and asserts the CONTRACT across SQL <-> API <-> frontend, so a future edit that
points a consumer at the wrong filter breaks CI here instead of silently
inflating an official number after release.

Six affiliation-policy checks, plus a data-quality health check (7):
  1. Overview     — endpoint returns the official metrics (papers/citations/Q1-Q4).
  2. Drill-down   — classified-papers row count == the Publications KPI.
  3. Excel        — each summary sheet's values match the endpoint that feeds it
                    (compare SHEET CONTENT to the API — NOT the file's total rows,
                    the workbook is a multi-sheet summary, not a paper list).
  4. Top Papers   — under affiliation=albaha, no FALSE (confirmed-elsewhere) row,
                    and every row carries `affiliation_verified` so a NULL can be
                    badged 'Pending' by the frontend.
  5. Filter       — albaha <= all for every official metric.
  6. Triangulation— a direct SQL COUNT (AffiliationVerified = TRUE, same scope)
                    == the KPI == the drill-down count. Proves SQL, API, and the
                    frontend contract all use the SAME policy.
  7. Data quality — regression checks for the two classes of bug this project
                    has actually hit in production: duplicate identifiers that
                    should be unique (DOI, normalized ISSN, researcher Scopus_ID)
                    and orphaned rows a broken ingestion step can leave behind
                    (papers with zero authors, journal rankings pointing at a
                    dead JournalID). The uq_paper_doi_normalized / on ResearchPaper
                    (DOI) and uq_journals_normalized_issn_print / on Journals
                    (ISSN_Print) already enforce the first two at the DB level
                    (migrations/20260804_*.sql) — these checks make a violation
                    visible in CI output instead of only surfacing as an insert
                    failure deep in a pipeline run.

Run (from backend/, uses the same .env / DATABASE_URL as everything else):
    python tools/integration_check.py
Exit 0 = all checks hold; 1 = a violation (use it as a release gate / in CI).
"""
import io
import os
import sys

# Django bootstrap (this file lives in backend/tools/).
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'litrix_backend.settings')
import django  # noqa: E402
django.setup()

from django.contrib.auth import get_user_model              # noqa: E402
from django.db import connection                             # noqa: E402
from rest_framework.test import APIRequestFactory, force_authenticate  # noqa: E402

from analytics import views, exports                          # noqa: E402
from analytics.stats import FOCUS_YEARS                       # noqa: E402

User = get_user_model()
rf = APIRequestFactory()
YEARS = list(FOCUS_YEARS)
YEARS_CSV = ','.join(str(y) for y in YEARS)

results = []  # (ok, label, detail)


def check(ok, label, detail=''):
    results.append((bool(ok), label, detail))


def _admin():
    for u in User.objects.all():
        try:
            if u.has_litrix_perm('view_all_researchers'):
                return u
        except Exception:
            continue
    return User.objects.filter(is_superuser=True).first()


def _call(view, path, params, user):
    req = rf.get(path, params)
    force_authenticate(req, user=user)
    return view(req)


def _to_int(v):
    """Parse a workbook cell value like '1,234' -> 1234."""
    if v is None:
        return None
    return int(str(v).replace(',', '').strip())


def main():
    admin = _admin()
    if admin is None:
        print('FAIL: no admin/superuser with view_all_researchers to test with.')
        sys.exit(1)

    albaha = {'affiliation': 'albaha', 'years': YEARS_CSV}
    all_   = {'affiliation': 'all',    'years': YEARS_CSV}

    # ---- 1. Overview endpoint (the official metrics) ----
    ov = _call(views.overview, '/api/overview/', albaha, admin)
    ok = ov.status_code == 200
    t = ov.data['totals'] if ok else {}
    metrics = ('papers', 'citations', 'q1_papers', 'q2_papers', 'q3_papers', 'q4_papers')
    fields_ok = ok and all(isinstance(t.get(m), int) and t[m] >= 0 for m in metrics) \
        and isinstance(t.get('pending_review'), int)
    check(fields_ok, 'Overview returns official metrics + pending_review',
          f'papers={t.get("papers")} cites={t.get("citations")} '
          f'Q1={t.get("q1_papers")} pending={t.get("pending_review")}')
    kpi_papers = t.get('papers')

    # ---- 2. Drill-down count == Publications KPI ----
    dd = _call(views.classified_papers, '/api/stats/classified-papers/', albaha, admin)
    dd_count = dd.data.get('count', len(dd.data.get('papers', []))) if dd.status_code == 200 else None
    check(dd_count == kpi_papers, 'Drill-down count == Publications KPI',
          f'KPI={kpi_papers} drill-down={dd_count}')

    # ---- 3. Excel: sheet content matches the endpoint that feeds it ----
    # The Overview sheet mirrors the dashboard cards (see exports.py): row 5 holds
    # the values at cols 1/3/5/7 = Researchers/Publications/Citations/h-index.
    # We compare SHEET VALUES to the API — never the file's total row count.
    xreq = rf.get('/api/export/excel/',
                  {'affiliation': 'albaha', 'years': YEARS_CSV, 'sheets': 'summary'})
    force_authenticate(xreq, user=admin)
    xr = exports.export_excel(xreq)
    content = xr.content if hasattr(xr, 'content') else b''.join(xr.streaming_content)
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb['Overview']
    xls_pubs  = _to_int(ws.cell(row=5, column=3).value)
    xls_cites = _to_int(ws.cell(row=5, column=5).value)
    check(xls_pubs == kpi_papers, 'Excel Overview sheet: Publications == API',
          f'excel={xls_pubs} api={kpi_papers}')
    check(xls_cites == t.get('citations'), 'Excel Overview sheet: Citations == API',
          f'excel={xls_cites} api={t.get("citations")}')

    # ---- 4. Top Papers: no FALSE under albaha; badge field present ----
    tp = ov.data['top_papers']
    check(all('affiliation_verified' in p for p in tp),
          'Top Papers expose affiliation_verified (badge field)',
          f'{sum(1 for p in tp if "affiliation_verified" in p)}/{len(tp)}')
    check(all(p.get('affiliation_verified') is not False for p in tp),
          'Top Papers has NO FALSE (confirmed-elsewhere) row under albaha',
          f'false={sum(1 for p in tp if p.get("affiliation_verified") is False)}')

    # ---- 5. Filter consistency: albaha <= all for every official metric ----
    ova = _call(views.overview, '/api/overview/', all_, admin)
    ta = ova.data['totals'] if ova.status_code == 200 else {}
    for m in metrics:
        check(t.get(m, 0) <= ta.get(m, 0), f'albaha.{m} <= all.{m}',
              f'albaha={t.get(m)} all={ta.get(m)}')

    # ---- 6. Triangulation: raw SQL COUNT(TRUE) == KPI == drill-down ----
    # Same scope as the overview KPI: papers in the window, attributed to at least
    # one author, with a CONFIRMED Al-Baha affiliation.
    with connection.cursor() as cur:
        cur.execute(
            'SELECT COUNT(DISTINCT rp."PaperID") FROM "ResearchPaper" rp '
            'WHERE rp."PubYear" = ANY(%s) '
            '  AND rp."AffiliationVerified" = TRUE '
            '  AND EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID")',
            [YEARS])
        sql_count = cur.fetchone()[0]
    check(sql_count == kpi_papers == dd_count,
          'SQL COUNT(TRUE) == KPI == drill-down (one policy end-to-end)',
          f'sql={sql_count} kpi={kpi_papers} drill={dd_count}')

    # ---- 7. Data quality health check ----
    with connection.cursor() as cur:
        cur.execute('''
            SELECT COUNT(*) FROM (
                SELECT lower(trim("DOI")) FROM "ResearchPaper"
                WHERE "DOI" IS NOT NULL AND trim("DOI") != ''
                GROUP BY 1 HAVING COUNT(*) > 1
            ) x
        ''')
        dup_doi = cur.fetchone()[0]
        check(dup_doi == 0, 'No duplicate DOI (case/whitespace-normalized)',
              f'groups={dup_doi}')

        cur.execute('''
            SELECT COUNT(*) FROM (
                SELECT regexp_replace(regexp_replace(upper(COALESCE("ISSN_Print", '')),
                                                       '^ISSN-?', ''), '[^A-Z0-9]', '', 'g') AS n
                FROM "Journals"
                WHERE "ISSN_Print" IS NOT NULL AND "ISSN_Print" != ''
                GROUP BY 1
                HAVING length(regexp_replace(regexp_replace(upper(COALESCE("ISSN_Print", '')),
                                                              '^ISSN-?', ''), '[^A-Z0-9]', '', 'g')) = 8
                   AND COUNT(*) > 1
            ) x
        ''')
        dup_issn = cur.fetchone()[0]
        check(dup_issn == 0, 'No duplicate Journal ISSN (format-normalized)',
              f'groups={dup_issn}')

        cur.execute('''
            SELECT COUNT(*) FROM (
                SELECT "Scopus_ID" FROM "Researcher"
                WHERE "Scopus_ID" IS NOT NULL AND trim("Scopus_ID") != ''
                GROUP BY 1 HAVING COUNT(*) > 1
            ) x
        ''')
        dup_scopus = cur.fetchone()[0]
        check(dup_scopus == 0, 'No duplicate researcher Scopus_ID',
              f'groups={dup_scopus}')

        cur.execute('''
            SELECT COUNT(*) FROM "ResearchPaper" p
            WHERE NOT EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = p."PaperID")
        ''')
        orphan_papers = cur.fetchone()[0]
        check(orphan_papers == 0, 'No orphan ResearchPaper (zero linked Authors)',
              f'orphans={orphan_papers}')

        cur.execute('''
            SELECT COUNT(*) FROM "JournalRankings" jr
            WHERE jr."JournalID" IS NOT NULL
              AND NOT EXISTS (SELECT 1 FROM "Journals" j WHERE j."JournalID" = jr."JournalID")
        ''')
        orphan_rankings = cur.fetchone()[0]
        check(orphan_rankings == 0, 'No orphan JournalRankings (dead JournalID)',
              f'orphans={orphan_rankings}')

    # ---- Report ----
    print('=' * 70)
    print(f' Litrix integration check - affiliation policy | scope {YEARS_CSV}')
    print('=' * 70)
    failed = 0
    for ok, label, detail in results:
        if not ok:
            failed += 1
        print(f'  [{"PASS" if ok else "FAIL"}] {label}' + (f'   ({detail})' if detail else ''))
    print('-' * 70)
    print(f'  {len(results) - failed}/{len(results)} checks hold.')
    sys.exit(1 if failed else 0)


if __name__ == '__main__':
    main()
