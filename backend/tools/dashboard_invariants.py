"""Dashboard affiliation-policy invariants — a release-gate regression guard.

Read-only integration test for the tri-state affiliation policy
(see analytics/stats.py :: verified_affil_clause / active_affil_clause).
It calls the REAL DRF endpoints (force-authenticated as an admin) against the
configured DB and asserts the CONTRACT between backend and frontend, so a future
edit that points a consumer at the wrong filter fails here instead of silently
inflating an official number.

Invariants checked:
  1. Overview KPI (papers) == classified-papers drill-down count      (albaha ON)
  2. Excel "Publications" value == the same KPI                        (albaha ON)
  3. Top Papers never contains a FALSE (confirmed-elsewhere) paper     (albaha ON)
     and every row carries `affiliation_verified` (so NULL can be badged)
  4. Monotonicity: albaha KPI <= all KPI for papers/citations/Q1-Q4

Run (from backend/, reads the same .env / DATABASE_URL as everything else):
    python tools/dashboard_invariants.py
Exit code 0 = all invariants hold; 1 = a violation (use it in a release gate/CI).
"""
import io
import os
import sys

# Django bootstrap (this file lives in backend/tools/).
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'litrix_backend.settings')
import django  # noqa: E402
django.setup()

from django.contrib.auth import get_user_model              # noqa: E402
from rest_framework.test import APIRequestFactory, force_authenticate  # noqa: E402

from analytics import views, exports                          # noqa: E402
from analytics.stats import FOCUS_YEARS                       # noqa: E402

User = get_user_model()
rf = APIRequestFactory()

# Shared scope so every endpoint compares on the SAME years as the dashboard.
YEARS_CSV = ','.join(str(y) for y in FOCUS_YEARS)

results = []  # (ok: bool, label: str, detail: str)


def _admin():
    for u in User.objects.all():
        try:
            if u.has_litrix_perm('view_all_researchers'):
                return u
        except Exception:
            continue
    return User.objects.filter(is_superuser=True).first()


def _call(view, path, params, user):
    req = rf.get(path, params)
    force_authenticate(req, user=user)
    return view(req)


def check(ok, label, detail=''):
    results.append((bool(ok), label, detail))


def main():
    admin = _admin()
    if admin is None:
        print('FAIL: no admin/superuser with view_all_researchers to test with.')
        sys.exit(1)

    base = {'affiliation': 'albaha', 'years': YEARS_CSV}
    all_ = {'affiliation': 'all', 'years': YEARS_CSV}

    ov = _call(views.overview, '/api/overview/', base, admin)
    assert ov.status_code == 200, f'overview {ov.status_code}'
    t = ov.data['totals']
    kpi_papers = t['papers']

    # 1) KPI papers == drill-down row count
    dd = _call(views.classified_papers, '/api/stats/classified-papers/', base, admin)
    assert dd.status_code == 200, f'classified_papers {dd.status_code}'
    dd_count = dd.data.get('count', len(dd.data.get('papers', [])))
    check(dd_count == kpi_papers,
          'KPI papers == drill-down count',
          f'KPI={kpi_papers} drill-down={dd_count}')

    # 2) Excel "Publications" value == KPI papers
    xreq = rf.get('/api/export/excel/', {'affiliation': 'albaha',
                                         'years': YEARS_CSV, 'sheets': 'summary'})
    force_authenticate(xreq, user=admin)
    xresp = exports.export_excel(xreq)
    from openpyxl import load_workbook
    # DRF/File response: pull the raw bytes however they're exposed.
    content = xresp.content if hasattr(xresp, 'content') else b''.join(xresp.streaming_content)
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb['Overview']
    # Overview card layout: row 5, col 3 is the Publications value (see exports.py).
    raw_val = ws.cell(row=5, column=3).value
    xlsx_papers = int(str(raw_val).replace(',', '').strip()) if raw_val is not None else -1
    check(xlsx_papers == kpi_papers,
          'Excel Publications == KPI papers',
          f'KPI={kpi_papers} excel={xlsx_papers}')

    # 3) Top Papers: no FALSE under albaha; every row carries the badge field
    tp = ov.data['top_papers']
    has_field = all('affiliation_verified' in p for p in tp)
    no_false = all(p.get('affiliation_verified') is not False for p in tp)
    check(has_field, 'Top Papers rows expose affiliation_verified (for badge)',
          f'{sum(1 for p in tp if "affiliation_verified" in p)}/{len(tp)} rows')
    check(no_false, 'Top Papers has NO confirmed-elsewhere (FALSE) paper under albaha',
          f'false_rows={sum(1 for p in tp if p.get("affiliation_verified") is False)}')

    # 4) Monotonicity: albaha KPI <= all KPI for each metric
    ov_all = _call(views.overview, '/api/overview/', all_, admin)
    assert ov_all.status_code == 200, f'overview(all) {ov_all.status_code}'
    ta = ov_all.data['totals']
    for k in ('papers', 'citations', 'q1_papers', 'q2_papers', 'q3_papers', 'q4_papers'):
        check(t[k] <= ta[k], f'albaha.{k} <= all.{k}', f'albaha={t[k]} all={ta[k]}')

    # Report
    print('=' * 68)
    print(f' Dashboard invariants — scope years: {YEARS_CSV}')
    print('=' * 68)
    failed = 0
    for ok, label, detail in results:
        mark = 'PASS' if ok else 'FAIL'
        if not ok:
            failed += 1
        print(f'  [{mark}] {label}' + (f'   ({detail})' if detail else ''))
    print('-' * 68)
    print(f'  {len(results) - failed}/{len(results)} invariants hold.')
    sys.exit(1 if failed else 0)


if __name__ == '__main__':
    main()
