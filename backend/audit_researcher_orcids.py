"""
Audit ORCID assignments across all researchers.

Reads Researcher.ORCID_ID (the actual column where 39 researchers'
ORCIDs live) and cross-references with what's in the papers.

OUTPUT
------
audit_orcid_assignments.xlsx with columns:
  UserID, FullName_Ar, FirstName, LastName, ORCID_ID,
  PapersWithORCID, MostCommonScrapedName, Verdict

Verdict values:
  LIKELY_CORRECT  - registered name appears in scraped names
  SUSPICIOUS      - scraped names don't match the registered name
  NO_PAPERS       - the ORCID is set but no papers in DB carry it
"""
import os, sys
import django
from collections import Counter
import re

if not os.environ.get("DJANGO_SETTINGS_MODULE"):
    os.environ["DJANGO_SETTINGS_MODULE"] = "litrix_backend.settings"
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

from django.db import connection
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def normalize(s):
    if not s: return ""
    s = s.lower()
    s = re.sub(r"[.,\-_']", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def tokens(s):
    return set(normalize(s).split())


def main():
    with connection.cursor() as cur:
        # Pull researchers from Researcher (the real source)
        cur.execute("""
            SELECT u."UserID", u."FullName_Ar", u."FirstName", u."LastName",
                   r."ORCID_ID"
            FROM "Researcher" r
            JOIN "Users" u ON u."UserID" = r."UserID"
            WHERE r."ORCID_ID" IS NOT NULL
              AND r."ORCID_ID" <> ''
            ORDER BY u."UserID"
        """)
        researchers = cur.fetchall()
        print(f"Researchers with ORCID set: {len(researchers)}")

        rows = []
        for user_id, name_ar, fn, ln, orcid in researchers:
            cur.execute("""
                SELECT ship->'author'->>'display_name' AS display_name
                FROM "ResearchPaper" rp
                CROSS JOIN LATERAL jsonb_array_elements(
                    rp."RawData_Log"->'authorships'
                ) AS ship
                WHERE REPLACE(
                        REPLACE(ship->'author'->>'orcid', 'https://orcid.org/', ''),
                        'http://orcid.org/', ''
                      ) = %s
            """, [orcid])
            names = [r[0] for r in cur.fetchall() if r[0]]
            paper_count = len(names)

            verdict = "NO_PAPERS"
            most_common = ""
            if names:
                counter = Counter(names)
                most_common = counter.most_common(1)[0][0]
                reg_en = normalize(f"{fn or ''} {ln or ''}")
                reg_tokens_en = tokens(reg_en)
                reg_tokens_ar = tokens(name_ar or "")
                scraped_tokens = tokens(most_common)
                overlap_en = (
                    len(reg_tokens_en & scraped_tokens) / max(1, len(reg_tokens_en))
                    if reg_tokens_en else 0
                )
                overlap_ar = (
                    len(reg_tokens_ar & scraped_tokens) / max(1, len(reg_tokens_ar))
                    if reg_tokens_ar else 0
                )
                if overlap_en >= 0.5 or overlap_ar >= 0.5:
                    verdict = "LIKELY_CORRECT"
                else:
                    verdict = "SUSPICIOUS"

            rows.append({
                "UserID":            user_id,
                "FullName_Ar":       name_ar or "",
                "FirstName":         fn or "",
                "LastName":          ln or "",
                "ORCID_ID":          orcid,
                "PapersWithORCID":   paper_count,
                "MostCommonScrapedName": most_common,
                "Verdict":           verdict,
            })

        verdicts = Counter(r["Verdict"] for r in rows)
        print("\nVerdict counts:")
        for v, n in verdicts.most_common():
            print(f"  {v:15s} {n}")

        # Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "ORCID Audit"

        ws["A1"] = "ORCID Assignment Audit - Litrix (Researcher.ORCID_ID)"
        ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1D1D1F")
        ws.merge_cells("A1:H1")

        ws["A2"] = ("SUSPICIOUS = the ORCID's papers list a different person's name. "
                    "Verify at orcid.org/<ID>")
        ws["A2"].font = Font(name="Arial", size=10, italic=True, color="6E6E73")
        ws.merge_cells("A2:H2")

        headers = ["UserID", "FullName_Ar", "FirstName", "LastName",
                   "ORCID_ID", "PapersWithORCID",
                   "MostCommonScrapedName", "Verdict"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = Font(name="Arial", bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", start_color="1D1D1F")

        order = {"SUSPICIOUS": 0, "NO_PAPERS": 1, "LIKELY_CORRECT": 2}
        rows.sort(key=lambda r: (order.get(r["Verdict"], 99),
                                  -r["PapersWithORCID"]))

        for i, r in enumerate(rows, 1):
            row_n = 4 + i
            ws.cell(row=row_n, column=1, value=r["UserID"])
            ws.cell(row=row_n, column=2, value=r["FullName_Ar"])
            ws.cell(row=row_n, column=3, value=r["FirstName"])
            ws.cell(row=row_n, column=4, value=r["LastName"])
            ws.cell(row=row_n, column=5, value=r["ORCID_ID"])
            ws.cell(row=row_n, column=6, value=r["PapersWithORCID"])
            ws.cell(row=row_n, column=7, value=r["MostCommonScrapedName"])
            cell_v = ws.cell(row=row_n, column=8, value=r["Verdict"])
            colors = {
                "SUSPICIOUS":     "FFE5E5",
                "NO_PAPERS":      "F5F5F7",
                "LIKELY_CORRECT": "E5F7E5",
            }
            cell_v.fill = PatternFill("solid", start_color=colors.get(r["Verdict"], "FFFFFF"))
            for col in range(1, 9):
                ws.cell(row=row_n, column=col).alignment = Alignment(
                    horizontal="left", vertical="center"
                )
                ws.cell(row=row_n, column=col).font = Font(name="Arial", size=10)

        widths = [8, 40, 18, 18, 22, 16, 30, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        out = "audit_orcid_assignments.xlsx"
        wb.save(out)
        print(f"\nSaved -> {out}")


if __name__ == "__main__":
    main()
