"""
Export the 18 Scholar orphan papers to Excel for review.

These are Source='Scholar' papers whose RawData_Log.citation_id (the
Scholar profile ID they were scraped from) does NOT match any current
Users.Scholar_ID. The original researcher was likely removed or had
their Scholar_ID changed.

OUTPUT: scholar_orphans_review.xlsx
"""
import os, sys
import django

if not os.environ.get("DJANGO_SETTINGS_MODULE"):
    os.environ["DJANGO_SETTINGS_MODULE"] = "litrix_backend.settings"
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

from django.db import connection
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


SQL = """
SELECT
    rp."PaperID",
    rp."Title",
    rp."PubYear",
    rp."DOI",
    j."JournalName",
    LEFT(rp."RawData_Log"->>'authors', 500)        AS authors_raw,
    rp."RawData_Log"->>'citation_id'               AS scraped_from_scholar_id,
    rp."ScrapedAt"
FROM "ResearchPaper" rp
LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
WHERE rp."Source" = 'Scholar'
  AND NOT EXISTS (
      SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID"
  )
ORDER BY rp."PubYear" DESC NULLS LAST, rp."PaperID"
"""


def main():
    with connection.cursor() as cur:
        cur.execute(SQL)
        rows = cur.fetchall()
    print(f"Scholar orphan papers to export: {len(rows)}")
    if not rows:
        print("Nothing to export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Scholar Orphans"

    ws["A1"] = "Scholar Orphan Papers - Review & Decide"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1D1D1F")
    ws.merge_cells("A1:J1")

    ws["A2"] = ("Scraped from a Scholar profile whose ID no longer exists in Users. "
                "Fill Decision = LINK / DELETE / KEEP.")
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="6E6E73")
    ws.merge_cells("A2:J2")

    headers = ["PaperID", "Title", "PubYear", "DOI", "Journal",
               "AuthorsRaw", "ScrapedFromScholarID",
               "Decision", "UserID", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1D1D1F")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for i, row in enumerate(rows, 1):
        r = 4 + i
        ws.cell(row=r, column=1, value=row[0])
        ws.cell(row=r, column=2, value=row[1])
        ws.cell(row=r, column=3, value=row[2])
        ws.cell(row=r, column=4, value=row[3] or "")
        ws.cell(row=r, column=5, value=row[4] or "")
        ws.cell(row=r, column=6, value=row[5] or "")
        ws.cell(row=r, column=7, value=row[6] or "")
        ws.cell(row=r, column=8, value="")  # Decision
        ws.cell(row=r, column=9, value="")  # UserID
        ws.cell(row=r, column=10, value="") # Notes
        for col in range(1, 11):
            c = ws.cell(row=r, column=col)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.font = Font(name="Arial", size=10)
        ws.cell(row=r, column=8).fill = PatternFill("solid", start_color="FFF4D6")
        ws.cell(row=r, column=9).fill = PatternFill("solid", start_color="FFF4D6")

    widths = [10, 55, 9, 28, 30, 50, 22, 12, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(5, 5 + len(rows)):
        ws.row_dimensions[r].height = 50

    out = "scholar_orphans_review.xlsx"
    wb.save(out)
    print(f"Saved -> {out}")
    print("\nNext:")
    print("  - If they're not Al-Baha papers: skip filling and run delete_scholar_orphans.py")
    print("  - If some are recoverable: fill Decision=LINK + UserID, then run apply_scholar_decisions.py")


if __name__ == "__main__":
    main()
