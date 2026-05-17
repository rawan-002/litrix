"""
Export the 92 Manual orphan papers to an Excel file for human review.

OUTPUT
------
manual_orphans_review.xlsx with columns:
    PaperID, Title, PubYear, DOI, Journal, AuthorsRaw,
    Decision    <- fill in: LINK / DELETE / KEEP
    UserID      <- fill in only if Decision = LINK
    Notes       <- optional free text

Open the file, walk through each row, then run apply_manual_decisions.py.

USAGE:
    python export_manual_orphans.py
"""
import os, sys
import django

if not os.environ.get("DJANGO_SETTINGS_MODULE"):
    os.environ["DJANGO_SETTINGS_MODULE"] = "litrix_backend.settings"
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

from django.db import connection
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


SQL = """
SELECT
    rp."PaperID",
    rp."Title",
    rp."PubYear",
    rp."DOI",
    j."JournalName",
    LEFT(rp."RawData_Log"->>'authors', 500) AS authors_raw,
    rp."ScrapedAt"
FROM "ResearchPaper" rp
LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
WHERE rp."Source" = 'Manual'
  AND NOT EXISTS (
      SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID"
  )
ORDER BY rp."PubYear" DESC NULLS LAST, rp."PaperID"
"""


def main():
    with connection.cursor() as cur:
        cur.execute(SQL)
        rows = cur.fetchall()
    print(f"Manual orphan papers to export: {len(rows)}")
    if not rows:
        print("Nothing to export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Manual Orphans"

    # Title row
    ws["A1"] = "Manual Orphan Papers - Review & Decide"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1D1D1F")
    ws.merge_cells("A1:I1")

    ws["A2"] = ("Fill 'Decision' with LINK / DELETE / KEEP. "
                "For LINK, also fill 'UserID' from the Researchers sheet of your Litrix export.")
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="6E6E73")
    ws.merge_cells("A2:I2")

    # Header
    headers = ["PaperID", "Title", "PubYear", "DOI", "Journal",
               "AuthorsRaw", "Decision", "UserID", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1D1D1F")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Data
    for i, row in enumerate(rows, 1):
        r = 4 + i
        ws.cell(row=r, column=1, value=row[0])  # PaperID
        ws.cell(row=r, column=2, value=row[1])  # Title
        ws.cell(row=r, column=3, value=row[2])  # PubYear
        ws.cell(row=r, column=4, value=row[3] or "")  # DOI
        ws.cell(row=r, column=5, value=row[4] or "")  # Journal
        ws.cell(row=r, column=6, value=row[5] or "")  # AuthorsRaw
        ws.cell(row=r, column=7, value="")  # Decision <- you fill
        ws.cell(row=r, column=8, value="")  # UserID <- you fill if LINK
        ws.cell(row=r, column=9, value="")  # Notes
        for col in range(1, 10):
            c = ws.cell(row=r, column=col)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.font = Font(name="Arial", size=10)
        # Highlight decision/userid columns lightly so they're easy to find
        ws.cell(row=r, column=7).fill = PatternFill("solid", start_color="FFF4D6")
        ws.cell(row=r, column=8).fill = PatternFill("solid", start_color="FFF4D6")

    widths = [10, 55, 9, 28, 30, 50, 12, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(5, 5 + len(rows)):
        ws.row_dimensions[r].height = 50

    out = "manual_orphans_review.xlsx"
    wb.save(out)
    print(f"Saved -> {out}")
    print("\nNext steps:")
    print("  1. Open the file, fill 'Decision' and 'UserID' columns.")
    print("  2. Save it (same filename).")
    print("  3. Run: python apply_manual_decisions.py")


if __name__ == "__main__":
    main()
