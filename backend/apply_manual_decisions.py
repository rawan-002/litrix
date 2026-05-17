"""
Apply decisions from manual_orphans_review.xlsx.

EXPECTED COLUMNS (case-insensitive):
    PaperID, Decision, UserID, Notes

DECISION VALUES:
    LINK   - link the paper to UserID (must be filled). MappingCriteria='admin_manual_review'.
    DELETE - cascade-delete the paper (snapshot to AuditLog like the OpenAlex script).
    KEEP   - leave as-is. Skipped.
    blank  - skipped, treated as KEEP.

USAGE:
    python apply_manual_decisions.py --dry-run
    python apply_manual_decisions.py
"""
import os, sys, json, argparse
import django

if not os.environ.get("DJANGO_SETTINGS_MODULE"):
    os.environ["DJANGO_SETTINGS_MODULE"] = "litrix_backend.settings"
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

from django.db import connection, transaction
from openpyxl import load_workbook


POSSIBLE_CHILDREN = [
    "ExternalAuthors", "PaperKeywords", "PaperGrants",
    "Citations", "CitationsHistory", "ReportPaperDecision",
]


def existing_child_tables(cur):
    cur.execute("""
        SELECT table_name FROM information_schema.tables
        WHERE table_schema = 'public' AND table_name = ANY(%s)
    """, [POSSIBLE_CHILDREN])
    return [r[0] for r in cur.fetchall()]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default="manual_orphans_review.xlsx")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = load_workbook(args.file)
    ws = wb.active

    # Header is on row 4 in our export.
    header_row = 4
    headers = {(c.value or "").strip().lower(): c.column
               for c in ws[header_row] if c.value}
    required = {"paperid", "decision"}
    if not required.issubset(headers.keys()):
        raise SystemExit(f"Missing required columns. Need: {required}")

    paper_col   = headers["paperid"]
    decision_col= headers["decision"]
    userid_col  = headers.get("userid")
    notes_col   = headers.get("notes")

    decisions = []
    for r in range(header_row + 1, ws.max_row + 1):
        pid = ws.cell(row=r, column=paper_col).value
        if not pid:
            continue
        dec = (ws.cell(row=r, column=decision_col).value or "").strip().upper()
        uid = ws.cell(row=r, column=userid_col).value if userid_col else None
        notes = ws.cell(row=r, column=notes_col).value if notes_col else None
        decisions.append((int(pid), dec, uid, notes))

    print(f"Rows read: {len(decisions)}")
    counts = {"LINK": 0, "DELETE": 0, "KEEP": 0, "OTHER": 0}
    for _, dec, _, _ in decisions:
        counts[dec if dec in counts else "OTHER"] += 1
    print(f"  LINK   : {counts['LINK']}")
    print(f"  DELETE : {counts['DELETE']}")
    print(f"  KEEP   : {counts['KEEP']}")
    print(f"  blank/other : {counts['OTHER']}")

    with transaction.atomic():
        with connection.cursor() as cur:
            tables = existing_child_tables(cur)

            n_linked = n_deleted = n_skipped = 0
            for paper_id, dec, uid, notes in decisions:
                if dec == "LINK":
                    if not uid:
                        print(f"  PaperID={paper_id} LINK without UserID - skipped.")
                        n_skipped += 1
                        continue
                    cur.execute("""
                        INSERT INTO "Authors" (
                            "UserID","PaperID","AuthorOrder",
                            "IsCorrespondingAuthor",
                            "MappingConfidence","MappingCriteria",
                            "AuthorNameRaw","Is_Verified"
                        )
                        VALUES (%s, %s, NULL, FALSE, 1.0,
                                'admin_manual_review', %s, TRUE)
                        ON CONFLICT ("UserID","PaperID") DO NOTHING
                    """, [int(uid), paper_id, (notes or "")[:255]])
                    n_linked += 1

                elif dec == "DELETE":
                    snapshot = {"PaperID": paper_id,
                                "Notes": notes,
                                "Reason": "manual_admin_decision"}
                    cur.execute(
                        'INSERT INTO "AuditLog" '
                        '("TenantID","UserID","Action","TargetType","TargetID",'
                        ' "Metadata","IpAddress","UserAgent") '
                        'VALUES (1, NULL, %s, %s, %s, %s::jsonb, NULL, %s)',
                        ["paper.delete.manual", "ResearchPaper", paper_id,
                         json.dumps(snapshot),
                         "backfill_script:apply_manual_decisions"]
                    )
                    for t in tables:
                        cur.execute(f'DELETE FROM "{t}" WHERE "PaperID" = %s',
                                    (paper_id,))
                    cur.execute('DELETE FROM "ResearchPaper" WHERE "PaperID" = %s',
                                (paper_id,))
                    n_deleted += 1
                else:
                    n_skipped += 1

            print(f"\nLinked : {n_linked}")
            print(f"Deleted: {n_deleted}")
            print(f"Skipped: {n_skipped}")

            if args.dry_run:
                transaction.set_rollback(True)
                print("\n--dry-run: rolled back. Nothing written.")
            else:
                print("\nDone. Changes committed.")


if __name__ == "__main__":
    main()
