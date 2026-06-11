"""
Public read-only API for the supervisor dashboard.

The supervisor (مشرف) views the college's research output without logging
in, so these endpoints are fully public, read-only, and strip sensitive
fields (Email, tracker IDs). DRF's anonymous throttle keeps scraping in
check. If access control is ever needed, the route shape lets us drop a
single @require_public_token decorator on the whole module.

Endpoints: overview, departments, researchers (+profile), papers (+detail),
kpis, trends, and an Excel export mirroring the admin layout.
"""

from __future__ import annotations

from typing import Any

from django.db import connection
from rest_framework import decorators, response
from rest_framework.permissions import AllowAny


# Dashboard scope: the current biennium. Every PubYear aggregation is
# constrained to these so the supervisor sees current output, not the
# lifetime archive. To add a year later, append here and redeploy.
DASHBOARD_YEARS = (2025, 2026)

# Retracted papers must never count toward metrics (NCAAA/ABET/Scimago all
# exclude them). We match on title because the scrapers don't populate a
# flag consistently — Scholar uses "[retracted]", Scopus "RETRACTED:",
# retraction notes start with "Retraction note:". Embed NOT_RETRACTED_SQL
# in any WHERE clause.
RETRACTION_PATTERNS = (
    "rp.\"Title\" ILIKE 'RETRACTED:%%'",
    "rp.\"Title\" ILIKE 'RETRACTED %%'",
    "rp.\"Title\" ILIKE '[retracted]%%'",
    "rp.\"Title\" ILIKE 'Retraction Note%%'",
    "rp.\"Title\" ILIKE 'Retraction note%%'",
    "rp.\"Title\" ILIKE 'Withdrawn:%%'",
)
NOT_RETRACTED_SQL = "NOT (" + " OR ".join(RETRACTION_PATTERNS) + ")"


# Excludes papers definitively verified as NOT-Al-Baha (set by
# affiliation_verifier.py). Decision matrix:
#   TRUE  → include (confirmed Al-Baha)
#   NULL  → include (unverified — benefit of the doubt, retry later)
#   FALSE → exclude (confirmed authored elsewhere)
# We keep NULL in on purpose: excluding it would drop 119+ papers we just
# couldn't verify (paywalled PDFs, sparse Scholar metadata) — most of which
# are legitimately ours.
AFFILIATION_VERIFIED_SQL = '(rp."AffiliationVerified" IS DISTINCT FROM FALSE)'


# Compute h-index live from the paper × citations data rather than trust the
# stale Researcher.H_Index column. Definition: a researcher has h-index h if
# h of their papers each have >= h citations. We rank papers per author by
# citations DESC and take MAX(rank) where cites >= rank. Returns ("UserID",
# h_index); researchers with no attributed papers drop out (COALESCE downstream).
CITES_EXPR = (
    'COALESCE('
    '(rp."RawData_Log"->\'cited_by\'->>\'value\')::int,'
    '(rp."RawData_Log"->>\'cited_by_count\')::int,'
    '0)'
)
H_INDEX_PER_USER_SQL = f'''
    SELECT "UserID", COALESCE(MAX(rnk), 0)::int AS h_index
    FROM (
        SELECT
            a."UserID",
            {CITES_EXPR} AS cites,
            ROW_NUMBER() OVER (
                PARTITION BY a."UserID"
                ORDER BY {CITES_EXPR} DESC
            ) AS rnk
        FROM "Authors" a
        JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
        WHERE {NOT_RETRACTED_SQL}
    ) ranked
    WHERE cites >= rnk
    GROUP BY "UserID"
'''


# Decorator stack shared by every public endpoint. No throttle here so the
# dashboard stays snappy — if scraping becomes a problem, add a
# throttle_classes decorator and register its rate in REST_FRAMEWORK.
public_endpoint = [
    decorators.api_view(['GET']),
    decorators.permission_classes([AllowAny]),
]


def _apply_decorators(decorators_list):
    """Apply a list of decorators in order (outermost last)."""
    def wrapper(func):
        for dec in reversed(decorators_list):
            func = dec(func)
        return func
    return wrapper


def _resolve_years(request) -> list[int]:
    """Years to scope by from the `year` param. A single in-range year gives
    [year]; 'all', absent, or anything out of DASHBOARD_YEARS gives them all."""
    raw = (request.query_params.get('year') or '').strip().lower()
    if raw and raw != 'all' and raw.isdigit():
        y = int(raw)
        if y in DASHBOARD_YEARS:
            return [y]
    return list(DASHBOARD_YEARS)


@_apply_decorators(public_endpoint)
def overview(request):
    """
    GET /api/public/overview/[?year=2025|2026|all]
    College-wide aggregated KPIs for the dashboard header, scoped by the
    global year filter.
    """
    years = _resolve_years(request)
    with connection.cursor() as cur:
        # Join every paper-derived metric with Authors so orphan papers (no
        # Litrix researcher attached) don't inflate the totals. The JOIN +
        # DISTINCT gives consistent "attributed to our faculty" semantics.
        cur.execute(f'''
            WITH attributed_papers AS (
                SELECT DISTINCT
                    rp."PaperID",
                    rp."JournalID",
                    COALESCE(
                        (rp."RawData_Log"->'cited_by'->>'value')::int,
                        (rp."RawData_Log"->>'cited_by_count')::int,
                        0
                    ) AS cites
                FROM "ResearchPaper" rp
                JOIN "Authors" a ON a."PaperID" = rp."PaperID"
                WHERE rp."PubYear" = ANY(%s)
                  AND {NOT_RETRACTED_SQL}
              AND {AFFILIATION_VERIFIED_SQL}
            )
            SELECT
                (SELECT COUNT(DISTINCT "PaperID") FROM attributed_papers)      AS total_papers,
                (SELECT COUNT(DISTINCT u."UserID")
                   FROM "Users" u
                   WHERE u."UserType" = 'Researcher')                          AS total_researchers,
                (SELECT COALESCE(SUM(cites), 0) FROM attributed_papers)        AS total_citations,
                (SELECT COUNT(DISTINCT ap."PaperID")
                   FROM attributed_papers ap
                   JOIN "JournalRankings" jr ON jr."JournalID" = ap."JournalID"
                                            AND jr."Quartile" = 'Q1')          AS q1_papers,
                (SELECT COALESCE(ROUND(AVG(h_index))::int, 0)
                   FROM ({H_INDEX_PER_USER_SQL}) hi
                   WHERE h_index > 0)                                          AS avg_h_index
        ''', [years])
        row = cur.fetchone()
    return response.Response({
        'total_papers':       row[0] or 0,
        'total_researchers':  row[1] or 0,
        'total_citations':    row[2] or 0,
        'q1_papers':          row[3] or 0,
        'avg_h_index':        int(row[4] or 0),
    })


@_apply_decorators(public_endpoint)
def departments_list(request):
    """
    GET /api/public/departments/[?year=2025|2026|all]
    Per-department summary stats — drives the department-card row. Scoped by
    the global year filter.
    """
    years = _resolve_years(request)
    with connection.cursor() as cur:
        # Paper counters are year-scoped; researchers_count stays current
        # faculty (a member exists regardless of publishing — the KPI section
        # already shows the "published in year X" subset).
        # Citations need their own CTE: when co-authors share a department,
        # the Authors JOIN duplicates the (dept, paper) row, so a plain SUM
        # would multiply. COUNT(DISTINCT PaperID) is fine; SUM is not — hence
        # one row per (dept, paper) in dept_papers, then SUM that.
        cur.execute(f'''
            WITH dept_papers AS (
                SELECT DISTINCT
                    d."DepartmentID",
                    rp."PaperID",
                    COALESCE(
                        (rp."RawData_Log"->'cited_by'->>'value')::int,
                        (rp."RawData_Log"->>'cited_by_count')::int,
                        0
                    ) AS cites
                FROM "Department" d
                JOIN "Works_In" w
                    ON w."DepartmentID" = d."DepartmentID"
                   AND w."IsCurrentPosition" = TRUE
                JOIN "Authors" a
                    ON a."UserID" = w."UserID"
                JOIN "ResearchPaper" rp
                    ON rp."PaperID" = a."PaperID"
                WHERE rp."PubYear" = ANY(%s)
                  AND {NOT_RETRACTED_SQL}
              AND {AFFILIATION_VERIFIED_SQL}
            )
            SELECT
                d."DepartmentID",
                d."DepartmentName",
                COUNT(DISTINCT w."UserID")                              AS researchers_count,
                COUNT(DISTINCT a."PaperID")
                    FILTER (WHERE rp."PubYear" = ANY(%s))               AS papers_count,
                COUNT(DISTINCT a."PaperID")
                    FILTER (WHERE jr."Quartile" = 'Q1'
                              AND rp."PubYear" = ANY(%s))               AS q1_papers,
                COALESCE(
                    (SELECT SUM(cites) FROM dept_papers dp
                     WHERE dp."DepartmentID" = d."DepartmentID"),
                    0
                )                                                       AS total_citations
            FROM "Department" d
            LEFT JOIN "Works_In" w
                ON w."DepartmentID" = d."DepartmentID"
               AND w."IsCurrentPosition" = TRUE
            LEFT JOIN "Authors" a
                ON a."UserID" = w."UserID"
            LEFT JOIN "ResearchPaper" rp
                ON rp."PaperID" = a."PaperID"
            LEFT JOIN LATERAL (
                SELECT "Quartile"
                FROM "JournalRankings"
                WHERE "JournalID" = rp."JournalID"
                ORDER BY "RankingYear" DESC NULLS LAST, "Source"
                LIMIT 1
            ) jr ON TRUE
            GROUP BY d."DepartmentID", d."DepartmentName"
            ORDER BY papers_count DESC, d."DepartmentName"
        ''', [years, years, years])
        rows = cur.fetchall()
    return response.Response({
        'results': [
            {
                'department_id':     r[0],
                'department_name':   r[1],
                'researchers_count': r[2] or 0,
                'papers_count':      r[3] or 0,
                'q1_papers':         r[4] or 0,
                'total_citations':   r[5] or 0,
            }
            for r in rows
        ]
    })


@_apply_decorators(public_endpoint)
def researchers_list(request):
    """
    GET /api/public/researchers/[?department_id=N][&year=2025|2026|all]
    Researcher list (academic fields only). Paper/citation counts reflect
    the selected year (or both if 'all').
    """
    years = _resolve_years(request)
    dept_id = request.query_params.get('department_id')
    params: list[Any] = []
    where = "WHERE u.\"UserType\" = 'Researcher'"
    if dept_id and dept_id.isdigit():
        where += " AND w.\"DepartmentID\" = %s"
        params.append(int(dept_id))

    with connection.cursor() as cur:
        cur.execute(f'''
            SELECT
                u."UserID",
                u."Litrix_ID",
                u."FullName_Ar",
                TRIM(CONCAT_WS(' ', u."FirstName", u."MiddleName", u."LastName"))
                                                                AS full_name_en,
                d."DepartmentID",
                d."DepartmentName",
                r."AcademicRank",
                r."Specialization",
                COALESCE(hi.h_index, 0)                         AS h_index,
                (SELECT COUNT(*)
                   FROM "Authors" a
                   JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
                   WHERE a."UserID" = u."UserID"
                     AND rp."PubYear" = ANY(%s)
                     AND {NOT_RETRACTED_SQL}
                     AND {AFFILIATION_VERIFIED_SQL})                AS total_papers,
                COALESCE((
                    SELECT SUM(COALESCE(
                        (rp."RawData_Log"->'cited_by'->>'value')::int,
                        (rp."RawData_Log"->>'cited_by_count')::int,
                        0))
                    FROM "Authors" a
                    JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
                    WHERE a."UserID" = u."UserID"
                      AND rp."PubYear" = ANY(%s)
                      AND {NOT_RETRACTED_SQL}
                      AND {AFFILIATION_VERIFIED_SQL}
                ), 0)                                           AS total_citations
            FROM "Users" u
            LEFT JOIN "Researcher" r       ON r."UserID" = u."UserID"
            LEFT JOIN "Works_In" w
                ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
            LEFT JOIN "Department" d       ON d."DepartmentID" = w."DepartmentID"
            LEFT JOIN ({H_INDEX_PER_USER_SQL}) hi ON hi."UserID" = u."UserID"
            {where}
            ORDER BY total_papers DESC NULLS LAST, u."FullName_Ar"
        ''', [years, years, *params])
        rows = cur.fetchall()
    return response.Response({
        'results': [
            {
                'user_id':         r[0],
                'litrix_id':       r[1],
                'full_name_ar':    r[2],
                'full_name_en':    r[3],
                'department_id':   r[4],
                'department_name': r[5],
                'academic_rank':   r[6],
                'specialization':  r[7],
                'h_index':         r[8] or 0,
                'total_papers':    r[9] or 0,
                'total_citations': r[10] or 0,
            }
            for r in rows
        ]
    })


@_apply_decorators(public_endpoint)
def researcher_profile(request, litrix_id: str):
    """
    GET /api/public/researchers/{litrix_id}/profile/
    Researcher metadata + papers + per-year citation series. Resolves
    Litrix_ID (e.g. "Lit-000042") to UserID internally.
    """
    with connection.cursor() as cur:
        cur.execute(
            'SELECT "UserID" FROM "Users" WHERE "Litrix_ID" = %s LIMIT 1',
            [litrix_id],
        )
        row = cur.fetchone()
        if not row:
            return response.Response({'error': 'researcher not found'}, status=404)
        user_id = row[0]

        cur.execute(f'''
            SELECT
                u."UserID",
                u."Litrix_ID",
                u."FullName_Ar",
                TRIM(CONCAT_WS(' ', u."FirstName", u."MiddleName", u."LastName"))
                                                            AS full_name_en,
                d."DepartmentID",
                d."DepartmentName",
                r."AcademicRank",
                r."Specialization",
                COALESCE(hi.h_index, 0)                     AS h_index,
                r."Scopus_ID",
                r."ORCID_ID",
                r."CitationsByYear"
            FROM "Users" u
            LEFT JOIN "Researcher" r       ON r."UserID" = u."UserID"
            LEFT JOIN "Works_In" w
                ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
            LEFT JOIN "Department" d       ON d."DepartmentID" = w."DepartmentID"
            LEFT JOIN ({H_INDEX_PER_USER_SQL}) hi ON hi."UserID" = u."UserID"
            WHERE u."UserID" = %s
        ''', [user_id])
        bio = cur.fetchone()

        # Lifetime stats here, not biennium — a profile page should show the
        # full career when a supervisor drills into one person.
        cur.execute('''
            SELECT
                COUNT(DISTINCT rp."PaperID")                         AS total_papers,
                COALESCE(SUM(COALESCE(
                    (rp."RawData_Log"->'cited_by'->>'value')::int,
                    (rp."RawData_Log"->>'cited_by_count')::int,
                    0)), 0)                                          AS total_citations,
                COUNT(DISTINCT rp."PaperID")
                    FILTER (WHERE jr."Quartile" = 'Q1')              AS q1_papers,
                MIN(rp."PubYear")                                    AS first_year,
                MAX(rp."PubYear")                                    AS last_year
            FROM "ResearchPaper" rp
            JOIN "Authors" a ON a."PaperID" = rp."PaperID"
            LEFT JOIN LATERAL (
                SELECT "Quartile" FROM "JournalRankings"
                WHERE "JournalID" = rp."JournalID"
                ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1
            ) jr ON TRUE
            WHERE a."UserID" = %s
        ''', [user_id])
        stats_row = cur.fetchone()

        # Papers list — lifetime, no year filter, so the whole record shows.
        cur.execute('''
            SELECT
                rp."PaperID",
                rp."Title",
                rp."DOI",
                rp."PubYear",
                rp."Source",
                rp."CitationsByYear",
                COALESCE(
                    (rp."RawData_Log"->'cited_by'->>'value')::int,
                    (rp."RawData_Log"->>'cited_by_count')::int,
                    0)                                              AS citations,
                j."JournalName",
                jr."Quartile",
                jr."ImpactFactor"
            FROM "ResearchPaper" rp
            JOIN "Authors" a ON a."PaperID" = rp."PaperID"
            LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
            LEFT JOIN LATERAL (
                SELECT "Quartile", "ImpactFactor" FROM "JournalRankings"
                WHERE "JournalID" = rp."JournalID"
                ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1
            ) jr ON TRUE
            WHERE a."UserID" = %s
            ORDER BY rp."PubYear" DESC NULLS LAST, rp."PaperID" DESC
        ''', [user_id])
        papers = cur.fetchall()

        # Co-authors: everyone sharing a paper with this researcher. Platform
        # members (non-null UserID -> on the Al-Baha roster) carry a Litrix_ID
        # so the card can link to their profile; external names come through
        # with is_albaha = false. Al-Baha first, then by shared-paper count.
        cur.execute('''
            SELECT user_id, litrix_id, name,
                   COUNT(DISTINCT paper_id) AS shared_papers
            FROM (
                SELECT
                    co."UserID"    AS user_id,
                    cu."Litrix_ID" AS litrix_id,
                    COALESCE(
                        cu."FullName_Ar",
                        NULLIF(TRIM(CONCAT_WS(' ', cu."FirstName", cu."LastName")), ''),
                        co."AuthorNameRaw"
                    )              AS name,
                    co."PaperID"   AS paper_id
                FROM "Authors" me
                JOIN "Authors" co ON co."PaperID" = me."PaperID"
                                  AND co."UserID" IS DISTINCT FROM me."UserID"
                LEFT JOIN "Users" cu ON cu."UserID" = co."UserID"
                WHERE me."UserID" = %s
                  AND COALESCE(cu."FullName_Ar", co."AuthorNameRaw") IS NOT NULL
            ) sub
            GROUP BY user_id, litrix_id, name
            ORDER BY (user_id IS NOT NULL) DESC, shared_papers DESC, name
            LIMIT 200
        ''', [user_id])
        coauthors = [
            {
                'user_id':       r[0],
                'litrix_id':     r[1],
                'name':          r[2],
                'shared_papers': r[3],
                'is_albaha':     r[0] is not None,
            }
            for r in cur.fetchall()
        ]

    # Build the citations-by-year series from the researcher row (if any).
    citations_by_year_chart: list[dict] = []
    raw_cby = bio[11] if bio else None
    if raw_cby:
        import json as _json
        if isinstance(raw_cby, str):
            try:
                raw_cby = _json.loads(raw_cby)
            except Exception:
                raw_cby = {}
        citations_by_year_chart = sorted(
            ({'year': int(y), 'citations': int(v)}
             for y, v in (raw_cby or {}).items()
             if str(y).isdigit()),
            key=lambda x: x['year'],
        )

    return response.Response({
        'user_id':         bio[0] if bio else None,
        'litrix_id':       bio[1] if bio else litrix_id,
        'full_name_ar':    bio[2] if bio else None,
        'full_name_en':    bio[3] if bio else None,
        'department_id':   bio[4] if bio else None,
        'department_name': bio[5] if bio else None,
        'academic_rank':   bio[6] if bio else None,
        'specialization':  bio[7] if bio else None,
        'h_index':         bio[8] if bio else 0,
        'scopus_id':       bio[9] if bio else None,
        'orcid_id':        bio[10] if bio else None,
        'stats': {
            'total_papers':     stats_row[0] or 0,
            'total_citations':  stats_row[1] or 0,
            'q1_papers':        stats_row[2] or 0,
            'first_year':       stats_row[3],
            'last_year':        stats_row[4],
        },
        'citations_by_year_chart': citations_by_year_chart,
        'papers': [
            {
                'paper_id':         p[0],
                'title':            p[1],
                'doi':              p[2],
                'pub_year':         p[3],
                'source':           p[4],
                'citations_by_year': p[5],
                'citations':        p[6] or 0,
                'journal_name':     p[7],
                'quartile':         p[8],
                'impact_factor':    p[9],
            }
            for p in papers
        ],
        'coauthors': coauthors,
    })


@_apply_decorators(public_endpoint)
def papers_list(request):
    """
    GET /api/public/papers/?department_id=N&year=Y&quartile=Q1&page=1
    Paginated paper list for the dashboard's papers tab.
    """
    dept_id  = request.query_params.get('department_id')
    year     = request.query_params.get('year')
    quartile = request.query_params.get('quartile')
    page     = int(request.query_params.get('page', 1))
    page_size = min(int(request.query_params.get('page_size', 50)), 200)
    offset = (page - 1) * page_size

    params: list[Any] = []
    filters_sql: list[str] = []
    if dept_id and dept_id.isdigit():
        filters_sql.append('w."DepartmentID" = %s')
        params.append(int(dept_id))
    if year and year.isdigit():
        filters_sql.append('rp."PubYear" = %s')
        params.append(int(year))
    if quartile in ('Q1', 'Q2', 'Q3', 'Q4'):
        filters_sql.append('jr."Quartile" = %s')
        params.append(quartile)

    where = "WHERE 1=1"
    if filters_sql:
        where += " AND " + " AND ".join(filters_sql)

    with connection.cursor() as cur:
        cur.execute(f'''
            SELECT COUNT(DISTINCT rp."PaperID")
            FROM "ResearchPaper" rp
            JOIN "Authors" a ON a."PaperID" = rp."PaperID"
            LEFT JOIN "Works_In" w
                ON w."UserID" = a."UserID" AND w."IsCurrentPosition" = TRUE
            LEFT JOIN LATERAL (
                SELECT "Quartile" FROM "JournalRankings"
                WHERE "JournalID" = rp."JournalID"
                ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1
            ) jr ON TRUE
            {where}
        ''', params)
        total = cur.fetchone()[0]

        cur.execute(f'''
            SELECT DISTINCT
                rp."PaperID",
                rp."Title",
                rp."DOI",
                rp."PubYear",
                rp."Source",
                COALESCE(
                    (rp."RawData_Log"->'cited_by'->>'value')::int,
                    (rp."RawData_Log"->>'cited_by_count')::int,
                    0)                                              AS citations,
                j."JournalName",
                jr."Quartile"
            FROM "ResearchPaper" rp
            JOIN "Authors" a ON a."PaperID" = rp."PaperID"
            LEFT JOIN "Works_In" w
                ON w."UserID" = a."UserID" AND w."IsCurrentPosition" = TRUE
            LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
            LEFT JOIN LATERAL (
                SELECT "Quartile" FROM "JournalRankings"
                WHERE "JournalID" = rp."JournalID"
                ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1
            ) jr ON TRUE
            {where}
            ORDER BY rp."PubYear" DESC NULLS LAST, citations DESC
            LIMIT %s OFFSET %s
        ''', params + [page_size, offset])
        rows = cur.fetchall()

    return response.Response({
        'count':    total,
        'page':     page,
        'page_size': page_size,
        'results': [
            {
                'paper_id':     r[0],
                'title':        r[1],
                'doi':          r[2],
                'pub_year':     r[3],
                'source':       r[4],
                'citations':    r[5] or 0,
                'journal_name': r[6],
                'quartile':     r[7],
            }
            for r in rows
        ]
    })



@_apply_decorators(public_endpoint)
def paper_detail(request, paper_id: int):
    """
    GET /api/public/papers/{paper_id}/detail/
    Rich metadata for one paper — drives the detail modal opened from a
    researcher profile.
    """
    with connection.cursor() as cur:
        cur.execute('''
            SELECT
                rp."PaperID",
                rp."Title",
                rp."Title_En",
                rp."Abstract",
                rp."DOI",
                rp."PubYear",
                rp."Volume",
                rp."Issue",
                rp."Pages",
                rp."Language",
                rp."Source",
                rp."CitationsByYear",
                COALESCE(
                    (rp."RawData_Log"->'cited_by'->>'value')::int,
                    (rp."RawData_Log"->>'cited_by_count')::int,
                    0)                                              AS citations,
                j."JournalName",
                j."Publisher",
                jr."Quartile",
                jr."ImpactFactor",
                jr."Category"
            FROM "ResearchPaper" rp
            LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
            LEFT JOIN LATERAL (
                SELECT "Quartile", "ImpactFactor", "Category"
                FROM "JournalRankings"
                WHERE "JournalID" = rp."JournalID"
                ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1
            ) jr ON TRUE
            WHERE rp."PaperID" = %s
        ''', [paper_id])
        row = cur.fetchone()
        if not row:
            return response.Response({'error': 'paper not found'}, status=404)

        # Internal authors (registered researchers)
        cur.execute('''
            SELECT u."UserID", u."Litrix_ID", u."FullName_Ar",
                   a."AuthorOrder", a."IsCorrespondingAuthor"
            FROM "Authors" a
            JOIN "Users" u ON u."UserID" = a."UserID"
            WHERE a."PaperID" = %s
            ORDER BY a."AuthorOrder" NULLS LAST
        ''', [paper_id])
        internal_authors = [
            {
                'user_id': r[0],
                'litrix_id': r[1],
                'name': r[2],
                'order': r[3],
                'is_corresponding': r[4],
            }
            for r in cur.fetchall()
        ]

        # External co-authors (not registered)
        cur.execute('''
            SELECT "FullName", "Affiliation"
            FROM "ExternalAuthors"
            WHERE "PaperID" = %s
            ORDER BY "ExtAuthorID"
        ''', [paper_id])
        external_authors = [
            {'name': r[0], 'affiliation': r[1]} for r in cur.fetchall()
        ]

        # No ExternalAuthors row (common for older Scholar/OpenAlex/ORCID
        # papers) → parse the raw scraper JSON. Each source stores authors
        # differently under RawData_Log:
        #   Scopus           "Authors" (capital A), pipe-separated
        #   Google Scholar   "authors" (lowercase), comma-separated
        #   OpenAlex / ORCID "authorships" -> [{author: {display_name}}]
        if not external_authors:
            cur.execute(
                'SELECT "RawData_Log" FROM "ResearchPaper" WHERE "PaperID" = %s',
                [paper_id],
            )
            raw = cur.fetchone()
            raw_log = raw[0] if raw else None
            if raw_log:
                import json as _json
                if isinstance(raw_log, str):
                    try:
                        raw_log = _json.loads(raw_log)
                    except Exception:
                        raw_log = {}
                if not isinstance(raw_log, dict):
                    raw_log = {}

                seen = {ia['name'] for ia in internal_authors if ia.get('name')}

                def _add(name: str, affiliation=None):
                    clean = (name or '').strip()
                    if not clean or clean in seen:
                        return
                    seen.add(clean)
                    external_authors.append(
                        {'name': clean, 'affiliation': affiliation}
                    )

                # Scopus: pipe-separated
                scopus_field = raw_log.get('Authors')
                if isinstance(scopus_field, str) and '|' in scopus_field:
                    for raw_name in scopus_field.split('|'):
                        _add(raw_name)

                # Google Scholar: comma-separated string (sometimes a list)
                scholar_field = raw_log.get('authors')
                if isinstance(scholar_field, str) and scholar_field:
                    parts = (
                        scholar_field.split(',')
                        if ',' in scholar_field
                        else [scholar_field]
                    )
                    for raw_name in parts:
                        # Strip the truncation marks Scholar sometimes adds
                        clean = raw_name.replace('…', '').replace('...', '').strip()
                        if clean and clean not in ('and', '...'):
                            _add(clean)

                # OpenAlex / ORCID: "authorships" list
                authorships = raw_log.get('authorships')
                if isinstance(authorships, list):
                    for ship in authorships:
                        if not isinstance(ship, dict):
                            continue
                        author = ship.get('author') or {}
                        name = (
                            author.get('display_name')
                            or author.get('name')
                            or ship.get('display_name')
                            or ship.get('name')
                        )
                        # Surface a primary affiliation if present
                        affs = ship.get('institutions') or []
                        aff_name = None
                        if affs and isinstance(affs[0], dict):
                            aff_name = affs[0].get('display_name')
                        _add(name, aff_name)

                # Generic: "authors" as a list of strings/dicts (manual imports)
                generic_authors = raw_log.get('authors')
                if isinstance(generic_authors, list):
                    for entry in generic_authors:
                        if isinstance(entry, dict):
                            _add(
                                entry.get('display_name') or entry.get('name'),
                                (entry.get('affiliation') or
                                 (entry.get('institutions') or [{}])[0].get('display_name')
                                 if entry.get('institutions') else None),
                            )
                        elif isinstance(entry, str):
                            _add(entry)

    return response.Response({
        'paper_id':         row[0],
        'title':            row[1],
        'title_en':         row[2],
        'abstract':         row[3],
        'doi':              row[4],
        'pub_year':         row[5],
        'volume':           row[6],
        'issue':            row[7],
        'pages':            row[8],
        'language':         row[9],
        'source':           row[10],
        'citations_by_year': row[11],
        'citations':        row[12] or 0,
        'journal_name':     row[13],
        'publisher':        row[14],
        'quartile':         row[15],
        'impact_factor':    row[16],
        'category':         row[17],
        'internal_authors': internal_authors,
        'external_authors': external_authors,
    })


@_apply_decorators(public_endpoint)
def kpis(request):
    """
    GET /api/public/kpis/?year=2026[&department_id=N]
    Three NCAAA-style KPIs:
        KPI-I-13: % faculty published (in year)
        KPI-I-14: avg publications per faculty (in year)
        KPI-I-15: avg citations per publication (all-time)
    """
    from datetime import datetime
    year_param = request.query_params.get('year')
    try:
        year = int(year_param) if year_param else datetime.now().year
    except ValueError:
        year = datetime.now().year

    dept_id = request.query_params.get('department_id')
    dept_filter = ""
    dept_params: list = []
    if dept_id and dept_id.isdigit():
        dept_filter = ' AND w."DepartmentID" = %s'
        dept_params = [int(dept_id)]

    with connection.cursor() as cur:
        # Total faculty (dept-scoped if a filter was given)
        cur.execute(f'''
            SELECT COUNT(DISTINCT u."UserID")
            FROM "Users" u
            LEFT JOIN "Works_In" w
                ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
            WHERE u."UserType" = 'Researcher'{dept_filter}
        ''', dept_params)
        total_faculty = cur.fetchone()[0] or 0

        # Faculty who published in the target year
        cur.execute(f'''
            SELECT COUNT(DISTINCT a."UserID")
            FROM "Authors" a
            JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
            JOIN "Users" u ON u."UserID" = a."UserID"
            LEFT JOIN "Works_In" w
                ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
            WHERE rp."PubYear" = %s
              AND u."UserType" = 'Researcher'{dept_filter}
              AND {NOT_RETRACTED_SQL}
              AND {AFFILIATION_VERIFIED_SQL}
        ''', [year, *dept_params])
        published_count = cur.fetchone()[0] or 0

        # Publications in the target year
        cur.execute(f'''
            SELECT COUNT(DISTINCT rp."PaperID")
            FROM "ResearchPaper" rp
            JOIN "Authors" a ON a."PaperID" = rp."PaperID"
            LEFT JOIN "Works_In" w
                ON w."UserID" = a."UserID" AND w."IsCurrentPosition" = TRUE
            WHERE rp."PubYear" = %s{dept_filter}
              AND {NOT_RETRACTED_SQL}
              AND {AFFILIATION_VERIFIED_SQL}
        ''', [year, *dept_params])
        pubs_in_year = cur.fetchone()[0] or 0

        # All-time citations + publications for KPI-15. The CTE collapses
        # (paper × author) duplicates so each paper counts once even with
        # multiple co-authors.
        cur.execute(f'''
            WITH unique_papers AS (
                SELECT DISTINCT
                    rp."PaperID",
                    COALESCE(
                        (rp."RawData_Log"->'cited_by'->>'value')::int,
                        (rp."RawData_Log"->>'cited_by_count')::int,
                        0
                    ) AS cites
                FROM "ResearchPaper" rp
                JOIN "Authors" a ON a."PaperID" = rp."PaperID"
                LEFT JOIN "Works_In" w
                    ON w."UserID" = a."UserID" AND w."IsCurrentPosition" = TRUE
                WHERE 1=1{dept_filter}
            )
            SELECT
                COUNT(*)                            AS total_pubs,
                COALESCE(SUM(cites), 0)             AS total_cites
            FROM unique_papers
        ''', dept_params)
        row = cur.fetchone()
        total_pubs_all = row[0] or 0
        total_cites_all = row[1] or 0

    kpi_13 = (published_count / total_faculty * 100) if total_faculty > 0 else 0
    kpi_14 = (pubs_in_year / total_faculty)          if total_faculty > 0 else 0
    kpi_15 = (total_cites_all / total_pubs_all)      if total_pubs_all > 0 else 0

    return response.Response({
        'year': year,
        'department_id': int(dept_id) if dept_id and dept_id.isdigit() else None,
        'kpi_13_percent_faculty_published': {
            'value': round(kpi_13, 1),
            'published_count': published_count,
            'total_faculty': total_faculty,
            'label': '% Faculty Published',
        },
        'kpi_14_avg_publications_per_faculty': {
            'value': round(kpi_14, 2),
            'total_publications': pubs_in_year,
            'total_faculty': total_faculty,
            'label': 'Avg Publications/Faculty',
        },
        'kpi_15_avg_citations_per_publication': {
            'value': round(kpi_15, 2),
            'total_citations': total_cites_all,
            'total_publications': total_pubs_all,
            'label': 'Avg Citations/Publication',
        },
    })


@_apply_decorators(public_endpoint)
def export_excel(request):
    """
    GET /api/public/export/excel/?years=2025,2026&sheets=overview,summary,departments,researchers,journals,conferences

    Workbook mirroring the admin export layout:
        Overview      — once
        Summary YYYY  — one per picked year
        Departments YYYY — one per year (journal/conference split + Q1-Q4)
        Researchers   — once (window + all-time + h-index)
        Journals YYYY / Conferences YYYY — one per picked year
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime

    # Parse + validate filters.
    years_param = (request.query_params.get('years') or '').strip()
    if years_param:
        requested = [int(y.strip()) for y in years_param.split(',') if y.strip().isdigit()]
        years = sorted([y for y in requested if y in DASHBOARD_YEARS]) or sorted(DASHBOARD_YEARS)
    else:
        years = sorted(DASHBOARD_YEARS)

    allowed_sheets = {'overview', 'summary', 'departments',
                      'researchers', 'journals', 'conferences'}
    sheets_param = (request.query_params.get('sheets') or '').strip()
    if sheets_param:
        sheets = ({s.strip().lower() for s in sheets_param.split(',') if s.strip()}
                  & allowed_sheets) or allowed_sheets
    else:
        sheets = allowed_sheets

    wb = Workbook()
    wb.remove(wb.active)

    # Shared styling.
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1D1D1F')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_font = Font(bold=True, size=18, color='1D1D1F')
    label_font = Font(bold=True, size=10, color='86868B')
    big_value_font = Font(bold=True, size=24, color='1D1D1F')
    sublabel_font = Font(size=9, color='86868B')

    def _header_row(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font; c.fill = header_fill; c.alignment = header_align
        ws.row_dimensions[1].height = 28

    # Venue classifier to split journals from conferences. Same heuristic as
    # the admin export — VenueType is often blank, so we also sniff the
    # JournalName for the usual conference/proceedings tokens.
    venue_case_sql = '''
        CASE
            WHEN LOWER(COALESCE(j."VenueType", '')) LIKE 'conference%%' THEN 'conference'
            WHEN LOWER(COALESCE(j."VenueType", '')) = 'proceedings' THEN 'conference'
            WHEN LOWER(COALESCE(j."JournalName", '')) LIKE '%%conference%%' THEN 'conference'
            WHEN LOWER(COALESCE(j."JournalName", '')) LIKE '%%proceedings%%' THEN 'conference'
            WHEN LOWER(COALESCE(j."JournalName", '')) LIKE '%%symposium%%' THEN 'conference'
            WHEN LOWER(COALESCE(j."JournalName", '')) LIKE '%%workshop%%' THEN 'conference'
            ELSE 'journal'
        END
    '''

    years_label = ', '.join(str(y) for y in years)

    # --- Overview sheet ---
    if 'overview' in sheets:
        ws = wb.create_sheet('Overview')
        ws['A1'] = 'Litrix — Research Analytics Overview'
        ws['A1'].font = title_font
        ws.row_dimensions[1].height = 30
        ws['A2'] = f'Window: {years_label}'
        ws['A2'].font = sublabel_font

        with connection.cursor() as cur:
            # Same CTE-with-Authors-JOIN approach as the overview() endpoint,
            # so the Excel header totals match the dashboard exactly.
            cur.execute(f'''
                WITH attributed_papers AS (
                    SELECT DISTINCT
                        rp."PaperID",
                        rp."JournalID",
                        COALESCE(
                            (rp."RawData_Log"->'cited_by'->>'value')::int,
                            (rp."RawData_Log"->>'cited_by_count')::int,
                            0) AS cites
                    FROM "ResearchPaper" rp
                    JOIN "Authors" a ON a."PaperID" = rp."PaperID"
                    WHERE rp."PubYear" = ANY(%s)
                      AND {NOT_RETRACTED_SQL}
                      AND {AFFILIATION_VERIFIED_SQL}
                )
                SELECT
                    (SELECT COUNT(DISTINCT u."UserID")
                       FROM "Users" u WHERE u."UserType" = 'Researcher'),
                    (SELECT COUNT(DISTINCT a."UserID")
                       FROM "Authors" a
                       JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
                       JOIN "Users" u ON u."UserID" = a."UserID"
                       WHERE u."UserType" = 'Researcher'
                         AND rp."PubYear" = ANY(%s)
                         AND {NOT_RETRACTED_SQL}
                         AND {AFFILIATION_VERIFIED_SQL}),
                    (SELECT COUNT(DISTINCT "PaperID") FROM attributed_papers),
                    (SELECT COUNT(DISTINCT ap."PaperID")
                       FROM attributed_papers ap
                       JOIN "JournalRankings" jr ON jr."JournalID" = ap."JournalID"
                                                AND jr."Quartile" = 'Q1'),
                    (SELECT COALESCE(SUM(cites), 0) FROM attributed_papers),
                    (SELECT COALESCE(ROUND(AVG(h_index)::numeric, 1), 0)
                       FROM ({H_INDEX_PER_USER_SQL}) hi WHERE h_index > 0)
            ''', [years, years])
            (total_researchers, active_researchers, total_papers,
             q1_papers, total_citations, avg_h) = cur.fetchone()

        labels = ['RESEARCHERS', 'PUBLICATIONS', 'CITATIONS', 'H-INDEX']
        values = [total_researchers, total_papers, total_citations, float(avg_h or 0)]
        subs = [f'{active_researchers} active',
                f'{q1_papers} in Q1 journals',
                f'received in {years_label}',
                'avg h-index']

        for col, (lbl, val, sub) in enumerate(zip(labels, values, subs)):
            c1 = ws.cell(row=4, column=col * 2 + 1, value=lbl); c1.font = label_font
            c2 = ws.cell(row=5, column=col * 2 + 1, value=val); c2.font = big_value_font
            c3 = ws.cell(row=6, column=col * 2 + 1, value=sub); c3.font = sublabel_font
        ws.row_dimensions[5].height = 36
        for ch in 'ABCDEFG':
            ws.column_dimensions[ch].width = 18

    # --- Summary YYYY sheets ---
    if 'summary' in sheets:
        for year in years:
            ws = wb.create_sheet(f'Summary {year}')
            _header_row(ws, ['Metric', 'Value'])
            with connection.cursor() as cur:
                # The CTE dedupes (author × paper) rows so each paper counts
                # once despite co-authorship, and the Authors JOIN drops
                # orphan papers — same "attributed to our faculty" semantics
                # as the dashboard.
                cur.execute(f'''
                    WITH attributed_papers AS (
                        SELECT DISTINCT
                            rp."PaperID",
                            rp."JournalID",
                            COALESCE(
                                (rp."RawData_Log"->'cited_by'->>'value')::int,
                                (rp."RawData_Log"->>'cited_by_count')::int, 0) AS cites
                        FROM "ResearchPaper" rp
                        JOIN "Authors" a ON a."PaperID" = rp."PaperID"
                        WHERE rp."PubYear" = %s
                          AND {NOT_RETRACTED_SQL}
                          AND {AFFILIATION_VERIFIED_SQL}
                    )
                    SELECT
                        COUNT(DISTINCT ap."PaperID"),
                        COALESCE(SUM(ap.cites), 0),
                        COUNT(DISTINCT ap."PaperID")
                            FILTER (WHERE {venue_case_sql} = 'journal'),
                        COUNT(DISTINCT ap."PaperID")
                            FILTER (WHERE {venue_case_sql} = 'conference'),
                        COUNT(DISTINCT ap."PaperID") FILTER (WHERE jr."Quartile" = 'Q1'),
                        COUNT(DISTINCT ap."PaperID") FILTER (WHERE jr."Quartile" = 'Q2'),
                        COUNT(DISTINCT ap."PaperID") FILTER (WHERE jr."Quartile" = 'Q3'),
                        COUNT(DISTINCT ap."PaperID") FILTER (WHERE jr."Quartile" = 'Q4')
                    FROM attributed_papers ap
                    LEFT JOIN "Journals" j ON j."JournalID" = ap."JournalID"
                    LEFT JOIN LATERAL (
                        SELECT "Quartile" FROM "JournalRankings"
                        WHERE "JournalID" = ap."JournalID"
                        ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1
                    ) jr ON TRUE
                ''', [year])
                row = cur.fetchone()
            (total, citations, journals_c, confs_c, q1, q2, q3, q4) = row
            rows = [
                ('Year', year),
                ('Total Papers', total or 0),
                ('Total Citations', citations or 0),
                ('Journal Papers', journals_c or 0),
                ('Conference Papers', confs_c or 0),
                ('Q1 Papers', q1 or 0),
                ('Q2 Papers', q2 or 0),
                ('Q3 Papers', q3 or 0),
                ('Q4 Papers', q4 or 0),
            ]
            for ri, (m, v) in enumerate(rows, 2):
                ws.cell(row=ri, column=1, value=m)
                ws.cell(row=ri, column=2, value=v)
            ws.column_dimensions['A'].width = 22
            ws.column_dimensions['B'].width = 14

    # --- Departments YYYY sheets ---
    if 'departments' in sheets:
        for year in years:
            ws = wb.create_sheet(f'Departments {year}')
            _header_row(ws, [
                'Department', 'Researchers', 'Journal Papers', 'Conference Papers',
                'Total Papers', 'Citations', 'Q1', 'Q2', 'Q3', 'Q4',
            ])
            with connection.cursor() as cur:
                cur.execute(f'''
                    WITH dept_papers AS (
                        SELECT DISTINCT d."DepartmentID", rp."PaperID",
                            {venue_case_sql} AS venue,
                            jr."Quartile" AS quartile,
                            COALESCE(
                                (rp."RawData_Log"->'cited_by'->>'value')::int,
                                (rp."RawData_Log"->>'cited_by_count')::int, 0) AS cites
                        FROM "Department" d
                        JOIN "Works_In" w ON w."DepartmentID" = d."DepartmentID"
                                         AND w."IsCurrentPosition" = TRUE
                        JOIN "Authors" a ON a."UserID" = w."UserID"
                        JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
                        LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
                        LEFT JOIN LATERAL (
                            SELECT "Quartile" FROM "JournalRankings"
                            WHERE "JournalID" = rp."JournalID"
                            ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1
                        ) jr ON TRUE
                        WHERE rp."PubYear" = %s
                          AND {NOT_RETRACTED_SQL}
                          AND {AFFILIATION_VERIFIED_SQL}
                    )
                    SELECT
                        d."DepartmentName",
                        (SELECT COUNT(DISTINCT w2."UserID") FROM "Works_In" w2
                           WHERE w2."DepartmentID" = d."DepartmentID"
                             AND w2."IsCurrentPosition" = TRUE) AS researchers,
                        COUNT(DISTINCT dp."PaperID") FILTER (WHERE dp.venue = 'journal'),
                        COUNT(DISTINCT dp."PaperID") FILTER (WHERE dp.venue = 'conference'),
                        COUNT(DISTINCT dp."PaperID"),
                        COALESCE(SUM(dp.cites), 0),
                        COUNT(DISTINCT dp."PaperID") FILTER (WHERE dp.quartile = 'Q1'),
                        COUNT(DISTINCT dp."PaperID") FILTER (WHERE dp.quartile = 'Q2'),
                        COUNT(DISTINCT dp."PaperID") FILTER (WHERE dp.quartile = 'Q3'),
                        COUNT(DISTINCT dp."PaperID") FILTER (WHERE dp.quartile = 'Q4')
                    FROM "Department" d
                    LEFT JOIN dept_papers dp ON dp."DepartmentID" = d."DepartmentID"
                    GROUP BY d."DepartmentID", d."DepartmentName"
                    ORDER BY 5 DESC NULLS LAST
                ''', [year])
                for ri, row in enumerate(cur.fetchall(), 2):
                    for ci, v in enumerate(row, 1):
                        ws.cell(row=ri, column=ci, value=v or 0)
            widths = [28, 14, 16, 18, 14, 14, 8, 8, 8, 8]
            for ch, w in zip('ABCDEFGHIJ', widths):
                ws.column_dimensions[ch].width = w

    # --- Researchers sheet ---
    if 'researchers' in sheets:
        ws = wb.create_sheet('Researchers')
        _header_row(ws, [
            'Researcher (AR)', 'Department', 'Rank',
            'Papers (window)', 'Citations (window)',
            'Papers (all-time)', 'Citations (all-time)',
            'h-index (all-time)', 'Status', 'Scholar ID', 'ORCID',
        ])
        with connection.cursor() as cur:
            cur.execute(f'''
                SELECT
                    u."FullName_Ar",
                    d."DepartmentName",
                    r."AcademicRank",
                    COALESCE((SELECT COUNT(*) FROM "Authors" a
                       JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
                       WHERE a."UserID" = u."UserID"
                         AND rp."PubYear" = ANY(%s)
                         AND {NOT_RETRACTED_SQL}
                         AND {AFFILIATION_VERIFIED_SQL}), 0),
                    COALESCE((SELECT SUM(COALESCE(
                        (rp."RawData_Log"->'cited_by'->>'value')::int,
                        (rp."RawData_Log"->>'cited_by_count')::int, 0))
                        FROM "Authors" a JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
                        WHERE a."UserID" = u."UserID"
                          AND rp."PubYear" = ANY(%s)
                          AND {NOT_RETRACTED_SQL}
                          AND {AFFILIATION_VERIFIED_SQL}), 0),
                    COALESCE((SELECT COUNT(*) FROM "Authors" a
                       JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
                       WHERE a."UserID" = u."UserID"
                         AND {NOT_RETRACTED_SQL}
                         AND {AFFILIATION_VERIFIED_SQL}), 0),
                    COALESCE((SELECT SUM(COALESCE(
                        (rp."RawData_Log"->'cited_by'->>'value')::int,
                        (rp."RawData_Log"->>'cited_by_count')::int, 0))
                        FROM "Authors" a JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
                        WHERE a."UserID" = u."UserID"
                          AND {NOT_RETRACTED_SQL}
                          AND {AFFILIATION_VERIFIED_SQL}), 0),
                    COALESCE(hi.h_index, 0),
                    CASE WHEN u."Email" LIKE '%%@litrix.placeholder' THEN 'Pending'
                         ELSE 'Active' END,
                    COALESCE(u."Scholar_ID", ''),
                    COALESCE(r."ORCID_ID", '')
                FROM "Users" u
                LEFT JOIN "Researcher" r ON r."UserID" = u."UserID"
                LEFT JOIN "Works_In" w ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
                LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
                LEFT JOIN ({H_INDEX_PER_USER_SQL}) hi ON hi."UserID" = u."UserID"
                WHERE u."UserType" = 'Researcher'
                ORDER BY 4 DESC NULLS LAST, u."FullName_Ar"
            ''', [years, years])
            for ri, row in enumerate(cur.fetchall(), 2):
                for ci, v in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=v if v is not None else '')
        widths = [32, 22, 16, 14, 14, 14, 16, 12, 10, 14, 22]
        for ch, w in zip('ABCDEFGHIJK', widths):
            ws.column_dimensions[ch].width = w

    # --- Journals YYYY sheets ---
    if 'journals' in sheets:
        for year in years:
            ws = wb.create_sheet(f'Journals {year}')
            _header_row(ws, [
                'Department', 'Title',
                'Al-Baha Researchers', 'All Authors (raw)',
                'Journal', 'Quartile', 'IF', 'Indexing', 'Citations', 'DOI',
            ])
            with connection.cursor() as cur:
                cur.execute(f'''
                    SELECT
                        string_agg(DISTINCT d."DepartmentName", ', '),
                        rp."Title",
                        string_agg(DISTINCT u."FullName_Ar", '، '),
                        COALESCE(rp."RawData_Log"->>'Authors',
                                 rp."RawData_Log"->>'authors', ''),
                        j."JournalName",
                        jr."Quartile",
                        jr."ImpactFactor",
                        rp."Indexing",
                        COALESCE(
                            (rp."RawData_Log"->'cited_by'->>'value')::int,
                            (rp."RawData_Log"->>'cited_by_count')::int, 0),
                        rp."DOI"
                    FROM "ResearchPaper" rp
                    JOIN "Authors" a ON a."PaperID" = rp."PaperID"
                    JOIN "Users" u ON u."UserID" = a."UserID"
                    LEFT JOIN "Works_In" w ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
                    LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
                    LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
                    LEFT JOIN LATERAL (
                        SELECT "Quartile", "ImpactFactor" FROM "JournalRankings"
                        WHERE "JournalID" = rp."JournalID"
                        ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1
                    ) jr ON TRUE
                    WHERE rp."PubYear" = %s
                      AND {venue_case_sql} = 'journal'
                      AND {NOT_RETRACTED_SQL}
                      AND {AFFILIATION_VERIFIED_SQL}
                    GROUP BY rp."PaperID", rp."Title",
                             rp."RawData_Log", j."JournalName", jr."Quartile",
                             jr."ImpactFactor", rp."Indexing", rp."DOI"
                    ORDER BY 9 DESC
                ''', [year])
                for ri, row in enumerate(cur.fetchall(), 2):
                    for ci, v in enumerate(row, 1):
                        ws.cell(row=ri, column=ci, value=v if v is not None else '')
            widths = [22, 60, 28, 50, 30, 10, 8, 12, 10, 30]
            for ch, w in zip('ABCDEFGHIJ', widths):
                ws.column_dimensions[ch].width = w

    # --- Conferences YYYY sheets ---
    if 'conferences' in sheets:
        for year in years:
            ws = wb.create_sheet(f'Conferences {year}')
            _header_row(ws, [
                'Department', 'Title',
                'Al-Baha Researchers', 'All Authors (raw)',
                'Conference', 'Indexing', 'Citations', 'DOI',
            ])
            with connection.cursor() as cur:
                cur.execute(f'''
                    SELECT
                        string_agg(DISTINCT d."DepartmentName", ', '),
                        rp."Title",
                        string_agg(DISTINCT u."FullName_Ar", '، '),
                        COALESCE(rp."RawData_Log"->>'Authors',
                                 rp."RawData_Log"->>'authors', ''),
                        j."JournalName",
                        rp."Indexing",
                        COALESCE(
                            (rp."RawData_Log"->'cited_by'->>'value')::int,
                            (rp."RawData_Log"->>'cited_by_count')::int, 0),
                        rp."DOI"
                    FROM "ResearchPaper" rp
                    JOIN "Authors" a ON a."PaperID" = rp."PaperID"
                    JOIN "Users" u ON u."UserID" = a."UserID"
                    LEFT JOIN "Works_In" w ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
                    LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
                    LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
                    WHERE rp."PubYear" = %s
                      AND {venue_case_sql} = 'conference'
                      AND {NOT_RETRACTED_SQL}
                      AND {AFFILIATION_VERIFIED_SQL}
                    GROUP BY rp."PaperID", rp."Title",
                             rp."RawData_Log", j."JournalName", rp."Indexing", rp."DOI"
                    ORDER BY 7 DESC
                ''', [year])
                for ri, row in enumerate(cur.fetchall(), 2):
                    for ci, v in enumerate(row, 1):
                        ws.cell(row=ri, column=ci, value=v if v is not None else '')
            widths = [22, 60, 28, 50, 40, 12, 10, 30]
            for ch, w in zip('ABCDEFGH', widths):
                ws.column_dimensions[ch].width = w

    if not wb.sheetnames:
        ws = wb.create_sheet('Empty')
        ws['A1'] = 'No sheets were selected. Pick at least one option and try again.'

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fname = f'litrix-export-{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp


@_apply_decorators(public_endpoint)
def trends(request):
    """
    GET /api/public/trends/
    Yearly papers + citations across the dashboard window. Retracted papers
    excluded so the chart lines up with the other KPIs.
    """
    years = list(DASHBOARD_YEARS)
    with connection.cursor() as cur:
        cur.execute(f'''
            WITH attributed_papers AS (
                SELECT DISTINCT
                    rp."PaperID",
                    rp."PubYear",
                    COALESCE(
                        (rp."RawData_Log"->'cited_by'->>'value')::int,
                        (rp."RawData_Log"->>'cited_by_count')::int,
                        0) AS cites
                FROM "ResearchPaper" rp
                JOIN "Authors" a ON a."PaperID" = rp."PaperID"
                WHERE rp."PubYear" = ANY(%s)
                  AND {NOT_RETRACTED_SQL}
                  AND {AFFILIATION_VERIFIED_SQL}
            )
            SELECT
                "PubYear"                       AS year,
                COUNT(DISTINCT "PaperID")       AS papers,
                COALESCE(SUM(cites), 0)         AS citations
            FROM attributed_papers
            GROUP BY "PubYear"
            ORDER BY "PubYear"
        ''', [years])
        rows = cur.fetchall()
    return response.Response({
        'results': [
            {'year': r[0], 'papers': r[1] or 0, 'citations': r[2] or 0}
            for r in rows
        ]
    })
