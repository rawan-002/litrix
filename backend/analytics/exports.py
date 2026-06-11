"""Excel export for the dashboard (admin + HoD scoped). Builds the .xlsx
workbook; every number runs through the same stats helpers as the overview so
the file matches what's on screen."""
from rest_framework import decorators, response
from django.http import HttpResponse
from django.db import connection

from .stats import (
    FOCUS_YEARS, _albaha_only, _resolve_years,
    _cites_expr, _affil_clause, _hod_scope_department_id,
)


def _excel_response(filename: str):
    """Helper: build an HttpResponse with the right xlsx headers."""
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@decorators.api_view(['GET'])
def export_excel(request):
    """GET /api/export/excel/?years=2025,2026&sheets=summary,departments,researchers,journals,conferences

    Builds an xlsx workbook from the years + sheets the dashboard's options
    modal picks (defaults to all of both). With everything selected you get a
    Summary, Departments, Journals, and Conferences sheet per year, plus a
    single Researchers sheet of contributions in the window.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.db import connection

    # The export carries per-researcher rows (names, Scholar/ORCID IDs) and full
    # paper lists, so limit it to roles that may view researchers. Otherwise a
    # plain Researcher would fall into the unscoped branch below and pull the
    # whole institution.
    _u = getattr(request, 'user', None)
    if not (_u and _u.is_authenticated and (
            _u.has_litrix_perm('view_all_researchers') or
            _u.has_litrix_perm('view_dept_researchers'))):
        return response.Response({'error': 'Forbidden'}, status=403)

    # A HoD may only export their own department, so every sheet below is
    # filtered to hod_dept_id. Admin/Dean get the full institution (hod_dept_id
    # is None, so the `scoped` clauses are skipped). The helper returns the
    # sentinel -1 for a HoD with no department at all.
    hod_dept_id = _hod_scope_department_id(request)
    if hod_dept_id == -1:
        return response.Response(
            {'error': 'You are not assigned to any department. '
                      'Contact an admin to set your department.'},
            status=403,
        )
    scoped = hod_dept_id is not None

    # Department label for the Overview sheet, so a HoD's file makes its
    # scope obvious.
    hod_dept_name = None
    if scoped:
        with connection.cursor() as _c:
            _c.execute(
                'SELECT "DepartmentName" FROM "Department" WHERE "DepartmentID" = %s',
                [hod_dept_id],
            )
            _row = _c.fetchone()
            hod_dept_name = _row[0] if _row else None

    # EXISTS predicate: the paper has at least one author whose current
    # position is in the HoD's department. Appended to paper queries when
    # scoped; rp_alias is the ResearchPaper alias in the target query.
    def _dept_paper_clause(rp_alias: str) -> str:
        return (
            f' AND EXISTS (SELECT 1 FROM "Authors" a_dept '
            f'             JOIN "Works_In" w_dept ON w_dept."UserID" = a_dept."UserID" '
            f'                                   AND w_dept."IsCurrentPosition" = TRUE '
            f'             WHERE a_dept."PaperID" = {rp_alias}."PaperID" '
            f'               AND w_dept."DepartmentID" = %s)'
        )

    years_param = request.query_params.get('years', '').strip()
    if years_param:
        try:
            years = [int(y) for y in years_param.split(',') if y.strip().isdigit()]
        except ValueError:
            years = list(FOCUS_YEARS)
    else:
        years = list(FOCUS_YEARS)
    if not years:
        years = list(FOCUS_YEARS)

    # In "Al-Baha only" mode (?affiliation=albaha) every paper/citation query
    # below drops papers confirmed authored elsewhere, so the workbook matches
    # the screen.
    albaha_only = _albaha_only(request)
    AFFIL     = _affil_clause(albaha_only, 'rp')
    AFFIL_RP2 = _affil_clause(albaha_only, 'rp2')
    AFFIL_RPW = _affil_clause(albaha_only, 'rp_w')

    sheets_param = request.query_params.get('sheets', '').strip()
    if sheets_param:
        sheets = {s.strip().lower() for s in sheets_param.split(',') if s.strip()}
    else:
        sheets = {'summary', 'departments', 'researchers', 'journals', 'conferences'}

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1D1D1F')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # KPI overview as the first sheet so it opens by default; mirrors the
    # dashboard cards (Researchers / Publications / Citations / h-index).
    if 'summary' in sheets or 'departments' in sheets or 'researchers' in sheets:
        ws_overview = wb.create_sheet('Overview', 0)
        big_value_font = Font(bold=True, size=28, color='1D1D1F')
        label_font = Font(bold=True, size=11, color='86868B')
        sublabel_font = Font(size=10, color='86868B')
        title_font = Font(bold=True, size=18, color='1D1D1F')
        bg_fill = PatternFill('solid', fgColor='FAFAFA')

        ws_overview['A1'] = 'Litrix — Research Analytics Overview'
        ws_overview['A1'].font = title_font
        ws_overview.row_dimensions[1].height = 36
        years_label = ', '.join(str(y) for y in sorted(years))
        scope_label = f' · Department: {hod_dept_name}' if hod_dept_name else ''
        ws_overview['A2'] = f'Window: {years_label}{scope_label}'
        ws_overview['A2'].font = sublabel_font

        # Same per-year semantics as the dashboard. Each KPI is its own small
        # query so the HoD filter drops in cleanly without juggling param order.
        year_strs = [str(y) for y in years]
        with connection.cursor() as cur:
            # Researchers + active (scoped to current dept positions).
            if scoped:
                cur.execute('''
                    SELECT COUNT(DISTINCT u."UserID"),
                           COUNT(DISTINCT u."UserID") FILTER (WHERE r."LastSyncedAt" IS NOT NULL)
                    FROM "Users" u
                    JOIN "Works_In" w ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
                    LEFT JOIN "Researcher" r ON r."UserID" = u."UserID"
                    WHERE u."UserType" = 'Researcher' AND w."DepartmentID" = %s
                ''', [hod_dept_id])
            else:
                cur.execute('''
                    SELECT COUNT(DISTINCT u."UserID"),
                           COUNT(DISTINCT u."UserID") FILTER (WHERE r."LastSyncedAt" IS NOT NULL)
                    FROM "Users" u
                    LEFT JOIN "Researcher" r ON r."UserID" = u."UserID"
                    WHERE u."UserType" = 'Researcher'
                ''')
            r_total, r_active = cur.fetchone()

            # Papers in window + Q1 (author-in-dept when scoped).
            papers_sql = (
                'SELECT COUNT(*), COUNT(*) FILTER (WHERE q = \'Q1\') FROM ('
                '  SELECT DISTINCT rp."PaperID", jr."Quartile" AS q'
                '  FROM "ResearchPaper" rp'
                '  LEFT JOIN LATERAL (SELECT "Quartile" FROM "JournalRankings"'
                '      WHERE "JournalID" = rp."JournalID"'
                '      ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1) jr ON TRUE'
                '  WHERE rp."PubYear" = ANY(%s)'
                '    AND EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID")'
            )
            papers_params = [years]
            if scoped:
                papers_sql += _dept_paper_clause('rp')
                papers_params.append(hod_dept_id)
            papers_sql += AFFIL
            papers_sql += ') sub'
            cur.execute(papers_sql, papers_params)
            p_total, p_q1 = cur.fetchone()

            # Citations received in the selected years, per-paper
            # (ResearchPaper.CitationsByYear) — same definition as the overview
            # dashboard so the two agree. EXISTS, not JOIN, so each paper counts
            # once across its co-authors.
            year_keys_expr = _cites_expr('rp', years)
            cit_sql = (
                f'SELECT COALESCE(SUM({year_keys_expr}), 0) FROM "ResearchPaper" rp '
                'WHERE EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID"'
            )
            cit_params = list(year_strs)
            if scoped:
                cit_sql += (
                    ' AND EXISTS (SELECT 1 FROM "Works_In" w '
                    '             WHERE w."UserID" = a."UserID" '
                    '               AND w."IsCurrentPosition" = TRUE '
                    '               AND w."DepartmentID" = %s)'
                )
                cit_params.append(hod_dept_id)
            cit_sql += ')'
            cit_sql += AFFIL
            cur.execute(cit_sql, cit_params)
            c_total = cur.fetchone()[0]

            # Average h-index, scoped to dept researchers.
            if scoped:
                cur.execute('''
                    SELECT ROUND(AVG(hi.h_index)::numeric, 1)
                    FROM v_researcher_h_index hi
                    JOIN "Works_In" w ON w."UserID" = hi."UserID"
                                     AND w."IsCurrentPosition" = TRUE
                    WHERE w."DepartmentID" = %s
                ''', [hod_dept_id])
            else:
                cur.execute('SELECT ROUND(AVG(h_index)::numeric, 1) FROM v_researcher_h_index')
            avg_h = cur.fetchone()[0]

        # Four cards in a row, each spanning two columns (A-B, C-D, E-F, G-H).
        cards = [
            ('Researchers',  str(r_total),                f'{r_active} active'),
            ('Publications', f'{p_total:,}',              f'{p_q1} in Q1 journals'),
            ('Citations',    f'{c_total:,}',              f'received in {years_label}'),
            ('h-index',      str(avg_h or 0),             'avg h-index'),
        ]
        for idx, (label, value, sub) in enumerate(cards):
            col_label = chr(ord('A') + idx * 2)
            col_value = chr(ord('A') + idx * 2)  # same col, multiple rows

            cell_label = ws_overview.cell(row=4, column=idx * 2 + 1, value=label.upper())
            cell_label.font = label_font
            cell_label.fill = bg_fill

            cell_val = ws_overview.cell(row=5, column=idx * 2 + 1, value=value)
            cell_val.font = big_value_font
            cell_val.fill = bg_fill
            ws_overview.row_dimensions[5].height = 44

            cell_sub = ws_overview.cell(row=6, column=idx * 2 + 1, value=sub)
            cell_sub.font = sublabel_font
            cell_sub.fill = bg_fill

        # Widen the KPI columns
        for col in ['A', 'C', 'E', 'G']:
            ws_overview.column_dimensions[col].width = 22
        for col in ['B', 'D', 'F', 'H']:
            ws_overview.column_dimensions[col].width = 4  # spacer

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
        ws.row_dimensions[1].height = 26

    def set_widths(ws, widths):
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = w

    def wrap_column(ws, col_idx: int):
        """Wrap text in every body cell of a column — used for the Abstract
        column so long paragraphs don't overflow into the next cell."""
        from openpyxl.styles import Alignment
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical='top',
            )

    if 'summary' in sheets:
        for year in sorted(years):
            ws = wb.create_sheet(f'Summary {year}')
            ws.append(['Metric', 'Value'])
            style_header(ws, 2)
            with connection.cursor() as cur:
                # Paper counts per year (published-in-year semantics)
                summary_sql = (
                    'SELECT '
                    '    COUNT(DISTINCT rp."PaperID"), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q1\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q2\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q3\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q4\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE j."VenueType" = \'Journal\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE j."VenueType" = \'Conference\') '
                    'FROM "ResearchPaper" rp '
                    'LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID" '
                    'LEFT JOIN LATERAL (SELECT "Quartile", "ImpactFactor" FROM "JournalRankings" '
                    '  WHERE "JournalID" = rp."JournalID" ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1) jr ON TRUE '
                    'WHERE rp."PubYear" = %s '
                    '  AND EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID")'
                )
                summary_params = [year]
                if scoped:
                    summary_sql += _dept_paper_clause('rp')
                    summary_params.append(hod_dept_id)
                summary_sql += AFFIL
                cur.execute(summary_sql, summary_params)
                p, q1, q2, q3, q4, jp, cp = cur.fetchone()

                # Citations received this year, per-paper — same definition as
                # the overview dashboard; EXISTS counts each paper once.
                cit_sql = (
                    f'SELECT COALESCE(SUM({_cites_expr("rp", [year])}), 0) '
                    'FROM "ResearchPaper" rp '
                    'WHERE EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID"'
                )
                cit_params = [str(year)]
                if scoped:
                    cit_sql += (
                        ' AND EXISTS (SELECT 1 FROM "Works_In" w '
                        '             WHERE w."UserID" = a."UserID" '
                        '               AND w."IsCurrentPosition" = TRUE '
                        '               AND w."DepartmentID" = %s)'
                    )
                    cit_params.append(hod_dept_id)
                cit_sql += ')'
                cit_sql += AFFIL
                cur.execute(cit_sql, cit_params)
                c = cur.fetchone()[0]
            for label, val in [
                ('Year', year),
                ('Total Papers', p),
                ('Total Citations', c),
                ('Journal Papers', jp),
                ('Conference Papers', cp),
                ('Q1 Papers', q1),
                ('Q2 Papers', q2),
                ('Q3 Papers', q3),
                ('Q4 Papers', q4),
            ]:
                ws.append([label, val])
            set_widths(ws, [25, 20])

    if 'departments' in sheets:
        for year in sorted(years):
            ws = wb.create_sheet(f'Departments {year}')
            ws.append([
                'Department', 'Researchers',
                'Journal Papers', 'Conference Papers', 'Total Papers',
                'Citations', 'Q1', 'Q2', 'Q3', 'Q4'
            ])
            style_header(ws, 10)
            with connection.cursor() as cur:
                # Citations received this year, per-paper, same definition as
                # the overview dashboard. DISTINCT on (dept, paper) keeps a paper
                # co-authored by two members of the same department from being
                # double-counted for that department.
                dept_sql = (
                    'WITH dept_cites AS ('
                    '    SELECT dept AS "DepartmentID", SUM(cites) AS cites FROM ('
                    '        SELECT DISTINCT w."DepartmentID" AS dept, rp2."PaperID" AS pid, '
                    f'               {_cites_expr("rp2", [year])} AS cites '
                    '        FROM "Works_In" w '
                    '        JOIN "Authors" a2 ON a2."UserID" = w."UserID" '
                    '        JOIN "ResearchPaper" rp2 ON rp2."PaperID" = a2."PaperID" '
                    '        WHERE w."IsCurrentPosition" = TRUE' + AFFIL_RP2 +
                    '    ) ded GROUP BY dept'
                    ') '
                    'SELECT '
                    '    d."DepartmentName", '
                    '    COUNT(DISTINCT u."UserID"), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE j."VenueType" = \'Journal\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE j."VenueType" = \'Conference\'), '
                    '    COUNT(DISTINCT rp."PaperID"), '
                    '    COALESCE(MAX(dc.cites), 0) AS citations, '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q1\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q2\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q3\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q4\') '
                    'FROM "Department" d '
                    'LEFT JOIN "Works_In" w ON w."DepartmentID" = d."DepartmentID" AND w."IsCurrentPosition" = TRUE '
                    'LEFT JOIN "Users" u ON u."UserID" = w."UserID" AND u."UserType" = \'Researcher\' '
                    'LEFT JOIN "Authors" a ON a."UserID" = u."UserID" '
                    'LEFT JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID" AND rp."PubYear" = %s' + AFFIL + ' '
                    'LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID" '
                    'LEFT JOIN LATERAL (SELECT "Quartile", "ImpactFactor" FROM "JournalRankings" '
                    '  WHERE "JournalID" = rp."JournalID" ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1) jr ON TRUE '
                    'LEFT JOIN dept_cites dc ON dc."DepartmentID" = d."DepartmentID" '
                )
                dept_params = [str(year), year]
                if scoped:
                    dept_sql += 'WHERE d."DepartmentID" = %s '
                    dept_params.append(hod_dept_id)
                dept_sql += 'GROUP BY d."DepartmentName" ORDER BY 5 DESC'
                cur.execute(dept_sql, dept_params)
                for row in cur.fetchall():
                    ws.append(list(row))
            set_widths(ws, [32, 12, 14, 16, 12, 12, 8, 8, 8, 8])

    if 'researchers' in sheets:
        ws = wb.create_sheet('Researchers')
        ws.append([
            'Researcher (AR)', 'Department', 'Rank',
            'Papers (window)', 'Citations (window)',
            'Papers (all-time)', 'Citations (all-time)',
            'h-index (all-time)', 'Status',
            'Scholar ID', 'ORCID'
        ])
        style_header(ws, 11)
        with connection.cursor() as cur:
            researchers_sql = '''
                SELECT
                    u."FullName_Ar",
                    d."DepartmentName",
                    r."AcademicRank",
                    -- Window stats. Papers = published in the focus years.
                    -- Citations = RECEIVED in the focus years across ALL the
                    -- researcher's papers (per-paper ResearchPaper.CitationsByYear),
                    -- matching the overview dashboard's definition.
                    COUNT(DISTINCT a_w."PaperID") FILTER (WHERE rp_w."PubYear" = ANY(%(years)s)) AS papers_window,
                    COALESCE(SUM({cite_window}), 0) AS citations_window,
                    -- All-time stats. Citations here stay the lifetime snapshot
                    -- (every citation the paper ever earned) — a distinct,
                    -- explicitly-labelled "all-time" column.
                    COUNT(DISTINCT a_w."PaperID") AS papers_all,
                    COALESCE(SUM(COALESCE((rp_w."RawData_Log"->'cited_by'->>'value')::int, 0)), 0) AS citations_all,
                    COALESCE(hi.h_index, 0) AS h_index,
                    -- Status with clear priority:
                    --   1. Has papers in the focus window  → Active
                    --   2. Has all-time papers (manual OR scraped) → Historical
                    --   3. Has a public profile but no papers yet → Pending Sync
                    --   4. Nothing at all → No Profile (manual outreach needed)
                    CASE
                        WHEN COUNT(DISTINCT a_w."PaperID")
                                FILTER (WHERE rp_w."PubYear" = ANY(%(years)s)) > 0
                            THEN 'Active'
                        WHEN COUNT(DISTINCT a_w."PaperID") > 0
                            THEN 'Historical'
                        WHEN u."Scholar_ID" IS NOT NULL
                          OR r."ORCID_ID" IS NOT NULL
                          OR r."OpenAlex_AuthorID" IS NOT NULL
                          OR r."Scopus_ID" IS NOT NULL
                            THEN 'Pending Sync'
                        ELSE 'No Profile'
                    END AS sync_status,
                    u."Scholar_ID",
                    r."ORCID_ID"
                FROM "Users" u
                JOIN "Researcher" r ON r."UserID" = u."UserID"
                LEFT JOIN "Works_In" w ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
                LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
                LEFT JOIN "Authors" a_w ON a_w."UserID" = u."UserID"
                LEFT JOIN "ResearchPaper" rp_w ON rp_w."PaperID" = a_w."PaperID"{affil_rpw}
                LEFT JOIN v_researcher_h_index hi ON hi."UserID" = u."UserID"
                WHERE u."UserType" = 'Researcher'
                {dept_filter}
                GROUP BY u."UserID", u."FullName_Ar", d."DepartmentName",
                         r."AcademicRank", hi.h_index, u."Scholar_ID",
                         r."ORCID_ID", r."OpenAlex_AuthorID", r."Scopus_ID",
                         r."LastSyncedAt"
                ORDER BY papers_window DESC, h_index DESC
            '''
            # Named params so the per-year keys coexist with the dict-style
            # %(years)s / %(dept)s params.
            cite_window_expr = ' + '.join([
                f'COALESCE((rp_w."CitationsByYear"->>%(y{i})s)::int, 0)'
                for i in range(len(years))
            ]) or '0'
            researchers_params = {'years': years}
            researchers_params.update({f'y{i}': str(y) for i, y in enumerate(years)})
            if scoped:
                researchers_sql = researchers_sql.format(
                    cite_window=cite_window_expr, affil_rpw=AFFIL_RPW,
                    dept_filter='AND w."DepartmentID" = %(dept)s')
                researchers_params['dept'] = hod_dept_id
            else:
                researchers_sql = researchers_sql.format(
                    cite_window=cite_window_expr, affil_rpw=AFFIL_RPW, dept_filter='')
            cur.execute(researchers_sql, researchers_params)
            for row in cur.fetchall():
                ws.append(list(row))
        set_widths(ws, [32, 22, 18, 12, 14, 14, 16, 14, 14, 18, 22])

    if 'journals' in sheets:
        for year in sorted(years):
            ws = wb.create_sheet(f'Journals {year}')
            # Two author columns: Al-Baha researchers only (Arabic), then all
            # authors combined (Arabic + foreign, no affiliations).
            ws.append([
                'Department', 'Title', 'Abstract',
                'Al-Baha Researchers', 'All Authors (raw)',
                'Journal', 'Quartile', 'IF', 'Indexing',
                'Citations', 'DOI'
            ])
            style_header(ws, 11)
            with connection.cursor() as cur:
                # v_paper_details has no paper_id, so to attach Abstract we
                # match ResearchPaper by DOI when available and fall back to a
                # case-insensitive title match when either side lacks a DOI.
                journals_sql = (
                    'SELECT '
                    '    v.department_name, v.title, '
                    '    rp."Abstract" AS abstract, '
                    '    v.authors_ar, v.all_authors_en, '
                    '    v.journal_name, v.quartile, v.impact_factor, v.indexing, '
                    '    v.citations, v.doi '
                    'FROM v_paper_details v '
                    'LEFT JOIN "ResearchPaper" rp ON '
                    '    (v.doi IS NOT NULL AND LOWER(rp."DOI") = LOWER(v.doi)) '
                    '    OR ((v.doi IS NULL OR rp."DOI" IS NULL) AND LOWER(rp."Title") = LOWER(v.title)) '
                    'WHERE v.pub_year = %s AND v.venue_type = \'Journal\''
                )
                journals_params = [year]
                if scoped:
                    journals_sql += ' AND v.department_id = %s'
                    journals_params.append(hod_dept_id)
                journals_sql += AFFIL
                journals_sql += ' ORDER BY v.department_name, v.citations DESC NULLS LAST'
                cur.execute(journals_sql, journals_params)
                for row in cur.fetchall():
                    ws.append(list(row))
            set_widths(ws, [22, 60, 80, 50, 50, 30, 10, 8, 12, 10, 30])
            wrap_column(ws, 3)  # Abstract column

    if 'conferences' in sheets:
        for year in sorted(years):
            ws = wb.create_sheet(f'Conferences {year}')
            ws.append([
                'Department', 'Title', 'Abstract',
                'Al-Baha Researchers', 'All Authors (raw)',
                'Conference', 'Indexing', 'Citations', 'DOI'
            ])
            style_header(ws, 9)
            with connection.cursor() as cur:
                conf_sql = (
                    'SELECT '
                    '    v.department_name, v.title, '
                    '    rp."Abstract" AS abstract, '
                    '    v.authors_ar, v.all_authors_en, '
                    '    v.journal_name, v.indexing, v.citations, v.doi '
                    'FROM v_paper_details v '
                    'LEFT JOIN "ResearchPaper" rp ON '
                    '    (v.doi IS NOT NULL AND LOWER(rp."DOI") = LOWER(v.doi)) '
                    '    OR ((v.doi IS NULL OR rp."DOI" IS NULL) AND LOWER(rp."Title") = LOWER(v.title)) '
                    'WHERE v.pub_year = %s AND v.venue_type = \'Conference\''
                )
                conf_params = [year]
                if scoped:
                    conf_sql += ' AND v.department_id = %s'
                    conf_params.append(hod_dept_id)
                conf_sql += AFFIL
                conf_sql += ' ORDER BY v.department_name, v.citations DESC NULLS LAST'
                cur.execute(conf_sql, conf_params)
                for row in cur.fetchall():
                    ws.append(list(row))
            set_widths(ws, [22, 60, 80, 50, 50, 30, 12, 10, 30])
            wrap_column(ws, 3)  # Abstract column

    if not wb.sheetnames:
        ws = wb.create_sheet('Empty')
        ws.append(['No sheets selected'])

    fname = f"litrix_export_{'-'.join(map(str, sorted(years)))}.xlsx"
    resp = _excel_response(fname)
    wb.save(resp)
    return resp
