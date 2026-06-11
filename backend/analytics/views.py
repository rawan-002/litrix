"""Read-only DRF endpoints for the dashboard and reports.

Everything here is ReadOnlyModelViewSet — the frontend never writes through
these; the scraper and bootstrap scripts own all writes. Filtering goes
through django-filter (e.g. ?department_id=2). The overview/export endpoints
are scoped to FOCUS_YEARS, so widen or narrow that list to move the window.
"""
import re

from rest_framework import viewsets, filters, decorators, response
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Sum, Count, Avg
from django.http import HttpResponse


# Canonical Litrix-ID is "Lit-" + 6 zero-padded digits. We accept anything
# loose (LIT-42, lit-0042, Lit-000042) and normalize before a DB lookup so
# manually-typed and differently-cased URLs still resolve.
LITRIX_ID_PATTERN = re.compile(r'^lit-(\d+)$', re.IGNORECASE)


def normalize_litrix_id(raw):
    """Return canonical Lit-NNNNNN if `raw` looks like a Litrix-ID, else None."""
    if not isinstance(raw, str):
        return None
    m = LITRIX_ID_PATTERN.match(raw.strip())
    if not m:
        return None
    return f'Lit-{int(m.group(1)):06d}'

from .models import (
    ResearcherStats, DepartmentStats, TopPaper, PublicationTrend,
    ResearchPaper,
)
from .serializers import (
    ResearcherStatsSerializer, DepartmentStatsSerializer,
    TopPaperSerializer, PublicationTrendSerializer,
    ResearchPaperSerializer,
)

from .stats import (
    YEAR_FLOOR, CHART_YEAR_FLOOR, FOCUS_YEARS,
    _resolve_years, _hod_scope_department_id, _albaha_only,
    _cites_expr, _affil_clause,
    _dept_cards_windowed, _researcher_rows_windowed,
)


@decorators.api_view(['GET'])
def paper_detail(request, paper_id):
    """GET /api/papers/<paper_id>/detail/

    Full paper details for the modal popup, pulled from ResearchPaper plus the
    original scraped RawData_Log: metadata, raw author string, citation graph,
    journal info, and the Al-Baha researchers attributed to the paper.
    """
    from django.db import connection
    with connection.cursor() as cur:
        cur.execute('''
            SELECT
                rp."PaperID",
                rp."Title",
                rp."Abstract",
                rp."DOI",
                rp."PubYear",
                rp."Source",
                rp."Indexing",
                rp."CitationsByYear",
                rp."RawData_Log"->>'authors'      AS raw_authors,
                rp."RawData_Log"->>'publication'  AS publication,
                rp."RawData_Log"->>'link'         AS link,
                rp."RawData_Log"->>'citation_id'  AS cites_id,
                COALESCE(
                    ("RawData_Log"->'cited_by'->>'value')::int,
                    ("RawData_Log"->>'cited_by_count')::int,
                    0
                ) AS total_citations,
                COALESCE(j."JournalName", rp."RawData_Log"->>'publication') AS journal_name,
                j."ISSN_Print",
                j."VenueType",
                jr."Quartile",
                jr."ImpactFactor",
                rp."RawData_Log"->'authorships'   AS authorships_jsonb
            FROM "ResearchPaper" rp
            LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
            LEFT JOIN LATERAL (    SELECT "Quartile", "ImpactFactor"    FROM "JournalRankings"    WHERE "JournalID" = j."JournalID"    ORDER BY "RankingYear" DESC NULLS LAST, "Source"    LIMIT 1) jr ON TRUE
            WHERE rp."PaperID" = %s
        ''', [paper_id])
        row = cur.fetchone()
        if not row:
            return response.Response(
                {'error': 'Paper not found'}, status=404
            )

        cby_raw = row[7]
        if isinstance(cby_raw, str):
            try:
                import json
                cby = json.loads(cby_raw)
            except Exception:
                cby = None
        else:
            cby = cby_raw

        # RawData_Log->'authorships' is the OpenAlex-shape array we store on
        # every scrape. For each entry grab the display name and every
        # institution string we can reach, then flag whether any of them is
        # Al-Baha University.
        import re as _re
        ALBAHA = _re.compile(r'(al[\s\-]?baha|albaha|الباحة)', _re.IGNORECASE)
        authorships_payload = []
        raw_authorships = row[18]
        if isinstance(raw_authorships, str):
            try:
                import json as _json
                raw_authorships = _json.loads(raw_authorships)
            except Exception:
                raw_authorships = None
        if isinstance(raw_authorships, list):
            for ship in raw_authorships:
                if not isinstance(ship, dict):
                    continue
                name = ((ship.get('author') or {}).get('display_name') or '').strip()
                insts = []
                for inst in (ship.get('institutions') or []):
                    n = (inst or {}).get('display_name')
                    if n:
                        insts.append(n)
                for raw in (ship.get('raw_affiliation_strings') or []):
                    if raw and raw not in insts:
                        insts.append(raw)
                single = ship.get('raw_affiliation_string')
                if single and single not in insts:
                    insts.append(single)
                at_albaha = any(ALBAHA.search(n) for n in insts)
                authorships_payload.append({
                    'name':         name,
                    'institutions': insts,
                    'at_albaha':    at_albaha,
                })

        paper = {
            'paper_id':        row[0],
            'title':           row[1],
            'abstract':        row[2],
            'doi':             row[3],
            'pub_year':        row[4],
            'source':          row[5],
            'indexing':        row[6],
            'citations_by_year': cby,
            'raw_authors':     row[8],
            'publication':     row[9],
            'link':            row[10],
            'cites_id':        row[11],
            'total_citations': row[12],
            'journal_name':    row[13],
            'issn_print':      row[14],
            'venue_type':      row[15],
            'quartile':        row[16],
            'impact_factor':   row[17],
            # Empty until backfill_authorships runs or a new scrape fills it;
            # the frontend falls back to raw_authors when this list is empty.
            'authorships':     authorships_payload,
        }

        # Al-Baha researchers attributed to this paper
        cur.execute('''
            SELECT u."UserID", u."FullName_Ar", d."DepartmentName"
            FROM "Authors" a
            JOIN "Users" u ON u."UserID" = a."UserID"
            LEFT JOIN "Works_In" w ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
            LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
            WHERE a."PaperID" = %s
            ORDER BY a."AuthorOrder" NULLS LAST
        ''', [paper_id])
        paper['albaha_authors'] = [
            {'user_id': r[0], 'full_name_ar': r[1], 'department_name': r[2]}
            for r in cur.fetchall()
        ]

    return response.Response(paper)


def _excel_response(filename: str):
    """Helper: build an HttpResponse with the right xlsx headers."""
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@decorators.api_view(['GET'])
def export_excel(request):
    """GET /api/export/excel/?years=2025,2026&sheets=summary,departments,researchers,journals,conferences

    Builds an xlsx workbook from the years + sheets the dashboard's options
    modal picks (defaults to all of both). With everything selected you get a
    Summary, Departments, Journals, and Conferences sheet per year, plus a
    single Researchers sheet of contributions in the window.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.db import connection

    # The export carries per-researcher rows (names, Scholar/ORCID IDs) and full
    # paper lists, so limit it to roles that may view researchers. Otherwise a
    # plain Researcher would fall into the unscoped branch below and pull the
    # whole institution.
    _u = getattr(request, 'user', None)
    if not (_u and _u.is_authenticated and (
            _u.has_litrix_perm('view_all_researchers') or
            _u.has_litrix_perm('view_dept_researchers'))):
        return response.Response({'error': 'Forbidden'}, status=403)

    # A HoD may only export their own department, so every sheet below is
    # filtered to hod_dept_id. Admin/Dean get the full institution (hod_dept_id
    # is None, so the `scoped` clauses are skipped). The helper returns the
    # sentinel -1 for a HoD with no department at all.
    hod_dept_id = _hod_scope_department_id(request)
    if hod_dept_id == -1:
        return response.Response(
            {'error': 'You are not assigned to any department. '
                      'Contact an admin to set your department.'},
            status=403,
        )
    scoped = hod_dept_id is not None

    # Department label for the Overview sheet, so a HoD's file makes its
    # scope obvious.
    hod_dept_name = None
    if scoped:
        with connection.cursor() as _c:
            _c.execute(
                'SELECT "DepartmentName" FROM "Department" WHERE "DepartmentID" = %s',
                [hod_dept_id],
            )
            _row = _c.fetchone()
            hod_dept_name = _row[0] if _row else None

    # EXISTS predicate: the paper has at least one author whose current
    # position is in the HoD's department. Appended to paper queries when
    # scoped; rp_alias is the ResearchPaper alias in the target query.
    def _dept_paper_clause(rp_alias: str) -> str:
        return (
            f' AND EXISTS (SELECT 1 FROM "Authors" a_dept '
            f'             JOIN "Works_In" w_dept ON w_dept."UserID" = a_dept."UserID" '
            f'                                   AND w_dept."IsCurrentPosition" = TRUE '
            f'             WHERE a_dept."PaperID" = {rp_alias}."PaperID" '
            f'               AND w_dept."DepartmentID" = %s)'
        )

    years_param = request.query_params.get('years', '').strip()
    if years_param:
        try:
            years = [int(y) for y in years_param.split(',') if y.strip().isdigit()]
        except ValueError:
            years = list(FOCUS_YEARS)
    else:
        years = list(FOCUS_YEARS)
    if not years:
        years = list(FOCUS_YEARS)

    # In "Al-Baha only" mode (?affiliation=albaha) every paper/citation query
    # below drops papers confirmed authored elsewhere, so the workbook matches
    # the screen.
    albaha_only = _albaha_only(request)
    AFFIL     = _affil_clause(albaha_only, 'rp')
    AFFIL_RP2 = _affil_clause(albaha_only, 'rp2')
    AFFIL_RPW = _affil_clause(albaha_only, 'rp_w')

    sheets_param = request.query_params.get('sheets', '').strip()
    if sheets_param:
        sheets = {s.strip().lower() for s in sheets_param.split(',') if s.strip()}
    else:
        sheets = {'summary', 'departments', 'researchers', 'journals', 'conferences'}

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1D1D1F')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # KPI overview as the first sheet so it opens by default; mirrors the
    # dashboard cards (Researchers / Publications / Citations / h-index).
    if 'summary' in sheets or 'departments' in sheets or 'researchers' in sheets:
        ws_overview = wb.create_sheet('Overview', 0)
        big_value_font = Font(bold=True, size=28, color='1D1D1F')
        label_font = Font(bold=True, size=11, color='86868B')
        sublabel_font = Font(size=10, color='86868B')
        title_font = Font(bold=True, size=18, color='1D1D1F')
        bg_fill = PatternFill('solid', fgColor='FAFAFA')

        ws_overview['A1'] = 'Litrix — Research Analytics Overview'
        ws_overview['A1'].font = title_font
        ws_overview.row_dimensions[1].height = 36
        years_label = ', '.join(str(y) for y in sorted(years))
        scope_label = f' · Department: {hod_dept_name}' if hod_dept_name else ''
        ws_overview['A2'] = f'Window: {years_label}{scope_label}'
        ws_overview['A2'].font = sublabel_font

        # Same per-year semantics as the dashboard. Each KPI is its own small
        # query so the HoD filter drops in cleanly without juggling param order.
        year_strs = [str(y) for y in years]
        with connection.cursor() as cur:
            # Researchers + active (scoped to current dept positions).
            if scoped:
                cur.execute('''
                    SELECT COUNT(DISTINCT u."UserID"),
                           COUNT(DISTINCT u."UserID") FILTER (WHERE r."LastSyncedAt" IS NOT NULL)
                    FROM "Users" u
                    JOIN "Works_In" w ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
                    LEFT JOIN "Researcher" r ON r."UserID" = u."UserID"
                    WHERE u."UserType" = 'Researcher' AND w."DepartmentID" = %s
                ''', [hod_dept_id])
            else:
                cur.execute('''
                    SELECT COUNT(DISTINCT u."UserID"),
                           COUNT(DISTINCT u."UserID") FILTER (WHERE r."LastSyncedAt" IS NOT NULL)
                    FROM "Users" u
                    LEFT JOIN "Researcher" r ON r."UserID" = u."UserID"
                    WHERE u."UserType" = 'Researcher'
                ''')
            r_total, r_active = cur.fetchone()

            # Papers in window + Q1 (author-in-dept when scoped).
            papers_sql = (
                'SELECT COUNT(*), COUNT(*) FILTER (WHERE q = \'Q1\') FROM ('
                '  SELECT DISTINCT rp."PaperID", jr."Quartile" AS q'
                '  FROM "ResearchPaper" rp'
                '  LEFT JOIN LATERAL (SELECT "Quartile" FROM "JournalRankings"'
                '      WHERE "JournalID" = rp."JournalID"'
                '      ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1) jr ON TRUE'
                '  WHERE rp."PubYear" = ANY(%s)'
                '    AND EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID")'
            )
            papers_params = [years]
            if scoped:
                papers_sql += _dept_paper_clause('rp')
                papers_params.append(hod_dept_id)
            papers_sql += AFFIL
            papers_sql += ') sub'
            cur.execute(papers_sql, papers_params)
            p_total, p_q1 = cur.fetchone()

            # Citations received in the selected years, per-paper
            # (ResearchPaper.CitationsByYear) — same definition as the overview
            # dashboard so the two agree. EXISTS, not JOIN, so each paper counts
            # once across its co-authors.
            year_keys_expr = _cites_expr('rp', years)
            cit_sql = (
                f'SELECT COALESCE(SUM({year_keys_expr}), 0) FROM "ResearchPaper" rp '
                'WHERE EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID"'
            )
            cit_params = list(year_strs)
            if scoped:
                cit_sql += (
                    ' AND EXISTS (SELECT 1 FROM "Works_In" w '
                    '             WHERE w."UserID" = a."UserID" '
                    '               AND w."IsCurrentPosition" = TRUE '
                    '               AND w."DepartmentID" = %s)'
                )
                cit_params.append(hod_dept_id)
            cit_sql += ')'
            cit_sql += AFFIL
            cur.execute(cit_sql, cit_params)
            c_total = cur.fetchone()[0]

            # Average h-index, scoped to dept researchers.
            if scoped:
                cur.execute('''
                    SELECT ROUND(AVG(hi.h_index)::numeric, 1)
                    FROM v_researcher_h_index hi
                    JOIN "Works_In" w ON w."UserID" = hi."UserID"
                                     AND w."IsCurrentPosition" = TRUE
                    WHERE w."DepartmentID" = %s
                ''', [hod_dept_id])
            else:
                cur.execute('SELECT ROUND(AVG(h_index)::numeric, 1) FROM v_researcher_h_index')
            avg_h = cur.fetchone()[0]

        # Four cards in a row, each spanning two columns (A-B, C-D, E-F, G-H).
        cards = [
            ('Researchers',  str(r_total),                f'{r_active} active'),
            ('Publications', f'{p_total:,}',              f'{p_q1} in Q1 journals'),
            ('Citations',    f'{c_total:,}',              f'received in {years_label}'),
            ('h-index',      str(avg_h or 0),             'avg h-index'),
        ]
        for idx, (label, value, sub) in enumerate(cards):
            col_label = chr(ord('A') + idx * 2)
            col_value = chr(ord('A') + idx * 2)  # same col, multiple rows

            cell_label = ws_overview.cell(row=4, column=idx * 2 + 1, value=label.upper())
            cell_label.font = label_font
            cell_label.fill = bg_fill

            cell_val = ws_overview.cell(row=5, column=idx * 2 + 1, value=value)
            cell_val.font = big_value_font
            cell_val.fill = bg_fill
            ws_overview.row_dimensions[5].height = 44

            cell_sub = ws_overview.cell(row=6, column=idx * 2 + 1, value=sub)
            cell_sub.font = sublabel_font
            cell_sub.fill = bg_fill

        # Widen the KPI columns
        for col in ['A', 'C', 'E', 'G']:
            ws_overview.column_dimensions[col].width = 22
        for col in ['B', 'D', 'F', 'H']:
            ws_overview.column_dimensions[col].width = 4  # spacer

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
        ws.row_dimensions[1].height = 26

    def set_widths(ws, widths):
        for col_idx, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = w

    def wrap_column(ws, col_idx: int):
        """Wrap text in every body cell of a column — used for the Abstract
        column so long paragraphs don't overflow into the next cell."""
        from openpyxl.styles import Alignment
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical='top',
            )

    if 'summary' in sheets:
        for year in sorted(years):
            ws = wb.create_sheet(f'Summary {year}')
            ws.append(['Metric', 'Value'])
            style_header(ws, 2)
            with connection.cursor() as cur:
                # Paper counts per year (published-in-year semantics)
                summary_sql = (
                    'SELECT '
                    '    COUNT(DISTINCT rp."PaperID"), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q1\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q2\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q3\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q4\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE j."VenueType" = \'Journal\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE j."VenueType" = \'Conference\') '
                    'FROM "ResearchPaper" rp '
                    'LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID" '
                    'LEFT JOIN LATERAL (SELECT "Quartile", "ImpactFactor" FROM "JournalRankings" '
                    '  WHERE "JournalID" = rp."JournalID" ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1) jr ON TRUE '
                    'WHERE rp."PubYear" = %s '
                    '  AND EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID")'
                )
                summary_params = [year]
                if scoped:
                    summary_sql += _dept_paper_clause('rp')
                    summary_params.append(hod_dept_id)
                summary_sql += AFFIL
                cur.execute(summary_sql, summary_params)
                p, q1, q2, q3, q4, jp, cp = cur.fetchone()

                # Citations received this year, per-paper — same definition as
                # the overview dashboard; EXISTS counts each paper once.
                cit_sql = (
                    f'SELECT COALESCE(SUM({_cites_expr("rp", [year])}), 0) '
                    'FROM "ResearchPaper" rp '
                    'WHERE EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID"'
                )
                cit_params = [str(year)]
                if scoped:
                    cit_sql += (
                        ' AND EXISTS (SELECT 1 FROM "Works_In" w '
                        '             WHERE w."UserID" = a."UserID" '
                        '               AND w."IsCurrentPosition" = TRUE '
                        '               AND w."DepartmentID" = %s)'
                    )
                    cit_params.append(hod_dept_id)
                cit_sql += ')'
                cit_sql += AFFIL
                cur.execute(cit_sql, cit_params)
                c = cur.fetchone()[0]
            for label, val in [
                ('Year', year),
                ('Total Papers', p),
                ('Total Citations', c),
                ('Journal Papers', jp),
                ('Conference Papers', cp),
                ('Q1 Papers', q1),
                ('Q2 Papers', q2),
                ('Q3 Papers', q3),
                ('Q4 Papers', q4),
            ]:
                ws.append([label, val])
            set_widths(ws, [25, 20])

    if 'departments' in sheets:
        for year in sorted(years):
            ws = wb.create_sheet(f'Departments {year}')
            ws.append([
                'Department', 'Researchers',
                'Journal Papers', 'Conference Papers', 'Total Papers',
                'Citations', 'Q1', 'Q2', 'Q3', 'Q4'
            ])
            style_header(ws, 10)
            with connection.cursor() as cur:
                # Citations received this year, per-paper, same definition as
                # the overview dashboard. DISTINCT on (dept, paper) keeps a paper
                # co-authored by two members of the same department from being
                # double-counted for that department.
                dept_sql = (
                    'WITH dept_cites AS ('
                    '    SELECT dept AS "DepartmentID", SUM(cites) AS cites FROM ('
                    '        SELECT DISTINCT w."DepartmentID" AS dept, rp2."PaperID" AS pid, '
                    f'               {_cites_expr("rp2", [year])} AS cites '
                    '        FROM "Works_In" w '
                    '        JOIN "Authors" a2 ON a2."UserID" = w."UserID" '
                    '        JOIN "ResearchPaper" rp2 ON rp2."PaperID" = a2."PaperID" '
                    '        WHERE w."IsCurrentPosition" = TRUE' + AFFIL_RP2 +
                    '    ) ded GROUP BY dept'
                    ') '
                    'SELECT '
                    '    d."DepartmentName", '
                    '    COUNT(DISTINCT u."UserID"), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE j."VenueType" = \'Journal\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE j."VenueType" = \'Conference\'), '
                    '    COUNT(DISTINCT rp."PaperID"), '
                    '    COALESCE(MAX(dc.cites), 0) AS citations, '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q1\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q2\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q3\'), '
                    '    COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q4\') '
                    'FROM "Department" d '
                    'LEFT JOIN "Works_In" w ON w."DepartmentID" = d."DepartmentID" AND w."IsCurrentPosition" = TRUE '
                    'LEFT JOIN "Users" u ON u."UserID" = w."UserID" AND u."UserType" = \'Researcher\' '
                    'LEFT JOIN "Authors" a ON a."UserID" = u."UserID" '
                    'LEFT JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID" AND rp."PubYear" = %s' + AFFIL + ' '
                    'LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID" '
                    'LEFT JOIN LATERAL (SELECT "Quartile", "ImpactFactor" FROM "JournalRankings" '
                    '  WHERE "JournalID" = rp."JournalID" ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1) jr ON TRUE '
                    'LEFT JOIN dept_cites dc ON dc."DepartmentID" = d."DepartmentID" '
                )
                dept_params = [str(year), year]
                if scoped:
                    dept_sql += 'WHERE d."DepartmentID" = %s '
                    dept_params.append(hod_dept_id)
                dept_sql += 'GROUP BY d."DepartmentName" ORDER BY 5 DESC'
                cur.execute(dept_sql, dept_params)
                for row in cur.fetchall():
                    ws.append(list(row))
            set_widths(ws, [32, 12, 14, 16, 12, 12, 8, 8, 8, 8])

    if 'researchers' in sheets:
        ws = wb.create_sheet('Researchers')
        ws.append([
            'Researcher (AR)', 'Department', 'Rank',
            'Papers (window)', 'Citations (window)',
            'Papers (all-time)', 'Citations (all-time)',
            'h-index (all-time)', 'Status',
            'Scholar ID', 'ORCID'
        ])
        style_header(ws, 11)
        with connection.cursor() as cur:
            researchers_sql = '''
                SELECT
                    u."FullName_Ar",
                    d."DepartmentName",
                    r."AcademicRank",
                    -- Window stats. Papers = published in the focus years.
                    -- Citations = RECEIVED in the focus years across ALL the
                    -- researcher's papers (per-paper ResearchPaper.CitationsByYear),
                    -- matching the overview dashboard's definition.
                    COUNT(DISTINCT a_w."PaperID") FILTER (WHERE rp_w."PubYear" = ANY(%(years)s)) AS papers_window,
                    COALESCE(SUM({cite_window}), 0) AS citations_window,
                    -- All-time stats. Citations here stay the lifetime snapshot
                    -- (every citation the paper ever earned) — a distinct,
                    -- explicitly-labelled "all-time" column.
                    COUNT(DISTINCT a_w."PaperID") AS papers_all,
                    COALESCE(SUM(COALESCE((rp_w."RawData_Log"->'cited_by'->>'value')::int, 0)), 0) AS citations_all,
                    COALESCE(hi.h_index, 0) AS h_index,
                    -- Status with clear priority:
                    --   1. Has papers in the focus window  → Active
                    --   2. Has all-time papers (manual OR scraped) → Historical
                    --   3. Has a public profile but no papers yet → Pending Sync
                    --   4. Nothing at all → No Profile (manual outreach needed)
                    CASE
                        WHEN COUNT(DISTINCT a_w."PaperID")
                                FILTER (WHERE rp_w."PubYear" = ANY(%(years)s)) > 0
                            THEN 'Active'
                        WHEN COUNT(DISTINCT a_w."PaperID") > 0
                            THEN 'Historical'
                        WHEN u."Scholar_ID" IS NOT NULL
                          OR r."ORCID_ID" IS NOT NULL
                          OR r."OpenAlex_AuthorID" IS NOT NULL
                          OR r."Scopus_ID" IS NOT NULL
                            THEN 'Pending Sync'
                        ELSE 'No Profile'
                    END AS sync_status,
                    u."Scholar_ID",
                    r."ORCID_ID"
                FROM "Users" u
                JOIN "Researcher" r ON r."UserID" = u."UserID"
                LEFT JOIN "Works_In" w ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
                LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
                LEFT JOIN "Authors" a_w ON a_w."UserID" = u."UserID"
                LEFT JOIN "ResearchPaper" rp_w ON rp_w."PaperID" = a_w."PaperID"{affil_rpw}
                LEFT JOIN v_researcher_h_index hi ON hi."UserID" = u."UserID"
                WHERE u."UserType" = 'Researcher'
                {dept_filter}
                GROUP BY u."UserID", u."FullName_Ar", d."DepartmentName",
                         r."AcademicRank", hi.h_index, u."Scholar_ID",
                         r."ORCID_ID", r."OpenAlex_AuthorID", r."Scopus_ID",
                         r."LastSyncedAt"
                ORDER BY papers_window DESC, h_index DESC
            '''
            # Named params so the per-year keys coexist with the dict-style
            # %(years)s / %(dept)s params.
            cite_window_expr = ' + '.join([
                f'COALESCE((rp_w."CitationsByYear"->>%(y{i})s)::int, 0)'
                for i in range(len(years))
            ]) or '0'
            researchers_params = {'years': years}
            researchers_params.update({f'y{i}': str(y) for i, y in enumerate(years)})
            if scoped:
                researchers_sql = researchers_sql.format(
                    cite_window=cite_window_expr, affil_rpw=AFFIL_RPW,
                    dept_filter='AND w."DepartmentID" = %(dept)s')
                researchers_params['dept'] = hod_dept_id
            else:
                researchers_sql = researchers_sql.format(
                    cite_window=cite_window_expr, affil_rpw=AFFIL_RPW, dept_filter='')
            cur.execute(researchers_sql, researchers_params)
            for row in cur.fetchall():
                ws.append(list(row))
        set_widths(ws, [32, 22, 18, 12, 14, 14, 16, 14, 14, 18, 22])

    if 'journals' in sheets:
        for year in sorted(years):
            ws = wb.create_sheet(f'Journals {year}')
            # Two author columns: Al-Baha researchers only (Arabic), then all
            # authors combined (Arabic + foreign, no affiliations).
            ws.append([
                'Department', 'Title', 'Abstract',
                'Al-Baha Researchers', 'All Authors (raw)',
                'Journal', 'Quartile', 'IF', 'Indexing',
                'Citations', 'DOI'
            ])
            style_header(ws, 11)
            with connection.cursor() as cur:
                # v_paper_details has no paper_id, so to attach Abstract we
                # match ResearchPaper by DOI when available and fall back to a
                # case-insensitive title match when either side lacks a DOI.
                journals_sql = (
                    'SELECT '
                    '    v.department_name, v.title, '
                    '    rp."Abstract" AS abstract, '
                    '    v.authors_ar, v.all_authors_en, '
                    '    v.journal_name, v.quartile, v.impact_factor, v.indexing, '
                    '    v.citations, v.doi '
                    'FROM v_paper_details v '
                    'LEFT JOIN "ResearchPaper" rp ON '
                    '    (v.doi IS NOT NULL AND LOWER(rp."DOI") = LOWER(v.doi)) '
                    '    OR ((v.doi IS NULL OR rp."DOI" IS NULL) AND LOWER(rp."Title") = LOWER(v.title)) '
                    'WHERE v.pub_year = %s AND v.venue_type = \'Journal\''
                )
                journals_params = [year]
                if scoped:
                    journals_sql += ' AND v.department_id = %s'
                    journals_params.append(hod_dept_id)
                journals_sql += AFFIL
                journals_sql += ' ORDER BY v.department_name, v.citations DESC NULLS LAST'
                cur.execute(journals_sql, journals_params)
                for row in cur.fetchall():
                    ws.append(list(row))
            set_widths(ws, [22, 60, 80, 50, 50, 30, 10, 8, 12, 10, 30])
            wrap_column(ws, 3)  # Abstract column

    if 'conferences' in sheets:
        for year in sorted(years):
            ws = wb.create_sheet(f'Conferences {year}')
            ws.append([
                'Department', 'Title', 'Abstract',
                'Al-Baha Researchers', 'All Authors (raw)',
                'Conference', 'Indexing', 'Citations', 'DOI'
            ])
            style_header(ws, 9)
            with connection.cursor() as cur:
                conf_sql = (
                    'SELECT '
                    '    v.department_name, v.title, '
                    '    rp."Abstract" AS abstract, '
                    '    v.authors_ar, v.all_authors_en, '
                    '    v.journal_name, v.indexing, v.citations, v.doi '
                    'FROM v_paper_details v '
                    'LEFT JOIN "ResearchPaper" rp ON '
                    '    (v.doi IS NOT NULL AND LOWER(rp."DOI") = LOWER(v.doi)) '
                    '    OR ((v.doi IS NULL OR rp."DOI" IS NULL) AND LOWER(rp."Title") = LOWER(v.title)) '
                    'WHERE v.pub_year = %s AND v.venue_type = \'Conference\''
                )
                conf_params = [year]
                if scoped:
                    conf_sql += ' AND v.department_id = %s'
                    conf_params.append(hod_dept_id)
                conf_sql += AFFIL
                conf_sql += ' ORDER BY v.department_name, v.citations DESC NULLS LAST'
                cur.execute(conf_sql, conf_params)
                for row in cur.fetchall():
                    ws.append(list(row))
            set_widths(ws, [22, 60, 80, 50, 50, 30, 12, 10, 30])
            wrap_column(ws, 3)  # Abstract column

    if not wb.sheetnames:
        ws = wb.create_sheet('Empty')
        ws.append(['No sheets selected'])

    fname = f"litrix_export_{'-'.join(map(str, sorted(years)))}.xlsx"
    resp = _excel_response(fname)
    wb.save(resp)
    return resp


class ResearcherViewSet(viewsets.ReadOnlyModelViewSet):
    """
    GET /api/researchers/             → list of all researchers
    GET /api/researchers/?department_id=2  → filtered
    GET /api/researchers/?search=محمد  → fuzzy search
    GET /api/researchers/?ordering=-h_index  → sort
    GET /api/researchers/{user_id}/   → single researcher detail
    GET /api/researchers/{user_id}/papers/ → researcher's papers
    """
    queryset = ResearcherStats.objects.all()
    serializer_class = ResearcherStatsSerializer
    filter_backends = [
        DjangoFilterBackend,
        filters.OrderingFilter,
        filters.SearchFilter,
    ]
    filterset_fields = ['department_id', 'academic_rank']
    search_fields = ['full_name_ar', 'full_name_en', 'scholar_id', 'orcid_id']
    ordering_fields = [
        'total_papers', 'total_citations', 'h_index',
        'q1_papers', 'last_pub_year',
    ]
    ordering = ['-h_index', '-total_papers']

    @decorators.action(detail=False, methods=['get'])
    def all(self, request):
        """
        GET /api/researchers/all/
        Returns ALL researchers as a flat list — for the sidebar listing.
        Lightweight: name + department + paper count only.
        """
        from django.db import connection
        with connection.cursor() as cur:
            cur.execute('''
                SELECT
                    u."UserID",
                    u."Litrix_ID",
                    u."FullName_Ar",
                    TRIM(CONCAT_WS(' ', u."FirstName", u."LastName")) AS full_name_en,
                    d."DepartmentName",
                    (SELECT COUNT(*) FROM "Authors" a WHERE a."UserID" = u."UserID") AS papers
                FROM "Users" u
                LEFT JOIN "Works_In" w ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
                LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
                WHERE u."UserType" = 'Researcher'
                ORDER BY papers DESC NULLS LAST, u."FullName_Ar"
            ''')
            results = [
                {
                    'user_id': r[0],
                    'litrix_id': r[1],
                    'full_name_ar': r[2],
                    'full_name_en': r[3],
                    'department_name': r[4],
                    'papers': r[5],
                }
                for r in cur.fetchall()
            ]
        return response.Response({'results': results})

    @decorators.action(detail=False, methods=['get'])
    def search(self, request):
        """
        GET /api/researchers/search/?q=محمد
        Lightweight search — hits Users table directly. No aggregation.
        Returns up to 15 results in <100ms.
        """
        q = (request.query_params.get('q') or '').strip()
        if not q or len(q) < 2:
            return response.Response({'results': []})

        from django.db import connection
        with connection.cursor() as cur:
            like = f'%{q}%'
            cur.execute('''
                SELECT
                    u."UserID",
                    u."Litrix_ID",
                    u."FullName_Ar",
                    TRIM(CONCAT_WS(' ', u."FirstName", u."LastName")) AS full_name_en,
                    d."DepartmentName",
                    u."Scholar_ID"
                FROM "Users" u
                LEFT JOIN "Works_In" w ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
                LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
                WHERE u."UserType" = 'Researcher'
                  AND (
                       u."FullName_Ar" ILIKE %s
                    OR u."FirstName" ILIKE %s
                    OR u."LastName" ILIKE %s
                  )
                ORDER BY u."FullName_Ar"
                LIMIT 15
            ''', [like, like, like])
            results = [
                {
                    'user_id': r[0],
                    'litrix_id': r[1],
                    'full_name_ar': r[2],
                    'full_name_en': r[3],
                    'department_name': r[4],
                    'scholar_id': r[5],
                }
                for r in cur.fetchall()
            ]
        return response.Response({'results': results})

    @decorators.action(detail=True, methods=['get'])
    def profile(self, request, pk=None):
        """GET /api/researchers/{id}/profile/

        `id` accepts either a numeric UserID or the public Lit-NNNNNN; we
        resolve to the integer PK at the boundary so the SQL keeps using the
        indexed key. Both are accepted for backward compat with tooling that
        stored numeric IDs (the frontend now uses Litrix_ID everywhere).

        One shot for the whole profile page — identity, aggregated stats,
        chart-ready per-year citations, and the papers list. The page renders
        all of these from the same data, so batching avoids three round-trips
        and a flash of empty-then-filled UI.
        """
        from django.db import connection

        # Litrix-ID gets normalized then looked up; a pure number is treated as
        # the UserID directly. Malformed input returns 400 rather than leaking a
        # 500 from a Postgres type error.
        canonical = normalize_litrix_id(pk)
        if canonical is not None:
            # Match on the numeric core so "LIT-0001", "Lit-000001" and "lit-1"
            # all resolve to the same user regardless of stored zero-padding.
            seq = int(LITRIX_ID_PATTERN.match(pk.strip()).group(1))
            with connection.cursor() as cur:
                cur.execute(
                    '''
                    SELECT "UserID" FROM "Users"
                    WHERE "Litrix_ID" IS NOT NULL
                      AND "Litrix_ID" ~* '^lit-[0-9]+$'
                      AND CAST(SUBSTRING("Litrix_ID" FROM 5) AS INTEGER) = %s
                    LIMIT 1
                    ''',
                    [seq],
                )
                row = cur.fetchone()
                if not row:
                    return response.Response(
                        {'error': f'Researcher {canonical} not found'},
                        status=404,
                    )
                resolved_user_id = row[0]
        else:
            try:
                resolved_user_id = int(pk)
            except (TypeError, ValueError):
                return response.Response(
                    {'error': 'Invalid researcher ID'}, status=400
                )

        with connection.cursor() as cur:
            cur.execute('''
                SELECT
                    u."UserID", u."FullName_Ar",
                    u."FirstName", u."LastName", u."Email",
                    u."Scholar_ID", u."ORCID",
                    r."OpenAlex_AuthorID", r."LastSyncedAt",
                    d."DepartmentID", d."DepartmentName",
                    u."Litrix_ID"
                FROM "Users" u
                LEFT JOIN "Researcher" r ON r."UserID" = u."UserID"
                LEFT JOIN "Works_In" w ON w."UserID" = u."UserID"
                                       AND w."IsCurrentPosition" = TRUE
                LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
                WHERE u."UserID" = %s
            ''', [resolved_user_id])
            row = cur.fetchone()
            if not row:
                return response.Response(
                    {'error': 'Researcher not found'}, status=404
                )
            identity = {
                'user_id':           row[0],
                'full_name_ar':      row[1],
                'first_name':        row[2],
                'last_name':         row[3],
                'email':             row[4],
                'scholar_id':        row[5],
                'orcid':             row[6],
                'openalex_author_id': row[7],
                'last_synced_at':    row[8],
                'department_id':     row[9],
                'department_name':   row[10],
                'litrix_id':         row[11],
            }

            # For citations, sum the per-paper Scholar cited_by.value (the most
            # accurate per-paper signal), falling back to OpenAlex's count.
            cur.execute('''
                SELECT
                    COUNT(DISTINCT rp."PaperID")                         AS total_papers,
                    COALESCE(SUM(COALESCE(
                        ("RawData_Log"->'cited_by'->>'value')::int,
                        ("RawData_Log"->>'cited_by_count')::int,
                        0)), 0)                                          AS total_citations,
                    COUNT(DISTINCT rp."PaperID")
                        FILTER (WHERE jr."Quartile" = 'Q1')              AS q1_papers,
                    COUNT(DISTINCT rp."PaperID")
                        FILTER (WHERE rp."Indexing" = 'Scopus'
                                   OR jr."Quartile" IS NOT NULL)         AS scopus_papers,
                    COUNT(DISTINCT rp."PaperID")
                        FILTER (WHERE rp."Indexing" = 'ISI')             AS isi_papers
                FROM "ResearchPaper" rp
                JOIN "Authors" a ON a."PaperID" = rp."PaperID"
                LEFT JOIN LATERAL (    SELECT "Quartile", "ImpactFactor"    FROM "JournalRankings"    WHERE "JournalID" = rp."JournalID"    ORDER BY "RankingYear" DESC NULLS LAST, "Source"    LIMIT 1) jr ON TRUE
                WHERE a."UserID" = %s
            ''', [resolved_user_id])
            stats_row = cur.fetchone()

            # When Scholar gives us an author-level per-year graph, prefer its
            # total over the per-paper aggregation above.
            try:
                cur.execute(
                    'SELECT "CitationsByYear" FROM "Researcher" WHERE "UserID" = %s',
                    [resolved_user_id],
                )
                rcby = cur.fetchone()
                if rcby and rcby[0]:
                    raw = rcby[0]
                    if isinstance(raw, str):
                        import json as _json
                        try:
                            raw = _json.loads(raw)
                        except Exception:
                            raw = {}
                    scholar_total = sum(
                        int(v) for v in (raw or {}).values()
                        if str(v).isdigit() or isinstance(v, int)
                    )
                    if scholar_total > 0:
                        # Override with Scholar's authoritative total
                        stats_row = (
                            stats_row[0],
                            scholar_total,
                            stats_row[2],
                            stats_row[3],
                            stats_row[4],
                        )
            except Exception:
                pass
            stats = {
                'total_papers':    stats_row[0],
                'total_citations': stats_row[1],
                'q1_papers':       stats_row[2],
                'scopus_papers':   stats_row[3],
                'isi_papers':      stats_row[4],
            }

            # Prefer Researcher.CitationsByYear — the author-level graph straight
            # from Scholar's cited_by.graph, which is what Scholar shows on the
            # profile. Fall back to summing the per-paper CitationsByYear.
            citations_by_year = []
            try:
                cur.execute('''
                    SELECT "CitationsByYear" FROM "Researcher"
                    WHERE "UserID" = %s
                ''', [resolved_user_id])
                rcby = cur.fetchone()
                if rcby and rcby[0]:
                    raw = rcby[0]
                    if isinstance(raw, str):
                        import json as _json
                        try:
                            raw = _json.loads(raw)
                        except Exception:
                            raw = {}
                    citations_by_year = sorted(
                        ({'year': int(y), 'citations': int(v)}
                         for y, v in (raw or {}).items()
                         if str(y).isdigit()),
                        key=lambda x: x['year'],
                    )
                else:
                    cur.execute('''
                        SELECT
                            yr.year::int AS year,
                            SUM(CASE WHEN yr.cnt ~ '^[0-9]+$'
                                     THEN yr.cnt::int
                                     ELSE 0 END) AS citations
                        FROM "ResearchPaper" rp
                        JOIN "Authors" a ON a."PaperID" = rp."PaperID"
                        CROSS JOIN LATERAL jsonb_each_text(
                            COALESCE(rp."CitationsByYear", '{}'::jsonb)
                        ) AS yr(year, cnt)
                        WHERE a."UserID" = %s
                          AND yr.year ~ '^[0-9]+$'
                        GROUP BY yr.year
                        ORDER BY yr.year
                    ''', [resolved_user_id])
                    citations_by_year = [
                        {'year': r[0], 'citations': r[1]}
                        for r in cur.fetchall()
                    ]

                # Clip to CHART_YEAR_FLOOR to match the admin chart window —
                # earlier years just stretch the X-axis and bury recent growth.
                citations_by_year = [
                    pt for pt in citations_by_year
                    if pt['year'] >= CHART_YEAR_FLOOR
                ]
            except Exception:
                citations_by_year = []

            # Papers with full metadata. When there's no linked JournalID, the
            # journal name falls back to Scholar's free-text "publication" string
            # (e.g. "Sustainability 14 (2), 829, 2022").
            cur.execute('''
                SELECT
                    rp."PaperID", rp."Title", rp."DOI", rp."PubYear",
                    rp."Source", rp."Indexing", rp."CitationsByYear",
                    COALESCE(
                        ("RawData_Log"->'cited_by'->>'value')::int,
                        ("RawData_Log"->>'cited_by_count')::int,
                        0
                    )                       AS citations,
                    COALESCE(j."JournalName", rp."RawData_Log"->>'publication')
                                            AS journal_name,
                    j."ISSN_Print",
                    j."VenueType",
                    jr."Quartile",
                    jr."ImpactFactor",
                    rp."AffiliationVerified"
                FROM "ResearchPaper" rp
                JOIN "Authors" a ON a."PaperID" = rp."PaperID"
                LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
                LEFT JOIN LATERAL (    SELECT "Quartile", "ImpactFactor"    FROM "JournalRankings"    WHERE "JournalID" = rp."JournalID"    ORDER BY "RankingYear" DESC NULLS LAST, "Source"    LIMIT 1) jr ON TRUE
                WHERE a."UserID" = %s
                ORDER BY rp."PubYear" DESC NULLS LAST, rp."PaperID" DESC
            ''', [resolved_user_id])
            paper_cols = [c[0].lower() for c in cur.description]
            papers = [dict(zip([
                'paper_id', 'title', 'doi', 'pub_year', 'source',
                'indexing', 'citations_by_year', 'citations',
                'journal_name', 'issn_print', 'venue_type',
                'quartile', 'impact_factor', 'affiliation_verified',
            ], r)) for r in cur.fetchall()]

        return response.Response({
            'identity':          identity,
            'stats':             stats,
            'citations_by_year': citations_by_year,
            'papers':            papers,
        })

    @decorators.action(detail=True, methods=['get'])
    def papers(self, request, pk=None):
        """
        GET /api/researchers/{user_id}/papers/
        Returns ALL papers authored by this researcher (newest first).
        """
        from django.db import connection
        with connection.cursor() as cur:
            cur.execute('''
                SELECT
                    rp."PaperID", rp."Title", rp."Title_En", rp."Abstract",
                    rp."Language", rp."DOI", rp."PubYear",
                    rp."Volume", rp."Issue", rp."Pages",
                    rp."Source", rp."IsVerified", rp."ScrapedAt"
                FROM "ResearchPaper" rp
                JOIN "Authors" a ON a."PaperID" = rp."PaperID"
                WHERE a."UserID" = %s
                ORDER BY rp."PubYear" DESC NULLS LAST, rp."PaperID" DESC
            ''', [pk])
            rows = cur.fetchall()
            cols = [c[0].lower() for c in cur.description]

        field_map = {
            'paperid': 'paper_id', 'title': 'title', 'title_en': 'title_en',
            'abstract': 'abstract', 'language': 'language', 'doi': 'doi',
            'pubyear': 'pub_year', 'volume': 'volume', 'issue': 'issue',
            'pages': 'pages', 'source': 'source',
            'isverified': 'is_verified', 'scrapedat': 'scraped_at',
        }
        data = [
            {field_map.get(c, c): row[i] for i, c in enumerate(cols)}
            for row in rows
        ]
        return response.Response(data)


class DepartmentViewSet(viewsets.ReadOnlyModelViewSet):
    """
    GET /api/departments/         → list with aggregated stats
    GET /api/departments/{id}/    → single department detail

    HoDs are scoped to their own department; Admins/Deans see all.

    Paper + citation figures are recomputed this-period (per-paper) before
    returning, so this page agrees with the overview dashboard rather than
    showing the all-time/lifetime numbers stored in v_department_stats.
    """
    serializer_class = DepartmentStatsSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = [
        'total_papers', 'total_citations', 'total_q1_papers',
        'avg_h_index', 'total_researchers',
    ]
    ordering = ['-total_papers']

    def get_queryset(self):
        qs = DepartmentStats.objects.all()
        scope = _hod_scope_department_id(self.request)
        if scope is not None:
            qs = qs.filter(department_id=scope)
        return qs

    def list(self, request, *args, **kwargs):
        resp = super().list(request, *args, **kwargs)
        data = resp.data
        rows = data.get('results') if isinstance(data, dict) else data
        if rows:
            albaha = _albaha_only(request)
            win = _dept_cards_windowed(list(FOCUS_YEARS), albaha)
            for r in rows:
                w = win.get(r.get('department_id'))
                if w:
                    r.update(w)
        return resp

    @decorators.action(detail=True, methods=['get'])
    def researchers(self, request, pk=None):
        """GET /api/departments/{id}/researchers/ — list of researchers."""
        # A HoD may only inspect their own department's researchers.
        scope = _hod_scope_department_id(request)
        if scope is not None and str(scope) != str(pk):
            return response.Response(
                {'error': 'You can only view your own department.'},
                status=403,
            )
        qs = ResearcherStats.objects.filter(department_id=pk).order_by(
            '-h_index', '-total_papers'
        )
        page = self.paginate_queryset(qs)
        rows = [dict(r) for r in ResearcherStatsSerializer(page or qs, many=True).data]
        win = _researcher_rows_windowed(
            list(FOCUS_YEARS), [r['user_id'] for r in rows], _albaha_only(request))
        for r in rows:
            w = win.get(r['user_id'])
            r.update(w if w else {'total_papers': 0, 'q1_papers': 0, 'total_citations': 0})
        return self.get_paginated_response(rows) if page else response.Response(rows)


class TopPaperViewSet(viewsets.ReadOnlyModelViewSet):
    """
    GET /api/papers/top/?limit=10   → most-cited papers
    GET /api/papers/top/?quartile=Q1
    GET /api/papers/top/?pub_year=2024
    """
    queryset = TopPaper.objects.all()
    serializer_class = TopPaperSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['quartile', 'pub_year', 'source']
    ordering_fields = ['citations', 'pub_year', 'impact_factor']
    ordering = ['-citations']


class PublicationTrendViewSet(viewsets.ReadOnlyModelViewSet):
    """
    GET /api/trends/                       → all departments × all years
    GET /api/trends/?department_id=2       → single department's trend
    """
    queryset = PublicationTrend.objects.all()
    serializer_class = PublicationTrendSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['department_id', 'year']
    ordering = ['department_name', 'year']

    def get_queryset(self):
        # A HoD only sees their own department's trend; Admin/Dean see all
        # (consistent with the overview + departments scoping).
        qs = PublicationTrend.objects.all()
        scope = _hod_scope_department_id(self.request)
        if scope is not None:
            qs = qs.filter(department_id=scope)
        return qs


@decorators.api_view(['GET'])
def yearly_breakdown(request):
    """GET /api/yearly-breakdown/?year=2025

    Department-level breakdown for one year: per-department journal/conference
    counts and citations, plus a flat paper list the frontend splits by
    venue_type into expandable lists under each department card.
    """
    year = request.query_params.get('year')
    if not year:
        return response.Response(
            {'error': 'year parameter required'}, status=400
        )
    try:
        year_int = int(year)
    except (TypeError, ValueError):
        return response.Response(
            {'error': 'year must be an integer'}, status=400
        )

    # This drill-down belongs to the institutional overview, so it gets the same
    # gate + HoD scoping: a plain Researcher can't pull it and a HoD sees only
    # their own department (it was unscoped before — every dept leaked).
    _u = getattr(request, 'user', None)
    if not (_u and _u.is_authenticated and (
            _u.has_litrix_perm('view_all_researchers') or
            _u.has_litrix_perm('view_dept_researchers'))):
        return response.Response({'error': 'Forbidden'}, status=403)
    hod_dept_id = _hod_scope_department_id(request)
    dept_clause = ' AND department_id = %s' if hod_dept_id else ''
    dept_param = [hod_dept_id] if hod_dept_id else []

    from django.db import connection
    with connection.cursor() as cur:
        cur.execute(f'''
            SELECT
                department_id,
                department_name,
                COUNT(*) FILTER (WHERE venue_type = 'Journal')    AS journal_papers,
                COUNT(*) FILTER (WHERE venue_type = 'Conference') AS conference_papers,
                COUNT(*)                                          AS total_papers,
                COALESCE(SUM(citations), 0)                       AS total_citations
            FROM v_paper_details
            WHERE pub_year = %s
              AND department_id IS NOT NULL{dept_clause}
            GROUP BY department_id, department_name
            ORDER BY total_papers DESC
        ''', [year_int] + dept_param)
        dept_rows = cur.fetchall()
        departments = [
            {
                'department_id':     r[0],
                'department_name':   r[1],
                'journal_papers':    r[2],
                'conference_papers': r[3],
                'total_papers':      r[4],
                'total_citations':   r[5],
            }
            for r in dept_rows
        ]

        cur.execute(f'''
            SELECT
                paper_id, title, doi, citations, journal_name,
                venue_type, quartile, impact_factor, indexing,
                department_id, department_name, authors_ar
            FROM v_paper_details
            WHERE pub_year = %s
              AND department_id IS NOT NULL{dept_clause}
            ORDER BY citations DESC NULLS LAST, paper_id
        ''', [year_int] + dept_param)
        cols = [c[0] for c in cur.description]
        papers = [dict(zip(cols, row)) for row in cur.fetchall()]

    return response.Response({
        'year': year_int,
        'departments': departments,
        'papers': papers,
    })


@decorators.api_view(['GET'])
def overview(request):
    """GET /api/stats/overview/ (optionally ?year=YYYY)

    One-shot payload for the Admin/Dean/HoD landing page; HoDs are auto-scoped
    to their own department.

    Pass ?affiliation=albaha to restrict every paper-derived metric to papers
    not confirmed authored elsewhere (AffiliationVerified IS DISTINCT FROM FALSE
    keeps confirmed-Al-Baha TRUE and not-yet-verified NULL, drops only
    confirmed-elsewhere FALSE). The default leaves the historical numbers
    untouched. Citation totals come from Scholar's author-level CitationsByYear
    graph and can't be paper-filtered, so the toggle only affects paper counts.
    """
    years = _resolve_years(request)

    albaha_only = _albaha_only(request)
    AFFIL_CLAUSE = _affil_clause(albaha_only, 'rp')
    # Same predicate for the citation CTEs that alias the paper table as rp_all.
    AFFIL_CLAUSE_RPALL = _affil_clause(albaha_only, 'rp_all')

    # Institutional dashboard data is for roles that may view researchers; a
    # plain Researcher would otherwise fall into the unscoped branch and get the
    # full institution view (the UI already routes them to their own dashboard).
    _u = getattr(request, 'user', None)
    if not (_u and _u.is_authenticated and (
            _u.has_litrix_perm('view_all_researchers') or
            _u.has_litrix_perm('view_dept_researchers'))):
        return response.Response({'error': 'Forbidden'}, status=403)

    # Resolve via the shared helper so HoD detection matches the rest of the
    # app. Returns None for Admin/Dean (full institution), a dept id for a HoD,
    # or -1 when a HoD has no department — the -1 sentinel scopes every query
    # below to nothing instead of leaking the whole institution.
    hod_dept_id = _hod_scope_department_id(request)

    # avg_h keeps the historical definition (average of dept averages).
    _dept_qs = DepartmentStats.objects.all()
    if hod_dept_id:
        _dept_qs = _dept_qs.filter(department_id=hod_dept_id)
    dept_agg = _dept_qs.aggregate(avg_h=Avg('avg_h_index'))

    # COUNT(DISTINCT UserID) over current positions. The old
    # Sum('total_researchers') across DepartmentStats rows double-counted anyone
    # holding positions in two departments (one row each).
    from django.db import connection as _hc_conn
    with _hc_conn.cursor() as _hc_cur:
        _hc_cur.execute('''
            SELECT COUNT(DISTINCT u."UserID") AS researchers,
                   COUNT(DISTINCT u."UserID")
                       FILTER (WHERE r."LastSyncedAt" IS NOT NULL) AS active
            FROM "Users" u
            JOIN "Works_In" w ON w."UserID" = u."UserID"
                             AND w."IsCurrentPosition" = TRUE
            LEFT JOIN "Researcher" r ON r."UserID" = u."UserID"
            WHERE u."UserType" = 'Researcher'
              AND (%s::int IS NULL OR w."DepartmentID" = %s::int)
        ''', [hod_dept_id, hod_dept_id])
        _hc_row = _hc_cur.fetchone()
    dept_agg['researchers'] = _hc_row[0]
    dept_agg['active'] = _hc_row[1]

    from django.db import connection
    with connection.cursor() as cur:
        # Papers KPI plus Q1/Scopus/ISI counts — papers published in window.
        # The author-in-dept clause is appended for HoDs only.
        kpi_sql = (
            'SELECT COUNT(DISTINCT rp."PaperID") AS papers, '
            '       COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = \'Q1\') AS q1, '
            '       COUNT(DISTINCT rp."PaperID") FILTER (WHERE rp."Indexing" = \'Scopus\' OR jr."Quartile" IS NOT NULL) AS scopus, '
            '       COUNT(DISTINCT rp."PaperID") FILTER (WHERE rp."Indexing" = \'ISI\') AS isi '
            'FROM "ResearchPaper" rp '
            'LEFT JOIN LATERAL (SELECT "Quartile","ImpactFactor" FROM "JournalRankings" '
            '  WHERE "JournalID" = rp."JournalID" ORDER BY "RankingYear" DESC NULLS LAST, "Source" LIMIT 1) jr ON TRUE '
            'WHERE rp."PubYear" = ANY(%s) '
            '  AND EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID"'
        )
        kpi_params = [years]
        if hod_dept_id:
            kpi_sql += (
                ' AND EXISTS (SELECT 1 FROM "Works_In" w2 '
                '              WHERE w2."UserID" = a."UserID" '
                '                AND w2."DepartmentID" = %s '
                '                AND w2."IsCurrentPosition" = TRUE)'
            )
            kpi_params.append(hod_dept_id)
        kpi_sql += ')'
        kpi_sql += AFFIL_CLAUSE
        cur.execute(kpi_sql, kpi_params)
        paper_count_row = cur.fetchone()

        # Citations received in the window, summed per-paper over the selected
        # years across ALL the scope's papers (any publication year) — not
        # citations of papers published in the window. This keeps the metric
        # meaningful for a single recent year: older highly-cited papers still
        # contribute what they earned that year, whereas papers published in
        # e.g. 2026 have barely been cited yet. EXISTS, not JOIN, counts each
        # paper once even when several scope authors share it.
        cites_year_expr = _cites_expr('rp', years)
        cit_sql = (
            f'SELECT COALESCE(SUM({cites_year_expr}), 0) AS citations '
            'FROM "ResearchPaper" rp '
            'WHERE EXISTS (SELECT 1 FROM "Authors" a WHERE a."PaperID" = rp."PaperID"'
        )
        cit_params = [str(y) for y in years]
        if hod_dept_id:
            cit_sql += (
                ' AND EXISTS (SELECT 1 FROM "Works_In" w2 '
                '              WHERE w2."UserID" = a."UserID" '
                '                AND w2."DepartmentID" = %s '
                '                AND w2."IsCurrentPosition" = TRUE)'
            )
            cit_params.append(hod_dept_id)
        cit_sql += ')'
        cit_sql += AFFIL_CLAUSE
        cur.execute(cit_sql, cit_params)
        citations_row = cur.fetchone()

        # paper_count_row is (papers, q1, scopus, isi); splice the citation total
        # in to get the (papers, citations, q1, scopus, isi) shape.
        paper_totals = (
            paper_count_row[0],
            citations_row[0],
            paper_count_row[1],
            paper_count_row[2],
            paper_count_row[3],
        )

        # Top researchers: papers PUBLISHED in window + citations
        # RECEIVED in window (per-year sum across ALL their papers).
        year_keys_expr_alias = _cites_expr('rp_all', years)
        cur.execute(f'''
            WITH papers_in_window AS (
                SELECT a."UserID",
                       COUNT(DISTINCT rp."PaperID") AS focus_papers
                FROM "Authors" a
                JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
                WHERE rp."PubYear" = ANY(%s){AFFIL_CLAUSE}
                GROUP BY a."UserID"
            ),
            citations_in_window AS (
                SELECT a."UserID",
                       COALESCE(SUM({year_keys_expr_alias}), 0) AS focus_citations
                FROM "Authors" a
                JOIN "ResearchPaper" rp_all ON rp_all."PaperID" = a."PaperID"
                WHERE 1=1{AFFIL_CLAUSE_RPALL}
                GROUP BY a."UserID"
            )
            SELECT
                u."UserID",
                u."FullName_Ar",
                d."DepartmentName",
                COALESCE(p.focus_papers, 0) AS focus_papers,
                COALESCE(c.focus_citations, 0) AS focus_citations
            FROM "Users" u
            LEFT JOIN papers_in_window    p ON p."UserID" = u."UserID"
            LEFT JOIN citations_in_window c ON c."UserID" = u."UserID"
            LEFT JOIN "Works_In"   w ON w."UserID" = u."UserID" AND w."IsCurrentPosition" = TRUE
            LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
            WHERE u."UserType" = 'Researcher'
              AND COALESCE(p.focus_papers, 0) > 0
              AND (%s::int IS NULL OR w."DepartmentID" = %s::int)
            ORDER BY focus_papers DESC, focus_citations DESC
            LIMIT 5
        ''', [years] + [str(y) for y in years] + [hod_dept_id, hod_dept_id])
        top_researchers_rows = cur.fetchall()

        # For a HoD, restrict to papers authored by someone currently in their
        # department, else this leaks the institution's most-cited papers onto a
        # department dashboard.
        top_papers_sql = 'SELECT * FROM v_top_papers WHERE pub_year = ANY(%s)'
        top_papers_params = [years]
        if hod_dept_id is not None:
            top_papers_sql += (
                ' AND EXISTS (SELECT 1 FROM "Authors" a '
                '             JOIN "Works_In" w_tp ON w_tp."UserID" = a."UserID" '
                '                                 AND w_tp."IsCurrentPosition" = TRUE '
                '             WHERE a."PaperID" = v_top_papers.paper_id '
                '               AND w_tp."DepartmentID" = %s)'
            )
            top_papers_params.append(hod_dept_id)
        if albaha_only:
            top_papers_sql += (
                ' AND (SELECT rp_af."AffiliationVerified" FROM "ResearchPaper" rp_af '
                'WHERE rp_af."PaperID" = v_top_papers.paper_id) IS DISTINCT FROM FALSE'
            )
        top_papers_sql += ' ORDER BY citations DESC NULLS LAST LIMIT 5'
        cur.execute(top_papers_sql, top_papers_params)
        top_paper_cols = [c[0] for c in cur.description]
        top_papers = [dict(zip(top_paper_cols, row)) for row in cur.fetchall()]

        # Papers published in window + citations received in window. The
        # citation aggregation lives in its own CTE to avoid the cartesian blowup
        # of joining authors→papers→citations; each dept-researcher contributes
        # once.
        year_keys_dept = _cites_expr('rp_all', years)
        cur.execute(f'''
            WITH dept_citations AS (
                SELECT dept AS "DepartmentID", SUM(cites) AS total_citations
                FROM (
                    SELECT DISTINCT
                        w."DepartmentID"     AS dept,
                        rp_all."PaperID"     AS pid,
                        ({year_keys_dept})   AS cites
                    FROM "Works_In" w
                    JOIN "Authors" a ON a."UserID" = w."UserID"
                    JOIN "ResearchPaper" rp_all ON rp_all."PaperID" = a."PaperID"
                    WHERE w."IsCurrentPosition" = TRUE{AFFIL_CLAUSE_RPALL}
                ) ded
                GROUP BY dept
            )
            SELECT
                d."DepartmentID"   AS department_id,
                d."DepartmentName" AS department_name,
                d."CollegeID"      AS college_id,
                COUNT(DISTINCT u."UserID") AS total_researchers,
                COUNT(DISTINCT u."UserID") FILTER (WHERE r."LastSyncedAt" IS NOT NULL) AS active_researchers,
                COUNT(DISTINCT rp."PaperID") AS total_papers,
                COALESCE(MAX(dc.total_citations), 0) AS total_citations,
                COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = 'Q1') AS total_q1_papers,
                COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = 'Q2') AS total_q2_papers,
                COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = 'Q3') AS total_q3_papers,
                COUNT(DISTINCT rp."PaperID") FILTER (WHERE jr."Quartile" = 'Q4') AS total_q4_papers,
                COUNT(DISTINCT rp."PaperID")
                    FILTER (WHERE rp."Indexing" = 'Scopus' OR jr."Quartile" IS NOT NULL) AS total_scopus_papers,
                COUNT(DISTINCT rp."PaperID") FILTER (WHERE rp."Indexing" = 'ISI') AS total_isi_papers,
                -- Venue split: 'Conference Proceedings' folds into Conference;
                -- Book/Book Series/NULL fold into Journal (same rule as
                -- v_department_stats after 20260607_dept_stats_split).
                COUNT(DISTINCT rp."PaperID")
                    FILTER (WHERE j."VenueType" ILIKE 'Conference%%') AS conference_papers,
                COUNT(DISTINCT rp."PaperID")
                    FILTER (WHERE rp."PaperID" IS NOT NULL
                            AND (j."VenueType" IS NULL
                                 OR j."VenueType" NOT ILIKE 'Conference%%')) AS journal_papers
            FROM "Department" d
            LEFT JOIN "Works_In" w ON w."DepartmentID" = d."DepartmentID" AND w."IsCurrentPosition" = TRUE
            LEFT JOIN "Users" u ON u."UserID" = w."UserID" AND u."UserType" = 'Researcher'
            LEFT JOIN "Researcher" r ON r."UserID" = u."UserID"
            LEFT JOIN "Authors" a ON a."UserID" = u."UserID"
            LEFT JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID" AND rp."PubYear" = ANY(%s){AFFIL_CLAUSE}
            LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
            LEFT JOIN LATERAL (    SELECT "Quartile", "ImpactFactor"    FROM "JournalRankings"    WHERE "JournalID" = rp."JournalID"    ORDER BY "RankingYear" DESC NULLS LAST, "Source"    LIMIT 1) jr ON TRUE
            LEFT JOIN dept_citations dc ON dc."DepartmentID" = d."DepartmentID"
            GROUP BY d."DepartmentID", d."DepartmentName", d."CollegeID"
            HAVING (%s::int IS NULL OR d."DepartmentID" = %s::int)
            ORDER BY total_papers DESC NULLS LAST
        ''', [str(y) for y in years] + [years] + [hod_dept_id, hod_dept_id])
        dept_cols = [c[0] for c in cur.description]
        departments = [dict(zip(dept_cols, row)) for row in cur.fetchall()]

        # The KPI strip and per-department table use the full `years` floor, but
        # the trend chart clips to CHART_YEAR_FLOOR — 16 sparse points crush the
        # recent-growth signal. A narrower explicit ?year= filter is honored via
        # the intersection; an empty intersection falls back to `years` so we
        # never return an empty chart.
        chart_years = [y for y in years if y >= CHART_YEAR_FLOOR] or years

        # Papers published per (dept, year) — DISTINCT so a paper shared by
        # several co-authored faculty isn't double-counted.
        cur.execute(f'''
            SELECT w."DepartmentID", rp."PubYear",
                   COUNT(DISTINCT rp."PaperID") AS papers
            FROM "Works_In" w
            JOIN "Authors" a ON a."UserID" = w."UserID"
            JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
            WHERE w."IsCurrentPosition" = TRUE
              AND rp."PubYear" = ANY(%s)
              AND (%s::int IS NULL OR w."DepartmentID" = %s::int){AFFIL_CLAUSE}
            GROUP BY w."DepartmentID", rp."PubYear"
        ''', [chart_years, hod_dept_id, hod_dept_id])
        papers_by_dept_year = {}
        for did, yr, n in cur.fetchall():
            papers_by_dept_year.setdefault(did, {})[int(yr)] = int(n)

        # Per-year citations per department from the per-paper graph (same source
        # as the KPI, so the affiliation filter applies). DISTINCT on (dept,
        # paper, year) keeps a paper co-authored within one department from being
        # double-counted; a paper spanning two departments still counts once per
        # department, which is intended.
        cur.execute(f'''
            SELECT dept, yr, SUM(cites) AS citations
            FROM (
                SELECT DISTINCT
                    w."DepartmentID"      AS dept,
                    rp."PaperID"          AS pid,
                    year_kv.key::int      AS yr,
                    (year_kv.value)::int  AS cites
                FROM "Works_In" w
                JOIN "Authors" a ON a."UserID" = w."UserID"
                JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
                CROSS JOIN LATERAL jsonb_each_text(
                    COALESCE(rp."CitationsByYear", '{{}}'::jsonb)
                ) AS year_kv
                WHERE w."IsCurrentPosition" = TRUE
                  AND year_kv.value ~ '^[0-9]+$'
                  AND year_kv.key::int = ANY(%s)
                  AND (%s::int IS NULL OR w."DepartmentID" = %s::int){AFFIL_CLAUSE}
            ) ded
            GROUP BY dept, yr
        ''', [chart_years, hod_dept_id, hod_dept_id])
        cites_by_dept_year = {}
        for did, yr, n in cur.fetchall():
            cites_by_dept_year.setdefault(did, {})[int(yr)] = int(n)

        # Attach by_year to each department row over chart_years, so the chart
        # axis stays reasonable even though the KPIs span 2011+.
        for d in departments:
            did = d['department_id']
            d['by_year'] = [
                {
                    'year': y,
                    'papers':    papers_by_dept_year.get(did, {}).get(y, 0),
                    'citations': cites_by_dept_year.get(did, {}).get(y, 0),
                }
                for y in sorted(chart_years)
            ]

    return response.Response({
        'focus_years': years,
        # Echo the active filter back so the UI can reflect the mode.
        'affiliation_filter': 'albaha' if albaha_only else 'all',
        # Axis-year list so a chart doesn't have to re-derive the window.
        'chart_years': sorted(chart_years),
        'totals': {
            'researchers':         dept_agg['researchers'] or 0,
            'active_researchers':  dept_agg['active'] or 0,
            'papers':              paper_totals[0] or 0,
            'citations':           paper_totals[1] or 0,
            'q1_papers':           paper_totals[2] or 0,
            'scopus_papers':       paper_totals[3] or 0,
            'isi_papers':          paper_totals[4] or 0,
            'avg_h_index':         float(dept_agg['avg_h'] or 0),
        },
        'top_researchers': [
            {
                'user_id':         r[0],
                'full_name_ar':    r[1],
                'department_name': r[2],
                'total_papers':    r[3],
                'total_citations': r[4],
                'h_index':         0,
            }
            for r in top_researchers_rows
        ],
        'top_papers':      top_papers,
        'departments':     departments,
    })


# Spotlight-style global search returning { profiles: [...], papers: [...] }.
# The permission gate is the key rule: Admin/Dean/HoD see the full corpus
# (including papers by external, non-registered authors) for institutional
# oversight, while a plain Researcher only sees papers with at least one
# registered system author. Both sides cap results to keep the modal snappy.
@decorators.api_view(['GET'])
def universal_search(request):
    q = (request.query_params.get('q') or '').strip()
    if len(q) < 2:
        return response.Response({'profiles': [], 'papers': []})

    # has_litrix_perm lives on accounts.User; getattr keeps this safe when
    # SimpleJWT hands us an AnonymousUser-like object — we default to False.
    user = getattr(request, 'user', None)
    has_full_access = bool(
        user
        and user.is_authenticated
        and (
            user.has_litrix_perm('view_all_researchers')
            or user.has_litrix_perm('view_dept_researchers')
        )
    )

    from django.db import connection
    like = f'%{q}%'

    PROFILE_LIMIT = 8
    PAPER_LIMIT   = 10

    # All roles now see every UserType — the search is intentionally
    # unrestricted (it used to limit Researchers to Researcher-type profiles).
    user_type_filter = ''

    with connection.cursor() as cur:
        # Cross-script aware: we match Arabic name, English first/last, the
        # email/Litrix_ID identifiers, and Authors.AuthorNameRaw. That last one
        # is the bridge — scrapers populate it with the English-script author
        # string per paper, so an English query finds an Arabic-only profile (and
        # vice versa) without any transliteration heuristics.
        cur.execute(f'''
            SELECT
                u."UserID",
                u."Litrix_ID",
                u."FullName_Ar",
                TRIM(CONCAT_WS(' ', u."FirstName", u."LastName")) AS full_name_en,
                u."UserType",
                d."DepartmentName",
                (SELECT COUNT(*) FROM "Authors" a WHERE a."UserID" = u."UserID") AS papers
            FROM "Users" u
            -- Pick ONE current Works_In per user. If the user is HoD
            -- of a department, that row is deprioritized so the
            -- researcher-side department wins. Stable tiebreak on
            -- StartDate ASC (oldest position first).
            LEFT JOIN LATERAL (
                SELECT wx."DepartmentID", dx."DepartmentName"
                FROM "Works_In" wx
                LEFT JOIN "Department" dx ON dx."DepartmentID" = wx."DepartmentID"
                WHERE wx."UserID" = u."UserID"
                  AND wx."IsCurrentPosition" = TRUE
                ORDER BY (dx."HeadID" = u."UserID") ASC NULLS FIRST,
                         wx."StartDate" ASC NULLS LAST
                LIMIT 1
            ) w ON TRUE
            LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
            WHERE 1=1
              {user_type_filter}
              AND (
                   u."FullName_Ar" ILIKE %s
                OR u."FirstName"   ILIKE %s
                OR u."LastName"    ILIKE %s
                OR u."Email"       ILIKE %s
                OR u."Litrix_ID"   ILIKE %s
                OR EXISTS (
                    SELECT 1 FROM "Authors" a
                    WHERE a."UserID" = u."UserID"
                      AND a."AuthorNameRaw" ILIKE %s
                )
              )
            ORDER BY papers DESC NULLS LAST, u."FullName_Ar"
            LIMIT %s
        ''', [like, like, like, like, like, like, PROFILE_LIMIT])
        profiles = [
            {
                'user_id':         r[0],
                'litrix_id':       r[1],
                'full_name_ar':    r[2],
                'full_name_en':    r[3],
                'user_type':       r[4],
                'department_name': r[5],
                'papers':          r[6],
            }
            for r in cur.fetchall()
        ]

        # Title match with the permission gate baked in: the EXISTS subquery
        # enforces "at least one system author" for restricted users, while the
        # full-access path skips it.
        if has_full_access:
            paper_filter = ''
        else:
            paper_filter = '''
                AND EXISTS (
                    SELECT 1 FROM "Authors" a
                    WHERE a."PaperID" = rp."PaperID"
                      AND a."UserID" IS NOT NULL
                )
            '''


        cur.execute(f'''
            SELECT
                rp."PaperID",
                rp."Title",
                rp."Title_En",
                rp."PubYear",
                rp."DOI",
                rp."Indexing",
                COALESCE(j."JournalName", rp."RawData_Log"->>'publication') AS journal_name,
                jr."Quartile",
                COALESCE(
                    (rp."RawData_Log"->'cited_by'->>'value')::int,
                    (rp."RawData_Log"->>'cited_by_count')::int,
                    0
                ) AS citations,
                -- compact summary of the first 3 system authors, for the card
                (
                    SELECT STRING_AGG(u2."FullName_Ar", '  ·  ')
                    FROM (
                        SELECT u3."FullName_Ar" FROM "Authors" a2
                        JOIN "Users" u3 ON u3."UserID" = a2."UserID"
                        WHERE a2."PaperID" = rp."PaperID"
                          AND u3."FullName_Ar" IS NOT NULL
                        ORDER BY a2."AuthorOrder" NULLS LAST
                        LIMIT 3
                    ) u2
                ) AS authors_summary
            FROM "ResearchPaper" rp
            LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
            LEFT JOIN LATERAL (    SELECT "Quartile", "ImpactFactor"    FROM "JournalRankings"    WHERE "JournalID" = rp."JournalID"    ORDER BY "RankingYear" DESC NULLS LAST, "Source"    LIMIT 1) jr ON TRUE
            WHERE (
                   rp."Title"    ILIKE %s
                OR rp."Title_En" ILIKE %s
                OR rp."DOI"      ILIKE %s
            )
              {paper_filter}
            ORDER BY rp."PubYear" DESC NULLS LAST, citations DESC
            LIMIT %s
        ''', [like, like, like, PAPER_LIMIT])
        papers = [
            {
                'paper_id':        r[0],
                'title':           r[1],
                'title_en':        r[2],
                'pub_year':        r[3],
                'doi':
           r[4],
                'indexing':        r[5],
                'journal_name':    r[6],
                'quartile':        r[7],
                'citations':       r[8],
                'authors_summary': r[9],
            }
            for r in cur.fetchall()
        ]

    return response.Response({
        'profiles':         profiles,
        'papers':           papers,
        'has_full_access':  has_full_access,
    })
