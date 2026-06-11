"""
Admin endpoints for the Reporting Campaigns workflow.

    GET    /api/campaigns/                  — list (filtered by perm)
    POST   /api/campaigns/                  — create draft
    GET    /api/campaigns/<id>/             — detail + submission stats
    PATCH  /api/campaigns/<id>/             — edit draft only
    POST   /api/campaigns/<id>/open/        — draft → active (fan-out)
    POST   /api/campaigns/<id>/close/       — active → closed
    GET    /api/campaigns/<id>/submissions/ — list submissions

State transitions are their own action endpoints rather than a PATCH on the
status field, so each transition can run its side effects (generating
submissions, queuing notifications) atomically. Opening is the heavy one: it
expands the audience into N ReportSubmission rows and queues the
reminder/final notifications in a single transaction — fail midway, nothing
lands.
"""
import json
from datetime import timedelta

from django.db import connection, transaction
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response

from .serializers import ReportCampaignSerializer
from accounts.views import audit


# Going through connection.cursor() (not the ORM), psycopg2 sometimes hands
# back JSONB as a raw str instead of a parsed dict, which blows up the first
# time we try **scope_filter. This makes every read site safe: None / dict /
# JSON string all come back as a dict.
def _coerce_jsonb(value):
    if value is None:
        return {}
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except (json.JSONDecodeError, ValueError):
            return {}
    return {}


# Permission helpers.
def _can_manage(user):
    return user.is_authenticated and user.has_litrix_perm('manage_campaigns')


def _can_view(user):
    return user.is_authenticated and (
        user.has_litrix_perm('view_campaign_reports')
        or user.has_litrix_perm('manage_campaigns')
    )


# Expand a ScopeFilter into a list of UserIDs.
def _expand_audience(cur, tenant_id, scope_type, scope_filter):
    """
    UserIDs matching the campaign's scope. Always Researchers in the tenant —
    Deans/HoDs/Admins never get verification submissions even if listed,
    since they have no Scholar/ORCID paper list to verify.
    """
    scope_filter = _coerce_jsonb(scope_filter)

    if scope_type == 'all':
        cur.execute(
            '''SELECT "UserID" FROM "Users"
               WHERE "UserType" = 'Researcher'
                 AND "IsActive" = TRUE
                 AND "TenantID" = %s''',
            [tenant_id],
        )
    elif scope_type == 'department':
        dept_ids = scope_filter.get('department_ids') or []
        if not dept_ids:
            return []
        cur.execute(
            '''SELECT DISTINCT u."UserID" FROM "Users" u
               JOIN "Works_In" w ON w."UserID" = u."UserID"
                                AND w."IsCurrentPosition" = TRUE
               WHERE u."UserType" = 'Researcher'
                 AND u."IsActive" = TRUE
                 AND u."TenantID" = %s
                 AND w."DepartmentID" = ANY(%s)''',
            [tenant_id, dept_ids],
        )
    elif scope_type == 'custom':
        user_ids = scope_filter.get('user_ids') or []
        if not user_ids:
            return []
        cur.execute(
            '''SELECT "UserID" FROM "Users"
               WHERE "UserType" = 'Researcher'
                 AND "IsActive" = TRUE
                 AND "TenantID" = %s
                 AND "UserID" = ANY(%s)''',
            [tenant_id, user_ids],
        )
    else:
        return []

    return [r[0] for r in cur.fetchall()]


# List + create.
@api_view(['GET', 'POST'])
def campaign_list_create(request):
    if request.method == 'GET':
        if not _can_view(request.user):
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        return _list_campaigns(request)

    if not _can_manage(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
    return _create_campaign(request)


def _list_campaigns(request):
    """
    Campaigns visible to the caller, with inline submission stats so the
    frontend can draw progress bars without a round-trip per campaign.
    """
    status_filter = (request.query_params.get('status') or '').strip()
    where = ['c."TenantID" = %s']
    params = [request.user.tenant_id]
    if status_filter:
        where.append('c."Status" = %s')
        params.append(status_filter)

    with connection.cursor() as cur:
        cur.execute(f'''
            SELECT
                c."CampaignID", c."Title", c."Description",
                c."TargetYears",
                c."OpensAt", c."ClosesAt",
                c."Status", c."ScopeType", c."ScopeFilter",
                c."CreatedByUserID", c."CreatedAt",
                c."ClosedAt", c."ArchivedAt",
                COALESCE(s.total, 0)     AS submissions_total,
                COALESCE(s.submitted, 0) AS submissions_submitted,
                COALESCE(s.late, 0)      AS submissions_late
            FROM "ReportCampaign" c
            LEFT JOIN (
                SELECT
                    "CampaignID",
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE "Status" = 'submitted') AS submitted,
                    COUNT(*) FILTER (WHERE "IsLate" = TRUE) AS late
                FROM "ReportSubmission"
                GROUP BY "CampaignID"
            ) s ON s."CampaignID" = c."CampaignID"
            WHERE {' AND '.join(where)}
            ORDER BY c."CreatedAt" DESC
        ''', params)
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]

    # Map the PascalCase columns to snake_case inline (cheaper than the ORM).
    out = []
    for r in rows:
        out.append({
            'campaign_id':           r['CampaignID'],
            'title':                 r['Title'],
            'description':           r['Description'],
            'target_years':          r['TargetYears'],
            'opens_at':              r['OpensAt'],
            'closes_at':             r['ClosesAt'],
            'status':                r['Status'],
            'scope_type':            r['ScopeType'],
            # Always a dict (or null) — never a raw JSON string.
            'scope_filter':          _coerce_jsonb(r['ScopeFilter']),
            'created_by_user_id':    r['CreatedByUserID'],
            'created_at':            r['CreatedAt'],
            'closed_at':             r['ClosedAt'],
            'archived_at':           r['ArchivedAt'],
            'submissions_total':     r['submissions_total'],
            'submissions_submitted': r['submissions_submitted'],
            'submissions_late':      r['submissions_late'],
        })

    return Response({'campaigns': out})


def _create_campaign(request):
    s = ReportCampaignSerializer(data=request.data)
    s.is_valid(raise_exception=True)
    d = s.validated_data

    with connection.cursor() as cur:
        cur.execute('''
            INSERT INTO "ReportCampaign"
                ("TenantID", "Title", "Description", "TargetYears",
                 "OpensAt", "ClosesAt", "Status",
                 "ScopeType", "ScopeFilter",
                 "CreatedByUserID")
            VALUES (%s, %s, %s, %s,
                    %s, %s, 'draft',
                    %s, %s::jsonb,
                    %s)
            RETURNING "CampaignID"
        ''', [
            request.user.tenant_id,
            d['title'],
            d.get('description'),
            d['target_years'],
            d['opens_at'],
            d['closes_at'],
            d.get('scope_type', 'all'),
            json.dumps(d.get('scope_filter') or {}),
            request.user.user_id,
        ])
        campaign_id = cur.fetchone()[0]

    audit(
        request.user.user_id, request.user.tenant_id,
        'campaign.create', 'ReportCampaign', campaign_id,
        {'title': d['title'], 'target_years': d['target_years']},
        request=request,
    )

    return Response({'campaign_id': campaign_id, 'status': 'draft'},
                    status=status.HTTP_201_CREATED)


# Detail + edit.
@api_view(['GET', 'PATCH'])
def campaign_detail(request, campaign_id):
    if request.method == 'GET':
        if not _can_view(request.user):
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        return _detail(request, campaign_id)

    if not _can_manage(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
    return _edit(request, campaign_id)


def _detail(request, campaign_id):
    with connection.cursor() as cur:
        cur.execute('''
            SELECT
                c."CampaignID", c."Title", c."Description",
                c."TargetYears", c."OpensAt", c."ClosesAt",
                c."Status", c."ScopeType", c."ScopeFilter",
                c."CreatedByUserID", c."CreatedAt",
                c."ClosedAt", c."ArchivedAt",
                COALESCE(s.total, 0), COALESCE(s.submitted, 0), COALESCE(s.late, 0)
            FROM "ReportCampaign" c
            LEFT JOIN (
                SELECT "CampaignID",
                       COUNT(*) AS total,
                       COUNT(*) FILTER (WHERE "Status" = 'submitted') AS submitted,
                       COUNT(*) FILTER (WHERE "IsLate" = TRUE) AS late
                FROM "ReportSubmission"
                GROUP BY "CampaignID"
            ) s ON s."CampaignID" = c."CampaignID"
            WHERE c."CampaignID" = %s AND c."TenantID" = %s
        ''', [campaign_id, request.user.tenant_id])
        row = cur.fetchone()
        if not row:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

    return Response({
        'campaign_id':           row[0],
        'title':                 row[1],
        'description':           row[2],
        'target_years':          row[3],
        'opens_at':              row[4],
        'closes_at':             row[5],
        'status':                row[6],
        'scope_type':            row[7],
        'scope_filter':          _coerce_jsonb(row[8]),
        'created_by_user_id':    row[9],
        'created_at':            row[10],
        'closed_at':             row[11],
        'archived_at':           row[12],
        'submissions_total':     row[13],
        'submissions_submitted': row[14],
        'submissions_late':      row[15],
    })


def _edit(request, campaign_id):
    """
    Edit a draft campaign. Once active, edits are refused — submissions and
    notifications have gone out, and silently changing target_years would
    invalidate everything.
    """
    with connection.cursor() as cur:
        cur.execute(
            'SELECT "Status" FROM "ReportCampaign" '
            'WHERE "CampaignID" = %s AND "TenantID" = %s',
            [campaign_id, request.user.tenant_id],
        )
        row = cur.fetchone()
        if not row:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        if row[0] != 'draft':
            return Response(
                {'error': f'Cannot edit a campaign in status={row[0]}. '
                          'Only drafts are editable.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

    # Partial update from whatever the client sent.
    allowed = {'title', 'description', 'target_years', 'opens_at',
               'closes_at', 'scope_type', 'scope_filter'}
    payload = {k: v for k, v in request.data.items() if k in allowed}
    if not payload:
        return Response({'error': 'Nothing to update'}, status=400)

    # Re-validate through the serializer.
    s = ReportCampaignSerializer(data=payload, partial=True)
    s.is_valid(raise_exception=True)
    d = s.validated_data

    sets = []
    params = []
    column_map = {
        'title':         '"Title"',
        'description':   '"Description"',
        'target_years':  '"TargetYears"',
        'opens_at':      '"OpensAt"',
        'closes_at':     '"ClosesAt"',
        'scope_type':    '"ScopeType"',
    }
    for k, col in column_map.items():
        if k in d:
            sets.append(f'{col} = %s')
            params.append(d[k])
    if 'scope_filter' in d:
        sets.append('"ScopeFilter" = %s::jsonb')
        params.append(json.dumps(d['scope_filter'] or {}))

    if not sets:
        return Response({'error': 'Nothing to update'}, status=400)

    params.extend([campaign_id, request.user.tenant_id])
    with connection.cursor() as cur:
        cur.execute(
            f'UPDATE "ReportCampaign" SET {", ".join(sets)} '
            f'WHERE "CampaignID" = %s AND "TenantID" = %s',
            params,
        )

    audit(
        request.user.user_id, request.user.tenant_id,
        'campaign.edit', 'ReportCampaign', campaign_id,
        list(d.keys()), request=request,
    )
    return Response({'message': 'Updated'})


@api_view(['POST'])
def campaign_open(request, campaign_id):
    """
    Draft → active, all in one transaction:
        - flip status (OpensAt clamped to NOW() if it was in the past)
        - one ReportSubmission per matched researcher
        - queue three notifications: open (now), reminder (3 days before
          close), final (24h before close), each skipped if its send time
          has already passed
    """
    if not _can_manage(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    now = timezone.now()

    with transaction.atomic():
        with connection.cursor() as cur:
            cur.execute(
                '''SELECT "CampaignID", "Title", "Status",
                          "OpensAt", "ClosesAt", "ScopeType", "ScopeFilter"
                   FROM "ReportCampaign"
                   WHERE "CampaignID" = %s AND "TenantID" = %s
                   FOR UPDATE''',
                [campaign_id, request.user.tenant_id],
            )
            row = cur.fetchone()
            if not row:
                return Response({'error': 'Not found'},
                                status=status.HTTP_404_NOT_FOUND)
            cid, title, current_status, opens_at, closes_at, scope_type, scope_filter = row
            scope_filter = _coerce_jsonb(scope_filter)

            if current_status != 'draft':
                return Response(
                    {'error': f'Cannot open a campaign in status={current_status}. '
                              'Only drafts can be opened.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if closes_at <= now:
                return Response(
                    {'error': 'ClosesAt has already passed. Update it first.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Activate. Clamp OpensAt to NOW() if it was in the past
            # (admin clicked Open Now).
            effective_opens = opens_at if opens_at and opens_at > now else now
            cur.execute(
                '''UPDATE "ReportCampaign"
                   SET "Status" = 'active', "OpensAt" = %s
                   WHERE "CampaignID" = %s''',
                [effective_opens, campaign_id],
            )

            # Expand audience and bulk-insert submissions. ON CONFLICT keeps
            # this safe to re-run if an archived campaign is reopened later.
            user_ids = _expand_audience(
                cur, request.user.tenant_id, scope_type, scope_filter,
            )
            if user_ids:
                cur.executemany(
                    '''INSERT INTO "ReportSubmission" ("CampaignID", "UserID")
                       VALUES (%s, %s)
                       ON CONFLICT ("CampaignID", "UserID") DO NOTHING''',
                    [(campaign_id, uid) for uid in user_ids],
                )

            # Schedule the three notifications. The worker resolves
            # recipients via RelatedCampaignID, so this audience JSON is
            # mostly informational — we fold scope_type in for the audit trail.
            audience_with_type = {
                **scope_filter,
                'campaign_id': campaign_id,
                'scope_type':  scope_type,
            }
            target_json = json.dumps(audience_with_type)

            campaign_title = title  # already a str
            # English copy on purpose — the UI labels this "Research Reports",
            # so the sidebar, page, and inbox stay consistent.
            schedule_rows = [
                ('campaign_open',
                 f'Research report opened: {campaign_title}',
                 f'The research report "{campaign_title}" is now open. '
                 f'Please review your publications list and submit your '
                 f'report before the closing date.',
                 now),
                ('campaign_reminder',
                 f'Reminder: 3 days left to submit — {campaign_title}',
                 f'You have not submitted your report for '
                 f'"{campaign_title}" yet. The window closes in 3 days.',
                 closes_at - timedelta(days=3)),
                ('campaign_final',
                 f'Final reminder: 24 hours left — {campaign_title}',
                 f'Last chance to submit your report for '
                 f'"{campaign_title}".',
                 closes_at - timedelta(hours=24)),
            ]

            for ntype, ntitle, nbody, send_at in schedule_rows:
                # Skip reminders already past due (opened too close to close).
                if send_at <= now and ntype != 'campaign_open':
                    continue
                cur.execute(
                    '''INSERT INTO "ScheduledNotification"
                       ("TenantID", "Title", "Body", "NotificationType",
                        "TargetAudience", "SendAt",
                        "RelatedCampaignID", "CreatedByUserID")
                       VALUES (%s, %s, %s, %s, %s::jsonb, %s, %s, %s)''',
                    [
                        request.user.tenant_id, ntitle, nbody, ntype,
                        target_json, send_at,
                        campaign_id, request.user.user_id,
                    ],
                )

    audit(
        request.user.user_id, request.user.tenant_id,
        'campaign.open', 'ReportCampaign', campaign_id,
        {'submissions_created': len(user_ids)},
        request=request,
    )
    return Response({
        'message':             'Campaign opened',
        'submissions_created': len(user_ids),
        'status':              'active',
    })


@api_view(['POST'])
def campaign_close(request, campaign_id):
    """
    Active → closed. Closing doesn't lock existing submissions (IsLate is set
    at submit time) — it just blocks new ones, enforced in
    my_reports_views.submit_submission.
    """
    if not _can_manage(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    with transaction.atomic():
        with connection.cursor() as cur:
            cur.execute(
                '''SELECT "Status" FROM "ReportCampaign"
                   WHERE "CampaignID" = %s AND "TenantID" = %s
                   FOR UPDATE''',
                [campaign_id, request.user.tenant_id],
            )
            row = cur.fetchone()
            if not row:
                return Response({'error': 'Not found'},
                                status=status.HTTP_404_NOT_FOUND)
            if row[0] != 'active':
                return Response(
                    {'error': f'Cannot close a campaign in status={row[0]}. '
                              'Only active campaigns can be closed.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            cur.execute(
                '''UPDATE "ReportCampaign"
                   SET "Status" = 'closed',
                       "ClosedAt" = NOW(),
                       "ClosesAt" = NOW()
                   WHERE "CampaignID" = %s''',
                [campaign_id],
            )

    audit(
        request.user.user_id, request.user.tenant_id,
        'campaign.close', 'ReportCampaign', campaign_id,
        request=request,
    )
    return Response({'message': 'Campaign closed', 'status': 'closed'})


# Per-campaign Excel export.
@api_view(['GET'])
def campaign_export(request, campaign_id):
    """
    GET /api/campaigns/<id>/export/

    A 5-sheet workbook:
        Overview         — metadata + aggregate counts
        Researchers      — one row per submission (all statuses)
        Confirmed Papers — papers researchers claimed
        Not Mine         — papers they rejected (feeds Authors cleanup)
        Missing Papers   — papers they added that we lacked (feeds ingest)

    The Not Mine / Missing sheets are the point of running a verification
    campaign — catching misattributions and papers we missed, not just the
    green checkmarks.
    """
    if not _can_view(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    with connection.cursor() as cur:
        # Confirm the campaign exists + load metadata for the cover sheet.
        cur.execute(
            '''SELECT "CampaignID", "Title", "Description",
                      "TargetYears", "OpensAt", "ClosesAt", "Status",
                      "ScopeType", "CreatedAt"
               FROM "ReportCampaign"
               WHERE "CampaignID" = %s AND "TenantID" = %s''',
            [campaign_id, request.user.tenant_id],
        )
        camp = cur.fetchone()
        if not camp:
            return Response({'error': 'Not found'},
                            status=status.HTTP_404_NOT_FOUND)

    wb = Workbook()
    wb.remove(wb.active)

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill('solid', fgColor='1D1D1F')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_font   = Font(bold=True, size=16, color='1D1D1F')
    label_font   = Font(bold=True, color='86868B')

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
        ws.row_dimensions[1].height = 24

    def set_widths(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

    # Sheet 1: Overview.
    ws = wb.create_sheet('Overview', 0)
    ws['A1'] = camp[1]  # Title
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 28
    if camp[2]:
        ws['A2'] = camp[2]  # Description
        ws['A2'].font = label_font

    meta = [
        ('Target Years',  ', '.join(str(y) for y in camp[3] or [])),
        ('Opens At',      camp[4].strftime('%Y-%m-%d %H:%M') if camp[4] else '—'),
        ('Closes At',     camp[5].strftime('%Y-%m-%d %H:%M') if camp[5] else '—'),
        ('Status',        camp[6]),
        ('Scope',         camp[7]),
        ('Created',       camp[8].strftime('%Y-%m-%d') if camp[8] else '—'),
    ]
    for i, (k, v) in enumerate(meta, start=4):
        ws.cell(row=i, column=1, value=k).font = label_font
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40

    # Aggregate stats.
    with connection.cursor() as cur:
        cur.execute(
            '''SELECT
                 COUNT(*)                                           AS total,
                 COUNT(*) FILTER (WHERE s."Status" = 'submitted')   AS submitted,
                 COUNT(*) FILTER (WHERE s."IsLate" = TRUE)          AS late,
                 (SELECT COUNT(*) FROM "ReportPaperDecision" rd
                   JOIN "ReportSubmission" rs ON rs."SubmissionID" = rd."SubmissionID"
                   WHERE rs."CampaignID" = %s AND rd."Decision" = 'confirmed')  AS confirmed,
                 (SELECT COUNT(*) FROM "ReportPaperDecision" rd
                   JOIN "ReportSubmission" rs ON rs."SubmissionID" = rd."SubmissionID"
                   WHERE rs."CampaignID" = %s AND rd."Decision" = 'not_mine')   AS not_mine,
                 (SELECT COUNT(*) FROM "ReportPaperDecision" rd
                   JOIN "ReportSubmission" rs ON rs."SubmissionID" = rd."SubmissionID"
                   WHERE rs."CampaignID" = %s AND rd."Decision" = 'missing')    AS missing
               FROM "ReportSubmission" s
               WHERE s."CampaignID" = %s''',
            [campaign_id] * 4,
        )
        stats = cur.fetchone()

    stats_block = [
        ('Researchers in scope', stats[0]),
        ('Submitted',            stats[1]),
        ('Late submissions',     stats[2]),
        ('Confirmed papers',     stats[3]),
        ('Not-mine decisions',   stats[4]),
        ('Missing-paper reports', stats[5]),
    ]
    for i, (k, v) in enumerate(stats_block, start=12):
        ws.cell(row=i, column=1, value=k).font = label_font
        ws.cell(row=i, column=2, value=v)

    # Sheet 2: Researchers — submission status per person.
    ws = wb.create_sheet('Researchers')
    ws.append([
        'Name (AR)', 'Name (EN)', 'Email', 'Litrix ID',
        'Department', 'Status', 'Submitted At', 'Late',
        'Confirmed', 'Not Mine', 'Missing',
    ])
    style_header(ws, 11)
    with connection.cursor() as cur:
        cur.execute(
            '''SELECT
                 u."FullName_Ar",
                 TRIM(CONCAT_WS(' ', u."FirstName", u."LastName")) AS full_name_en,
                 u."Email", u."Litrix_ID",
                 d."DepartmentName",
                 s."Status",
                 s."SubmittedAt",
                 s."IsLate",
                 (SELECT COUNT(*) FROM "ReportPaperDecision" rd
                   WHERE rd."SubmissionID" = s."SubmissionID"
                     AND rd."Decision" = 'confirmed') AS c_count,
                 (SELECT COUNT(*) FROM "ReportPaperDecision" rd
                   WHERE rd."SubmissionID" = s."SubmissionID"
                     AND rd."Decision" = 'not_mine')  AS n_count,
                 (SELECT COUNT(*) FROM "ReportPaperDecision" rd
                   WHERE rd."SubmissionID" = s."SubmissionID"
                     AND rd."Decision" = 'missing')   AS m_count
               FROM "ReportSubmission" s
               JOIN "Users" u ON u."UserID" = s."UserID"
               LEFT JOIN "Works_In" w ON w."UserID" = u."UserID"
                                     AND w."IsCurrentPosition" = TRUE
               LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
               WHERE s."CampaignID" = %s
               ORDER BY
                 CASE s."Status"
                   WHEN 'submitted' THEN 0
                   WHEN 'in_progress' THEN 1
                   WHEN 'reopened' THEN 2
                   ELSE 3
                 END,
                 u."FullName_Ar"''',
            [campaign_id],
        )
        for r in cur.fetchall():
            ws.append([
                r[0], r[1], r[2], r[3], r[4], r[5],
                r[6].strftime('%Y-%m-%d %H:%M') if r[6] else '',
                'Yes' if r[7] else '',
                r[8], r[9], r[10],
            ])
    set_widths(ws, [25, 25, 30, 12, 22, 14, 18, 6, 10, 10, 10])

    # Sheet 3: Confirmed Papers.
    ws = wb.create_sheet('Confirmed Papers')
    ws.append([
        'Researcher (AR)', 'Department', 'Paper Title', 'Year',
        'Journal', 'Quartile', 'Indexing', 'Citations', 'DOI',
    ])
    style_header(ws, 9)
    with connection.cursor() as cur:
        cur.execute(
            '''SELECT
                 u."FullName_Ar",
                 d."DepartmentName",
                 rp."Title",
                 rp."PubYear",
                 COALESCE(j."JournalName", rp."RawData_Log"->>'publication') AS journal,
                 jr."Quartile",
                 rp."Indexing",
                 COALESCE(
                     ("RawData_Log"->'cited_by'->>'value')::int,
                     ("RawData_Log"->>'cited_by_count')::int,
                     0
                 ) AS citations,
                 rp."DOI"
               FROM "ReportPaperDecision" rd
               JOIN "ReportSubmission" rs ON rs."SubmissionID" = rd."SubmissionID"
               JOIN "Users" u             ON u."UserID"        = rs."UserID"
               JOIN "ResearchPaper" rp    ON rp."PaperID"      = rd."PaperID"
               LEFT JOIN "Works_In" w ON w."UserID" = u."UserID"
                                     AND w."IsCurrentPosition" = TRUE
               LEFT JOIN "Department"      d ON d."DepartmentID" = w."DepartmentID"
               LEFT JOIN "Journals"        j ON j."JournalID"    = rp."JournalID"
               LEFT JOIN "JournalRankings" jr ON jr."JournalID"  = rp."JournalID"
               WHERE rs."CampaignID" = %s AND rd."Decision" = 'confirmed'
               ORDER BY u."FullName_Ar", rp."PubYear" DESC NULLS LAST''',
            [campaign_id],
        )
        for r in cur.fetchall():
            ws.append(list(r))
    set_widths(ws, [25, 22, 60, 8, 28, 10, 10, 10, 30])

    # Sheet 4: Not Mine — admin cleanup queue.
    ws = wb.create_sheet('Not Mine')
    ws.append([
        'Researcher (AR)', 'Department', 'Paper Title',
        'Year', 'Journal', 'DOI', 'Researcher Note',
    ])
    style_header(ws, 7)
    with connection.cursor() as cur:
        cur.execute(
            '''SELECT
                 u."FullName_Ar",
                 d."DepartmentName",
                 rp."Title",
                 rp."PubYear",
                 COALESCE(j."JournalName", rp."RawData_Log"->>'publication'),
                 rp."DOI",
                 rd."Note"
               FROM "ReportPaperDecision" rd
               JOIN "ReportSubmission" rs ON rs."SubmissionID" = rd."SubmissionID"
               JOIN "Users" u             ON u."UserID"        = rs."UserID"
               JOIN "ResearchPaper" rp    ON rp."PaperID"      = rd."PaperID"
               LEFT JOIN "Works_In" w ON w."UserID" = u."UserID"
                                     AND w."IsCurrentPosition" = TRUE
               LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
               LEFT JOIN "Journals"   j ON j."JournalID"    = rp."JournalID"
               WHERE rs."CampaignID" = %s AND rd."Decision" = 'not_mine'
               ORDER BY u."FullName_Ar", rp."PubYear" DESC NULLS LAST''',
            [campaign_id],
        )
        for r in cur.fetchall():
            ws.append(list(r))
    set_widths(ws, [25, 22, 60, 8, 28, 30, 40])

    # Sheet 5: Missing Papers — researcher-reported additions.
    ws = wb.create_sheet('Missing Papers')
    ws.append([
        'Researcher (AR)', 'Department', 'Title',
        'Year', 'DOI', 'Researcher Note', 'Resolved?',
    ])
    style_header(ws, 7)
    with connection.cursor() as cur:
        cur.execute(
            '''SELECT
                 u."FullName_Ar",
                 d."DepartmentName",
                 rd."MissingTitle",
                 rd."MissingYear",
                 rd."MissingDOI",
                 rd."Note",
                 CASE WHEN rd."MissingResolvedAt" IS NULL THEN 'Pending'
                      ELSE 'Resolved' END
               FROM "ReportPaperDecision" rd
               JOIN "ReportSubmission" rs ON rs."SubmissionID" = rd."SubmissionID"
               JOIN "Users" u             ON u."UserID"        = rs."UserID"
               LEFT JOIN "Works_In" w ON w."UserID" = u."UserID"
                                     AND w."IsCurrentPosition" = TRUE
               LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
               WHERE rs."CampaignID" = %s AND rd."Decision" = 'missing'
               ORDER BY u."FullName_Ar", rd."MissingYear" DESC NULLS LAST''',
            [campaign_id],
        )
        for r in cur.fetchall():
            ws.append(list(r))
    set_widths(ws, [25, 22, 60, 8, 30, 40, 12])

    # Stream the workbook back.
    safe_title = ''.join(
        c if (c.isalnum() or c in ' -_') else '_'
        for c in (camp[1] or 'campaign')
    ).strip().replace(' ', '_')[:60]
    filename = f'campaign_{campaign_id}_{safe_title}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    audit(
        request.user.user_id, request.user.tenant_id,
        'campaign.export', 'ReportCampaign', campaign_id,
        request=request,
    )
    return response


@api_view(['GET'])
def campaign_submission_detail(request, campaign_id, submission_id):
    """
    GET /api/campaigns/<campaign_id>/submissions/<submission_id>/

    Admin-side mirror of /api/my-reports/<id>/ — same payload shape so the
    frontend can reuse its rendering. A separate surface (rather than
    widening my-reports) because that endpoint hard-filters every query to
    the caller's own UserID; an admin shouldn't have to act as the researcher
    to read a report. Perms are verified here at the boundary.
    """
    if not _can_view(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    with connection.cursor() as cur:
        cur.execute(
            '''SELECT s."SubmissionID", s."CampaignID", s."UserID", s."Status",
                      s."StartedAt", s."SubmittedAt", s."IsLate",
                      c."Title", c."TargetYears", c."OpensAt", c."ClosesAt",
                      c."Status" AS campaign_status,
                      u."FullName_Ar", u."Email", u."Litrix_ID",
                      d."DepartmentName"
               FROM "ReportSubmission" s
               JOIN "ReportCampaign"   c ON c."CampaignID" = s."CampaignID"
               JOIN "Users"            u ON u."UserID"     = s."UserID"
               LEFT JOIN "Works_In" w ON w."UserID" = u."UserID"
                                     AND w."IsCurrentPosition" = TRUE
               LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
               WHERE s."SubmissionID" = %s
                 AND s."CampaignID"   = %s
                 AND c."TenantID"     = %s''',
            [submission_id, campaign_id, request.user.tenant_id],
        )
        row = cur.fetchone()
        if not row:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        (_, _, user_id, sub_status, started_at, submitted_at, is_late,
         c_title, target_years, c_opens, c_closes, c_status,
         name_ar, email, litrix_id, dept_name) = row

        # Paper list (same shape as the my-reports endpoint).
        cur.execute(
            '''SELECT
                 rp."PaperID", rp."Title", rp."Title_En", rp."DOI",
                 rp."PubYear", rp."Source", rp."Indexing",
                 COALESCE(j."JournalName", rp."RawData_Log"->>'publication')
                     AS journal_name,
                 jr."Quartile",
                 COALESCE(
                     ("RawData_Log"->'cited_by'->>'value')::int,
                     ("RawData_Log"->>'cited_by_count')::int,
                     0
                 ) AS citations,
                 d."DecisionID", d."Decision", d."Note", d."DecidedAt"
               FROM "Authors" a
               JOIN "ResearchPaper" rp ON rp."PaperID" = a."PaperID"
               LEFT JOIN "Journals" j ON j."JournalID" = rp."JournalID"
               LEFT JOIN "JournalRankings" jr ON jr."JournalID" = rp."JournalID"
               LEFT JOIN "ReportPaperDecision" d
                 ON d."SubmissionID" = %s AND d."PaperID" = rp."PaperID"
               WHERE a."UserID" = %s
                 AND rp."PubYear" = ANY(%s)
               ORDER BY rp."PubYear" DESC NULLS LAST, rp."Title"''',
            [submission_id, user_id, target_years],
        )
        papers = []
        for r in cur.fetchall():
            papers.append({
                'paper_id':     r[0],
                'title':        r[1],
                'title_en':     r[2],
                'doi':          r[3],
                'pub_year':     r[4],
                'source':       r[5],
                'indexing':     r[6],
                'journal_name': r[7],
                'quartile':     r[8],
                'citations':    r[9],
                'decision': {
                    'decision_id': r[10],
                    'decision':    r[11],
                    'note':        r[12],
                    'decided_at':  r[13],
                } if r[10] else None,
            })

        # Missing-paper entries.
        cur.execute(
            '''SELECT "DecisionID", "MissingTitle", "MissingDOI",
                      "MissingYear", "Note", "DecidedAt",
                      "MissingResolvedAt", "MissingResolvedToPaperID"
               FROM "ReportPaperDecision"
               WHERE "SubmissionID" = %s AND "Decision" = 'missing'
               ORDER BY "DecidedAt" DESC''',
            [submission_id],
        )
        missing = [
            {
                'decision_id':  r[0],
                'title':        r[1],
                'doi':          r[2],
                'year':         r[3],
                'note':         r[4],
                'decided_at':   r[5],
                'resolved_at':  r[6],
                'resolved_to_paper_id': r[7],
            }
            for r in cur.fetchall()
        ]

    return Response({
        'submission': {
            'submission_id': submission_id,
            'user_id':       user_id,
            'status':        sub_status,
            'started_at':    started_at,
            'submitted_at':  submitted_at,
            'is_late':       is_late,
        },
        'researcher': {
            'full_name_ar':    name_ar,
            'email':           email,
            'litrix_id':       litrix_id,
            'department_name': dept_name,
        },
        'campaign': {
            'campaign_id':   campaign_id,
            'title':         c_title,
            'target_years':  target_years,
            'opens_at':      c_opens,
            'closes_at':     c_closes,
            'status':        c_status,
        },
        'papers':  papers,
        'missing': missing,
    })


@api_view(['GET'])
def campaign_submissions(request, campaign_id):
    """
    GET /api/campaigns/<id>/submissions/
    Returns one row per (campaign × researcher) for admin oversight.
    """
    if not _can_view(request.user):
        return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

    with connection.cursor() as cur:
        # Confirm the campaign exists + tenant matches.
        cur.execute(
            'SELECT 1 FROM "ReportCampaign" '
            'WHERE "CampaignID" = %s AND "TenantID" = %s',
            [campaign_id, request.user.tenant_id],
        )
        if not cur.fetchone():
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        # Join Users + Department for a list-ready payload.
        cur.execute('''
            SELECT
                s."SubmissionID", s."UserID", s."Status",
                s."StartedAt", s."SubmittedAt", s."IsLate",
                u."FullName_Ar", u."Email", u."Litrix_ID",
                d."DepartmentName",
                (SELECT COUNT(*) FROM "ReportPaperDecision" rd
                  WHERE rd."SubmissionID" = s."SubmissionID"
                    AND rd."Decision" = 'confirmed')   AS confirmed_count,
                (SELECT COUNT(*) FROM "ReportPaperDecision" rd
                  WHERE rd."SubmissionID" = s."SubmissionID"
                    AND rd."Decision" = 'not_mine')    AS not_mine_count,
                (SELECT COUNT(*) FROM "ReportPaperDecision" rd
                  WHERE rd."SubmissionID" = s."SubmissionID"
                    AND rd."Decision" = 'missing')     AS missing_count
            FROM "ReportSubmission" s
            JOIN "Users" u ON u."UserID" = s."UserID"
            LEFT JOIN "Works_In" w ON w."UserID" = u."UserID"
                                  AND w."IsCurrentPosition" = TRUE
            LEFT JOIN "Department" d ON d."DepartmentID" = w."DepartmentID"
            WHERE s."CampaignID" = %s
            ORDER BY
                CASE s."Status"
                    WHEN 'pending'     THEN 0
                    WHEN 'in_progress' THEN 1
                    WHEN 'reopened'    THEN 2
                    WHEN 'submitted'   THEN 3
                END,
                u."FullName_Ar"
        ''', [campaign_id])
        cols = [d[0] for d in cur.description]
        submissions = [dict(zip(cols, r)) for r in cur.fetchall()]

    return Response({'submissions': submissions})
