"""
============================================================================
APPLY MANUAL REVIEW
============================================================================
Reads a color-coded Litrix Excel export where the supervisor has manually
marked each paper's affiliation status, then applies those decisions to
the database AffiliationVerified column.

WHY THIS EXISTS
---------------
Automated verifiers (OpenAlex, Crossref, PDF, Publisher HTML) get ~70% of
papers right but miss the rest — paywalls, JS-rendered pages, malformed
metadata. The supervisor's manual review is the source of truth for the
remaining cases.

This script bridges the gap: it lifts her color-coded decisions out of
the Excel and writes them to the DB atomically, with full audit trail.

COLOR MAP (configurable via constants below)
--------------------------------------------
    Light Blue  (FFCFE2F3) → AffiliationVerified = TRUE  (Al-Baha confirmed)
    Deeper Blue (FFC9DAF8) → AffiliationVerified = TRUE  (Al-Baha confirmed)
    Red         (FFF4CCCC) → AffiliationVerified = FALSE (NOT Al-Baha)
    Green       (FFD9EAD3) → AffiliationVerified = FALSE (NOT Al-Baha)

MATCHING STRATEGY
-----------------
For each row in the Excel:
  1. If DOI present  → match DB row by DOI (case-insensitive, trimmed)
  2. If DOI missing  → match by exact Title (fall back to title only when
                        DOI is empty)
  3. If still no match → report unmatched in dry-run, skip in apply mode

SAFETY
------
* Default mode = DRY-RUN. Nothing touches the DB until --apply.
* Each UPDATE is transactional. Crash mid-run never corrupts state.
* Skips papers that already have the same decision (idempotent).
* Preserves AI-verifier decisions in VerificationDetails for audit.

CLI USAGE
---------
    # Inspect what would happen, no DB writes:
    python apply_manual_review.py path/to/file.xlsx

    # Apply for real:
    python apply_manual_review.py path/to/file.xlsx --apply

    # Show diff only (verified rows the user OVERRODE a previous AI decision):
    python apply_manual_review.py path/to/file.xlsx --show-overrides
============================================================================
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from typing import Any, Optional

import psycopg2
from psycopg2.extras import RealDictCursor

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: 'openpyxl' library not installed. Run: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# CONFIG
# ============================================================================

# DB connection (loaded from Django settings same way as affiliation_verifier)
DB_KWARGS: dict[str, Any] = {}
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    DB_KWARGS = {'dsn': DATABASE_URL}
else:
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'litrix_backend.settings')
        import django
        django.setup()
        from django.conf import settings
        db = settings.DATABASES['default']
        DB_KWARGS = {
            'host':     db['HOST'],
            'port':     int(db.get('PORT') or 5432),
            'dbname':   db['NAME'],
            'user':     db['USER'],
            'password': db['PASSWORD'],
        }
        opts = db.get('OPTIONS', {}) or {}
        if opts.get('sslmode'):
            DB_KWARGS['sslmode'] = opts['sslmode']
    except Exception as e:
        print(f'ERROR: Could not load DB config: {e}')
        sys.exit(1)


# Color → Decision mapping
# Format: openpyxl returns the cell's fill.start_color.rgb as 8-char hex
# (the first 2 chars are alpha — typically 'FF').
COLOR_MAP = {
    'FFCFE2F3': True,    # Light Blue  → Al-Baha
    'FFC9DAF8': True,    # Deeper Blue → Al-Baha
    'FFF4CCCC': False,   # Red         → NOT Al-Baha
    'FFD9EAD3': False,   # Green       → NOT Al-Baha (user said forgot to recolor)
}

# Which sheets in the workbook to process
SHEETS_TO_PROCESS = ['Journals 2025', 'Conferences 2025',
                     'Journals 2026', 'Conferences 2026']


# ============================================================================
# EXCEL PARSING
# ============================================================================

def extract_decisions_from_excel(xlsx_path: str) -> list[dict]:
    """
    Walks every row in the configured sheets and returns a list of:
        {
            'sheet':   'Journals 2025',
            'row':     45,
            'doi':     '10.1109/...',  (lowercase, stripped)
            'title':   'Some paper title',
            'researcher': 'الباحث المنسوب',
            'color':   'FFCFE2F3',
            'decision': True | False | None,  # None = unrecognized color
        }
    """
    wb = load_workbook(xlsx_path)
    decisions: list[dict] = []

    for sheet_name in SHEETS_TO_PROCESS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Locate the relevant columns by header text
        headers = [c.value for c in ws[1]]
        try:
            title_idx     = headers.index('Title')
            doi_idx       = headers.index('DOI')
            researcher_idx = headers.index('Al-Baha Researchers')
        except ValueError as e:
            print(f"  ! {sheet_name}: missing required column ({e}), skipping")
            continue

        for row_num in range(2, ws.max_row + 1):
            # The fill color of the first cell represents the row's decision.
            # (Openpyxl can color one cell or all — we proxy on column A
            # because the user marks whole rows.)
            cell = ws.cell(row=row_num, column=1)
            color_rgb = None
            try:
                if cell.fill and cell.fill.start_color:
                    color_rgb = cell.fill.start_color.rgb
                    if not isinstance(color_rgb, str):
                        color_rgb = None
            except (AttributeError, TypeError):
                pass

            # Skip empty/header-style rows
            title = ws.cell(row=row_num, column=title_idx + 1).value
            if not title or str(title).strip() == '':
                continue

            doi = ws.cell(row=row_num, column=doi_idx + 1).value
            researcher = ws.cell(row=row_num, column=researcher_idx + 1).value

            decisions.append({
                'sheet':      sheet_name,
                'row':        row_num,
                'doi':        (str(doi).strip().lower() if doi else ''),
                'title':      str(title).strip(),
                'researcher': str(researcher or '').strip(),
                'color':      color_rgb,
                'decision':   COLOR_MAP.get(color_rgb),  # True/False/None
            })

    return decisions


# ============================================================================
# DB OPERATIONS
# ============================================================================

def build_paper_indexes(conn) -> tuple[dict, dict]:
    """
    Single bulk query: pulls every paper once and builds two in-memory dicts
    so the matching loop runs at O(1) per lookup instead of full-scanning
    the table 172 times.

    Returns: (doi_index, title_index)
        doi_index   = {lowercased_trimmed_doi   → paper dict}
        title_index = {lowercased_trimmed_title → paper dict}
    """
    print("  Building paper indexes (single bulk fetch)...")
    doi_index: dict[str, dict] = {}
    title_index: dict[str, dict] = {}
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute('''
            SELECT
                "PaperID", "Title", "DOI", "Source", "PubYear",
                "AffiliationVerified", "VerificationSource"
            FROM "ResearchPaper"
        ''')
        for row in cur:
            doi = (row['DOI'] or '').strip().lower()
            title = (row['Title'] or '').strip().lower()
            if doi:
                doi_index[doi] = dict(row)
            if title and len(title) >= 10:
                title_index[title] = dict(row)
    print(f"  Loaded {len(doi_index)} DOI keys, {len(title_index)} Title keys")
    return doi_index, title_index


def find_paper_by_doi(doi_index: dict, doi: str) -> Optional[dict]:
    """O(1) dict lookup."""
    if not doi:
        return None
    return doi_index.get(doi.strip().lower())


def find_paper_by_title(title_index: dict, title: str) -> Optional[dict]:
    """O(1) dict lookup."""
    if not title or len(title) < 10:
        return None
    return title_index.get(title.strip().lower())


def update_paper_decision(conn, paper_id: int, decision: bool, evidence: dict):
    """Records the user's decision and preserves any prior AI verification."""
    # Merge the prior VerificationDetails with the manual override metadata
    # so the audit trail is preserved.
    with conn.cursor() as cur:
        cur.execute(
            '''
            UPDATE "ResearchPaper"
            SET "AffiliationVerified"  = %s,
                "VerificationSource"   = 'manual-review',
                "VerifiedAt"           = NOW(),
                "VerificationDetails"  = COALESCE("VerificationDetails", '{}'::jsonb)
                                         || %s::jsonb
            WHERE "PaperID" = %s
            ''',
            [decision, json.dumps(evidence, ensure_ascii=False, default=str), paper_id],
        )


def update_paper_title(conn, paper_id: int, new_title: str, old_title: str) -> tuple[bool, str]:
    """
    Updates the Title column with conflict handling for the unique constraint.

    Returns: (success, message)
    """
    with conn.cursor() as cur:
        # Use SAVEPOINT so a uniqueness violation doesn't abort the whole transaction
        cur.execute('SAVEPOINT before_title_update')
        try:
            cur.execute(
                '''
                UPDATE "ResearchPaper"
                SET "Title" = %s,
                    "VerificationDetails" = COALESCE("VerificationDetails", '{}'::jsonb)
                                            || jsonb_build_object(
                                                'title_corrected', true,
                                                'previous_title', %s,
                                                'corrected_at', NOW()::text
                                            )
                WHERE "PaperID" = %s
                ''',
                [new_title, old_title, paper_id],
            )
            cur.execute('RELEASE SAVEPOINT before_title_update')
            return True, 'updated'
        except psycopg2.errors.UniqueViolation as e:
            cur.execute('ROLLBACK TO SAVEPOINT before_title_update')
            return False, f'unique constraint: new title already exists on another paper'
        except Exception as e:
            cur.execute('ROLLBACK TO SAVEPOINT before_title_update')
            return False, str(e)[:200]


def normalize_for_compare(text: str) -> str:
    """Normalize text for comparison: lowercase, collapse whitespace, strip."""
    import re
    if not text:
        return ''
    return re.sub(r'\s+', ' ', text.strip().lower())


# ============================================================================
# REPORT
# ============================================================================

def banner(text: str):
    print()
    print('=' * 80)
    print(f' {text} '.center(80))
    print('=' * 80)


def summarize(decisions: list[dict]):
    """Counts decisions by color."""
    from collections import Counter
    counter = Counter(d['decision'] for d in decisions)
    print(f"  Total rows:           {len(decisions)}")
    print(f"  Al-Baha (TRUE):       {counter[True]}")
    print(f"  NOT Al-Baha (FALSE):  {counter[False]}")
    print(f"  Unrecognized color:   {counter[None]}")

    if counter[None] > 0:
        print(f"\n  Color codes not in COLOR_MAP (review needed):")
        unrecognized = set(d['color'] for d in decisions if d['decision'] is None)
        for c in sorted(unrecognized, key=lambda x: str(x)):
            print(f"    {c!r}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description='Apply color-coded manual review to DB')
    parser.add_argument('xlsx_path', help='Path to the Litrix export .xlsx file')
    parser.add_argument('--apply', action='store_true', help='Actually write to DB (default = dry-run)')
    parser.add_argument('--show-overrides', action='store_true',
                        help='Highlight cases where user disagreed with the AI verifier')
    parser.add_argument('--fix-titles', action='store_true',
                        help='Also correct Title columns where Excel differs from DB')
    parser.add_argument('--titles-only', action='store_true',
                        help='Only fix titles, skip affiliation updates')
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx_path):
        print(f'ERROR: file not found: {args.xlsx_path}')
        sys.exit(1)

    banner('STEP 1 — Reading Excel')
    decisions = extract_decisions_from_excel(args.xlsx_path)
    summarize(decisions)

    if not decisions:
        print('Nothing to do.')
        return

    banner('STEP 2 — Matching to DB')

    conn = psycopg2.connect(connect_timeout=15, **DB_KWARGS)
    doi_index, title_index = build_paper_indexes(conn)

    stats = {
        'matched_doi':       0,
        'matched_title':     0,
        'unmatched':         0,
        'already_correct':   0,
        'will_change':       0,
        'overrides_ai':      0,
        'unrecognized':      0,
        'title_corrections': 0,
        'title_identical':   0,
    }

    actions: list[dict] = []        # affiliation updates
    title_actions: list[dict] = []  # title corrections

    for d in decisions:
        if d['decision'] is None:
            stats['unrecognized'] += 1
            continue

        paper = None
        match_type = None
        if d['doi']:
            paper = find_paper_by_doi(doi_index, d['doi'])
            if paper:
                match_type = 'doi'

        if not paper and d['title']:
            paper = find_paper_by_title(title_index, d['title'])
            if paper:
                match_type = 'title'

        if not paper:
            stats['unmatched'] += 1
            continue

        if match_type == 'doi':
            stats['matched_doi'] += 1
        else:
            stats['matched_title'] += 1

        # Title correction check (only when matched by DOI — title matching
        # would be circular here since we'd compare title to title).
        if args.fix_titles and match_type == 'doi' and d['title']:
            db_title = (paper['Title'] or '').strip()
            excel_title = d['title'].strip()
            if normalize_for_compare(db_title) != normalize_for_compare(excel_title):
                title_actions.append({
                    'paper_id':  paper['PaperID'],
                    'doi':       paper['DOI'],
                    'old_title': db_title,
                    'new_title': excel_title,
                    'researcher': d['researcher'],
                })
                stats['title_corrections'] += 1
            else:
                stats['title_identical'] += 1

        # If --titles-only, skip the affiliation update logic
        if args.titles_only:
            continue

        # Did the AI verifier already mark this differently?
        prior = paper['AffiliationVerified']
        if prior == d['decision']:
            stats['already_correct'] += 1
            continue

        # Need to update
        stats['will_change'] += 1
        if prior is not None and prior != d['decision']:
            stats['overrides_ai'] += 1

        actions.append({
            'paper_id':   paper['PaperID'],
            'title':      paper['Title'][:80],
            'doi':        paper['DOI'],
            'prior':      prior,
            'new':        d['decision'],
            'sheet':      d['sheet'],
            'row':        d['row'],
            'match_type': match_type,
            'color':      d['color'],
            'researcher': d['researcher'],
        })

    banner('STEP 3 — Match Summary')
    print(f"  Matched by DOI:       {stats['matched_doi']}")
    print(f"  Matched by Title:     {stats['matched_title']}")
    print(f"  Unmatched (skipped):  {stats['unmatched']}")
    if not args.titles_only:
        print(f"  Already correct:      {stats['already_correct']}")
        print(f"  Will be UPDATEd:      {stats['will_change']}")
        if stats['overrides_ai']:
            print(f"  ↳ AI overrides:       {stats['overrides_ai']} (you disagreed with verifier)")
    if args.fix_titles or args.titles_only:
        print(f"  Title identical:      {stats['title_identical']}")
        print(f"  Title corrections:    {stats['title_corrections']}")
    if stats['unrecognized']:
        print(f"  Unrecognized colors:  {stats['unrecognized']}")

    # Show preview of title diffs
    if (args.fix_titles or args.titles_only) and title_actions:
        banner('TITLE CORRECTIONS PREVIEW')
        for a in title_actions[:30]:  # cap so output isn't huge
            print(f"  #{a['paper_id']}  ({a['researcher'][:30]})")
            print(f"    OLD: {a['old_title'][:100]}")
            print(f"    NEW: {a['new_title'][:100]}")
            print()
        if len(title_actions) > 30:
            print(f"  ... and {len(title_actions) - 30} more")

    if args.show_overrides and stats['overrides_ai']:
        banner('AI OVERRIDES (you disagreed with the verifier)')
        for a in actions:
            if a['prior'] is not None and a['prior'] != a['new']:
                print(f"  #{a['paper_id']:6d}  AI={str(a['prior']):<5}  →  You={str(a['new']):<5}")
                print(f"           {a['title']}")
                print(f"           {a['researcher']}  |  {a['doi']}")
                print()

    total_to_change = stats['will_change'] + len(title_actions)
    if total_to_change == 0:
        banner('NOTHING TO APPLY')
        print("All decisions already match the DB — no UPDATEs needed.")
        conn.close()
        return

    if not args.apply:
        banner('DRY-RUN COMPLETE')
        if not args.titles_only:
            print(f"Would UPDATE {stats['will_change']} affiliation decisions.")
        if args.fix_titles or args.titles_only:
            print(f"Would CORRECT {len(title_actions)} titles.")
        print("Re-run with --apply to write changes to the database.")
        conn.close()
        return

    # ─── APPLY ────────────────────────────────────────────────────────
    # Affiliation decisions first
    if not args.titles_only and actions:
        banner('STEP 4a — APPLYING AFFILIATION DECISIONS')
        applied = 0
        failed  = 0
        for a in actions:
            evidence = {
                'manual_review': {
                    'sheet':      a['sheet'],
                    'row':        a['row'],
                    'color':      a['color'],
                    'match_type': a['match_type'],
                    'overrode_ai': a['prior'] is not None,
                    'prior_decision': str(a['prior']),
                }
            }
            try:
                update_paper_decision(conn, a['paper_id'], a['new'], evidence)
                conn.commit()
                applied += 1
                print(f"  ✓ #{a['paper_id']:6d}  →  {'Al-Baha' if a['new'] else 'NOT Al-Baha'}  ({a['title'][:50]}...)")
            except Exception as e:
                conn.rollback()
                failed += 1
                print(f"  ✗ #{a['paper_id']:6d}  FAILED: {e}")
        print(f"\n  Applied: {applied}  Failed: {failed}")

    # Title corrections second
    if (args.fix_titles or args.titles_only) and title_actions:
        banner('STEP 4b — APPLYING TITLE CORRECTIONS')
        applied = 0
        failed  = 0
        for a in title_actions:
            try:
                ok, msg = update_paper_title(conn, a['paper_id'], a['new_title'], a['old_title'])
                if ok:
                    conn.commit()
                    applied += 1
                    print(f"  ✓ #{a['paper_id']:6d}  title corrected")
                else:
                    failed += 1
                    print(f"  ✗ #{a['paper_id']:6d}  SKIPPED ({msg})")
            except Exception as e:
                conn.rollback()
                failed += 1
                print(f"  ✗ #{a['paper_id']:6d}  FAILED: {e}")
        print(f"\n  Applied: {applied}  Failed: {failed}")

    banner('ALL APPLY COMPLETE')

    conn.close()


if __name__ == '__main__':
    main()
# ORPHAN_BLOCK_TO_DELETE_BELOW
"""
r == d['decision']:
            stats['already_correct'] += 1
            continue

        # Need to update
        stats['will_change'] += 1
        if prior is not None and prior != d['decision']:
            stats['overrides_ai'] += 1

        actions.append({
            'paper_id':   paper['PaperID'],
            'title':      paper['Title'][:80],
            'doi':        paper['DOI'],
            'prior':      prior,
            'new':        d['decision'],
            'sheet':      d['sheet'],
            'row':        d['row'],
            'match_type': match_type,
            'color':      d['color'],
            'researcher': d['researcher'],
        })

    banner('STEP 3 — Match Summary')
    print(f"  Matched by DOI:       {stats['matched_doi']}")
    print(f"  Matched by Title:     {stats['matched_title']}")
    print(f"  Unmatched (skipped):  {stats['unmatched']}")
    if not args.titles_only:
        print(f"  Already correct:      {stats['already_correct']}")
        print(f"  Will be UPDATEd:      {stats['will_change']}")
        if stats['overrides_ai']:
            print(f"  ↳ AI overrides:       {stats['overrides_ai']} (you disagreed with verifier)")
    if args.fix_titles or args.titles_only:
        print(f"  Title identical:      {stats['title_identical']}")
        print(f"  Title corrections:    {stats['title_corrections']}")
    if stats['unrecognized']:
        print(f"  Unrecognized colors:  {stats['unrecognized']}")

    # Show preview of title diffs
    if (args.fix_titles or args.titles_only) and title_actions:
        banner('TITLE CORRECTIONS PREVIEW')
        for a in title_actions[:30]:
            print(f"  #{a['paper_id']}  ({a['researcher'][:30]})")
            print(f"    OLD: {a['old_title'][:100]}")
            print(f"    NEW: {a['new_title'][:100]}")
            print()
        if len(title_actions) > 30:
            print(f"  ... and {len(title_actions) - 30} more")

    if args.show_overrides and stats['overrides_ai']:
        banner('AI OVERRIDES (you disagreed with the verifier)')
        for a in actions:
            if a['prior'] is not None and a['prior'] != a['new']:
                print(f"  #{a['paper_id']:6d}  AI={str(a['prior']):<5}  →  You={str(a['new']):<5}")
                print(f"           {a['title']}")
                print(f"           {a['researcher']}  |  {a['doi']}")
                print()

    total_to_change = stats['will_change'] + len(title_actions)
    if total_to_change == 0:
        banner('NOTHING TO APPLY')
        print("All decisions already match the DB — no UPDATEs needed.")
        conn.close()
        return

    if not args.apply:
        banner('DRY-RUN COMPLETE')
        if not args.titles_only:
            print(f"Would UPDATE {stats['will_change']} affiliation decisions.")
        if args.fix_titles or args.titles_only:
            print(f"Would CORRECT {len(title_actions)} titles.")
        print("Re-run with --apply to write changes to the database.")
        conn.close()
        return

    # Apply affiliation decisions
    if not args.titles_only and actions:
        banner('STEP 4a — APPLYING AFFILIATION DECISIONS')
        applied = 0
        failed  = 0
        for a in actions:
            evidence = {
                'manual_review': {
                    'sheet':      a['sheet'],
                    'row':        a['row'],
                    'color':      a['color'],
                    'match_type': a['match_type'],
                    'overrode_ai': a['prior'] is not None,
                    'prior_decision': str(a['prior']),
                }
            }
            try:
                update_paper_decision(conn, a['paper_id'], a['new'], evidence)
                conn.commit()
                applied += 1
                print(f"  ✓ #{a['paper_id']:6d}  →  {'Al-Baha' if a['new'] else 'NOT Al-Baha'}  ({a['title'][:50]}...)")
            except Exception as e:
                conn.rollback()
                failed += 1
                print(f"  ✗ #{a['paper_id']:6d}  FAILED: {e}")
        print(f"\n  Applied: {applied}  Failed: {failed}")

    # Apply title corrections
    if (args.fix_titles or args.titles_only) and title_actions:
        banner('STEP 4b — APPLYING TITLE CORRECTIONS')
        applied = 0
        failed  = 0
        for a in title_actions:
            try:
                ok, msg = update_paper_title(conn, a['paper_id'], a['new_title'], a['old_title'])
                if ok:
                    conn.commit()
                    applied += 1
                    print(f"  ✓ #{a['paper_id']:6d}  title corrected")
                else:
                    failed += 1
                    print(f"  ✗ #{a['paper_id']:6d}  SKIPPED ({msg})")
            except Exception as e:
                conn.rollback()
                failed += 1
                print(f"  ✗ #{a['paper_id']:6d}  FAILED: {e}")
        print(f"\n  Applied: {applied}  Failed: {failed}")

    banner('ALL APPLY COMPLETE')
    conn.close()


if __name__ == '__main__':
    main()


"""
